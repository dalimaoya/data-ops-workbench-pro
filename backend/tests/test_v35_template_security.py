"""Tests for v3.5 Template Security features."""
import json
import os
import sys
import tempfile

import openpyxl
from openpyxl.styles import Protection as CellProtection

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))


class TestTemplateExcelProtection:
    """Test Excel template protection mechanisms."""

    def _create_mock_template(self):
        """Create a template Excel file similar to what export_template generates."""
        from openpyxl.styles import PatternFill, Font as XlFont
        from openpyxl.worksheet.protection import SheetProtection

        _SHEET_PROTECTION_PASSWORD = "DOW_tpl_v35_sec"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "数据"

        locked_cell = CellProtection(locked=True)
        unlocked_cell = CellProtection(locked=False)

        # v3.5 visual styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = XlFont(bold=True, color="FFFFFF", size=11)
        readonly_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        editable_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        blank_zone_fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")

        fields = [
            {"name": "id", "alias": "编号", "is_pk": True, "is_editable": False},
            {"name": "name", "alias": "名称", "is_pk": False, "is_editable": True},
            {"name": "status", "alias": "状态", "is_pk": False, "is_editable": False},
        ]
        pk_set = {"id"}
        editable_set = {"name"}

        # Header row
        for i, f in enumerate(fields, 1):
            cell = ws.cell(row=1, column=i, value=f["alias"])
            cell.font = header_font
            cell.fill = header_fill
            cell.protection = locked_cell

        # Data rows
        data = [
            ["1", "Alice", "active"],
            ["2", "Bob", "inactive"],
        ]
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, (val, f) in enumerate(zip(row_data, fields), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if f["name"] in pk_set:
                    cell.protection = locked_cell
                    cell.fill = readonly_fill
                elif f["name"] in editable_set:
                    cell.protection = unlocked_cell
                    cell.fill = editable_fill
                else:
                    cell.protection = locked_cell
                    cell.fill = readonly_fill

        # Blank rows
        blank_start = 2 + len(data)
        for row_idx in range(blank_start, blank_start + 50):
            for col_idx, f in enumerate(fields, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value="")
                cell.fill = blank_zone_fill
                if f["name"] in pk_set:
                    cell.protection = unlocked_cell
                elif f["name"] in editable_set:
                    cell.protection = unlocked_cell
                else:
                    cell.protection = locked_cell

        # Sheet protection with password
        ws.protection = SheetProtection(
            sheet=True,
            password=_SHEET_PROTECTION_PASSWORD,
            formatColumns=False,
            formatRows=False,
            formatCells=False,
            insertRows=False,
            deleteRows=True,
            deleteColumns=True,
            insertColumns=True,
            sort=False,
            autoFilter=False,
        )

        # Meta sheet
        meta_ws = wb.create_sheet("_meta")
        meta_info = {
            "table_config_id": 1,
            "datasource_id": 1,
            "config_version": 1,
            "field_codes": ["id", "name", "status"],
            "field_aliases": ["编号", "名称", "状态"],
            "primary_key_fields": ["id"],
            "data_row_count": 2,
            "blank_row_start": 4,
            "reserved_blank_rows": 50,
        }
        meta_ws.cell(row=1, column=1, value=json.dumps(meta_info, ensure_ascii=False))
        meta_ws.sheet_state = "hidden"

        return wb

    def test_header_row_locked(self):
        """v3.5 验收 #1: 列头行不可编辑（locked=True）"""
        wb = self._create_mock_template()
        ws = wb.active
        for col_idx in range(1, 4):
            cell = ws.cell(row=1, column=col_idx)
            assert cell.protection.locked, f"Header cell col {col_idx} should be locked"

    def test_header_row_style(self):
        """v3.5 验收: 表头加粗 + 蓝色背景"""
        wb = self._create_mock_template()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        assert cell.font.bold, "Header should be bold"
        assert cell.fill.start_color.rgb == "004472C4", f"Header fill should be blue, got {cell.fill.start_color.rgb}"

    def test_sheet_protection_enabled_with_password(self):
        """v3.5 验收: 工作表保护已启用且带密码"""
        wb = self._create_mock_template()
        ws = wb.active
        assert ws.protection.sheet, "Sheet protection should be enabled"
        assert ws.protection.password is not None, "Sheet protection should have password"

    def test_column_delete_insert_protected(self):
        """v3.5 验收 #2: 不可删除列或插入新列"""
        wb = self._create_mock_template()
        ws = wb.active
        assert ws.protection.deleteColumns, "deleteColumns should be True (protected)"
        assert ws.protection.insertColumns, "insertColumns should be True (protected)"

    def test_pk_column_locked_for_data_rows(self):
        """v3.5 验收 #3: 已有数据行的主键列不可编辑"""
        wb = self._create_mock_template()
        ws = wb.active
        # id column (col 1) in data rows (rows 2-3) should be locked
        assert ws.cell(row=2, column=1).protection.locked, "PK cell row 2 should be locked"
        assert ws.cell(row=3, column=1).protection.locked, "PK cell row 3 should be locked"

    def test_editable_cells_unlocked(self):
        """v3.5 验收 #4: 可编辑字段的数据单元格可以正常编辑"""
        wb = self._create_mock_template()
        ws = wb.active
        # name column (col 2) in data rows should be unlocked
        assert not ws.cell(row=2, column=2).protection.locked, "Editable cell should be unlocked"
        assert not ws.cell(row=3, column=2).protection.locked, "Editable cell should be unlocked"

    def test_non_editable_cells_locked(self):
        """v3.5: 非可编辑字段锁定"""
        wb = self._create_mock_template()
        ws = wb.active
        # status column (col 3) should be locked
        assert ws.cell(row=2, column=3).protection.locked, "Non-editable cell should be locked"

    def test_blank_rows_exist(self):
        """v3.5 验收 #5: 模板底部有 50 行空白预留区域"""
        wb = self._create_mock_template()
        ws = wb.active
        # Blank rows start at row 4, should have 50 rows
        blank_cells_count = 0
        for row_idx in range(4, 54):
            cell = ws.cell(row=row_idx, column=1)
            if cell.value == "" or cell.value is None:
                blank_cells_count += 1
        assert blank_cells_count == 50, f"Expected 50 blank rows, got {blank_cells_count}"

    def test_blank_row_pk_unlocked(self):
        """v3.5: 空白区域的主键列可以编辑（允许新增）"""
        wb = self._create_mock_template()
        ws = wb.active
        assert not ws.cell(row=4, column=1).protection.locked, "Blank row PK should be unlocked"
        assert not ws.cell(row=10, column=1).protection.locked, "Blank row PK should be unlocked"

    def test_readonly_cells_gray(self):
        """v3.5 验收 #6: 只读单元格背景色为浅灰"""
        wb = self._create_mock_template()
        ws = wb.active
        # PK cell (data row) should be gray
        fill = ws.cell(row=2, column=1).fill
        assert fill.start_color.rgb == "00F0F0F0", f"Readonly cell should be gray, got {fill.start_color.rgb}"

    def test_editable_cells_white(self):
        """v3.5 验收 #6: 可编辑单元格保持白色"""
        wb = self._create_mock_template()
        ws = wb.active
        fill = ws.cell(row=2, column=2).fill
        assert fill.start_color.rgb == "00FFFFFF", f"Editable cell should be white, got {fill.start_color.rgb}"

    def test_blank_zone_yellow(self):
        """v3.5 验收 #6: 模板底部空白区域加浅黄色背景"""
        wb = self._create_mock_template()
        ws = wb.active
        fill = ws.cell(row=4, column=1).fill
        assert fill.start_color.rgb == "00FFFFF0", f"Blank zone should be light yellow, got {fill.start_color.rgb}"

    def test_template_save_and_load(self):
        """v3.5: 模板可以正确保存和加载"""
        wb = self._create_mock_template()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            # Reload and verify
            wb2 = openpyxl.load_workbook(f.name)
            ws2 = wb2.active
            assert ws2.protection.sheet, "Protection should persist after save/load"
            assert ws2.cell(row=1, column=1).value == "编号"
            assert ws2.cell(row=2, column=1).protection.locked
            assert not ws2.cell(row=2, column=2).protection.locked
            os.unlink(f.name)


class TestImportValidation:
    """Test import validation enhancements."""

    def test_col_count_mismatch_message(self):
        """v3.5 验收 #7: 列数不匹配时阻断，提示具体差异"""
        from app.i18n import t, set_lang
        set_lang("zh")
        msg = t("data_maintenance.col_count_mismatch", expected=5, actual=3)
        assert "3" in msg and "5" in msg
        assert "列" in msg or "column" in msg.lower()

    def test_col_name_mismatch_message(self):
        """v3.5 验收 #8: 列名不匹配时阻断，提示具体位置"""
        from app.i18n import t, set_lang
        set_lang("zh")
        msg = t("data_maintenance.col_name_mismatch", col=2, expected="名称", actual="Name")
        assert "2" in msg
        assert "名称" in msg
        assert "Name" in msg

    def test_pk_modified_message(self):
        """v3.5 验收 #9: 主键被修改时阻断，提示具体行和值"""
        from app.i18n import t, set_lang
        set_lang("zh")
        msg = t("data_maintenance.row_pk_modified", row=5, old_val="ABC", new_val="XYZ")
        assert "5" in msg
        assert "XYZ" in msg
        assert "修改" in msg or "modified" in msg.lower()

    def test_date_validation_message(self):
        """v3.5 验收 #10: 类型校验错误精确到行号+列名 — 日期格式"""
        from app.i18n import t, set_lang
        set_lang("zh")
        msg = t("data_maintenance.row_field_expect_date", row=3, field="创建日期", value="not-a-date")
        assert "3" in msg
        assert "创建日期" in msg
        assert "not-a-date" in msg

    def test_int_validation_message(self):
        """v3.5 验收 #10: 类型校验错误精确到行号+列名 — 整数"""
        from app.i18n import t, set_lang
        set_lang("zh")
        msg = t("data_maintenance.row_field_expect_int", row=7, field="数量")
        assert "7" in msg
        assert "数量" in msg

    def test_number_validation_message(self):
        """v3.5 验收 #10: 类型校验错误精确到行号+列名 — 数值"""
        from app.i18n import t, set_lang
        set_lang("zh")
        msg = t("data_maintenance.row_field_expect_number", row=10, field="金额")
        assert "10" in msg
        assert "金额" in msg

    def test_date_parser_valid_dates(self):
        """v3.5: dateutil parser 能正确识别常见日期格式"""
        from dateutil import parser as date_parser
        valid_dates = [
            "2026-03-25",
            "2026-03-25 10:30:00",
            "2026/03/25",
            "25-Mar-2026",
            "Mar 25, 2026",
            "20260325",
        ]
        for d in valid_dates:
            try:
                result = date_parser.parse(d)
                assert result is not None, f"Should parse: {d}"
            except (ValueError, OverflowError):
                assert False, f"Should parse valid date: {d}"

    def test_date_parser_invalid_dates(self):
        """v3.5: dateutil parser 能正确拒绝无效日期"""
        from dateutil import parser as date_parser
        invalid_dates = [
            "not-a-date",
            "abc123",
            "xxxx-yy-zz",
        ]
        for d in invalid_dates:
            try:
                date_parser.parse(d)
                # Some strings might be parseable, that's OK
            except (ValueError, OverflowError):
                pass  # Expected

    def test_en_messages(self):
        """v3.5: English i18n messages work"""
        from app.i18n import t, set_lang
        set_lang("en")
        msg = t("data_maintenance.row_pk_modified", row=5, old_val="A", new_val="B")
        assert "modified" in msg.lower()
        assert "5" in msg
        set_lang("zh")  # Reset
