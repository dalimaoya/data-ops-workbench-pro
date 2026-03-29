"""Batch manage API — batch table onboarding + AI suggest + batch confirm + batch export."""

import io
import json
import os
import re
import time
import zipfile
import uuid
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel

from app.database import get_db
from app.models import (
    TableConfig, FieldConfig, DatasourceConfig, TemplateExportLog, _now_bjt,
)
from app.utils.crypto import decrypt_password
from app.utils.remote_db import list_columns, list_tables, fetch_sample_data, compute_structure_hash
from app.utils.audit import log_operation
from app.utils.auth import get_current_user, require_role
from app.utils.permissions import get_permitted_datasource_ids
from app.models import UserAccount
from app.ai.ai_engine import AIEngine
from app.ai.rules_engine import suggest_semantic_name, is_system_field, is_readonly_field
from app.i18n import t

router = APIRouter(prefix="/api/batch-manage", tags=["批量纳管"])

# ── Regex helpers (reused from ai_suggest) ──
_NUMERIC_TYPE_PATTERNS = re.compile(
    r"(int|integer|bigint|smallint|tinyint|mediumint|serial|float|double|decimal|numeric|real|number)",
    re.IGNORECASE,
)
_AUTO_INC_PATTERNS = re.compile(
    r"(auto_increment|serial|nextval|identity)", re.IGNORECASE,
)
_TIMESTAMP_PATTERNS = re.compile(
    r"(datetime|timestamp|date|time)", re.IGNORECASE,
)
_SYSTEM_KEYWORDS = {
    "created_at", "updated_at", "created_by", "updated_by", "is_deleted",
    "create_time", "update_time", "create_by", "update_by",
    "gmt_create", "gmt_modified",
}


def _is_numeric_type(db_data_type: str) -> bool:
    return bool(_NUMERIC_TYPE_PATTERNS.search(db_data_type))


def _is_auto_increment(db_data_type: str, column_default: Optional[str] = None) -> bool:
    if _AUTO_INC_PATTERNS.search(db_data_type):
        return True
    if column_default and _AUTO_INC_PATTERNS.search(column_default):
        return True
    return False


# ── Schemas ──

class BatchManageTablesRequest(BaseModel):
    datasource_id: int
    table_names: List[str]
    auto_ai_suggest: bool = True
    sample_count: int = 50


class FieldConfigItem(BaseModel):
    field_name: str
    field_alias: Optional[str] = None
    db_data_type: str
    field_order_no: int = 0
    is_primary_key: int = 0
    is_editable: int = 1
    is_required: int = 0
    is_system_field: int = 0
    is_displayed: int = 1
    include_in_export: int = 1
    include_in_import: int = 1
    enum_options_json: Optional[str] = None
    sample_value: Optional[str] = None
    editable_roles: Optional[str] = None
    remark: Optional[str] = None


class TableConfirmItem(BaseModel):
    table_name: str
    display_name: Optional[str] = None
    primary_key: str
    fields: List[FieldConfigItem]


class BatchConfirmRequest(BaseModel):
    datasource_id: int
    tables: List[TableConfirmItem]


class BatchExportRequest(BaseModel):
    datasource_id: int
    table_ids: List[int]
    format: str = "zip"  # "zip" or "multi_sheet"


# ── Helpers ──

def _get_ds(db: Session, ds_id: int) -> DatasourceConfig:
    ds = db.query(DatasourceConfig).filter(
        DatasourceConfig.id == ds_id, DatasourceConfig.is_deleted == 0
    ).first()
    if not ds:
        raise HTTPException(404, t("batch_manage.datasource_not_found"))
    return ds


def _gen_code(db: Session) -> str:
    today = _now_bjt().strftime("%Y%m%d")
    prefix = f"TB_{today}_"
    last = (
        db.query(TableConfig)
        .filter(TableConfig.table_config_code.like(f"{prefix}%"))
        .order_by(TableConfig.id.desc())
        .first()
    )
    seq = 1
    if last:
        try:
            seq = int(last.table_config_code.split("_")[-1]) + 1
        except ValueError:
            pass
    return f"{prefix}{seq:03d}"


def _analyze_column_values(col_idx: int, rows: list, is_numeric: bool):
    """Analyze sampled values for a single column."""
    values = []
    for row in rows:
        v = row[col_idx]
        if v is not None:
            values.append(v)

    result = {"total": len(rows), "non_null": len(values)}
    if not values:
        return result

    if is_numeric:
        nums = []
        for v in values:
            try:
                nums.append(float(v))
            except (ValueError, TypeError):
                pass
        if nums:
            result["min"] = min(nums)
            result["max"] = max(nums)
            result["avg"] = round(sum(nums) / len(nums), 2)
    else:
        str_values = [str(v) for v in values]
        unique = list(dict.fromkeys(str_values))
        result["unique_count"] = len(unique)
        if len(unique) <= 20:
            freq = {}
            for sv in str_values:
                freq[sv] = freq.get(sv, 0) + 1
            sorted_vals = sorted(unique, key=lambda x: freq[x], reverse=True)
            result["enum_candidates"] = sorted_vals

    return result


def _build_llm_prompt(uncovered_fields: list, sample_info: dict, table_name: str) -> str:
    """Build a prompt for LLM to enhance field semantic names."""
    prompt = f"""你是一个数据库字段语义分析助手。请根据表名「{table_name}」、字段名、数据类型和样本数据，为以下字段推荐中文语义名和表的中文别名。

要求：
1. 返回严格的 JSON 对象格式
2. 包含 table_display_name（表的中文别名）和 fields 数组
3. fields 中每个元素包含 field_name, display_name, reason
4. display_name 是简洁的中文名称（2-6个字）
5. reason 是推荐理由（一句话）

字段列表：
"""
    for f in uncovered_fields:
        info = sample_info.get(f["field_name"], {})
        sample_vals = ""
        if "enum_candidates" in info:
            sample_vals = f"，样本值: {info['enum_candidates'][:5]}"
        elif "min" in info:
            sample_vals = f"，范围: {info['min']}-{info['max']}"
        prompt += f"- {f['field_name']} (类型: {f['db_data_type']}{sample_vals})\n"

    prompt += "\n请返回 JSON 对象 {\"table_display_name\": \"...\", \"fields\": [...]}，不要输出其他内容："
    return prompt


async def _process_single_table(
    ds: DatasourceConfig,
    table_name: str,
    sample_count: int,
    db: Session,
    auto_ai: bool,
    managed_tables: set,
):
    """Process a single table: fetch columns, sample data, AI suggest."""
    pwd = decrypt_password(ds.password_encrypted)
    use_db = ds.database_name
    use_schema = ds.schema_name

    is_managed = table_name in managed_tables

    # Fetch columns
    try:
        columns = list_columns(
            db_type=ds.db_type, host=ds.host, port=ds.port,
            user=ds.username, password=pwd,
            table_name=table_name,
            database=use_db, schema=use_schema,
            charset=ds.charset or "utf8",
            timeout=ds.connect_timeout_seconds or 10,
        )
    except Exception as e:
        return {
            "table_name": table_name,
            "status": "error",
            "error": t("batch_manage.fetch_fields_failed", error=str(e)[:200]),
            "is_managed": is_managed,
            "fields": [],
            "ai_suggestions": {},
        }

    # Detect primary key
    pk_fields = []
    for col in columns:
        if col.get("is_primary_key"):
            pk_fields.append(col["field_name"])
    if not pk_fields and columns:
        # Heuristic: first column named 'id' or first column
        for col in columns:
            if col["field_name"].lower() == "id":
                pk_fields = [col["field_name"]]
                break
        if not pk_fields:
            pk_fields = [columns[0]["field_name"]]

    # Sample data
    sample_rows = []
    col_names = []
    try:
        col_names, sample_rows = fetch_sample_data(
            db_type=ds.db_type, host=ds.host, port=ds.port,
            user=ds.username, password=pwd,
            table_name=table_name,
            database=use_db, schema=use_schema,
            charset=ds.charset or "utf8",
            timeout=ds.connect_timeout_seconds or 10,
            limit=sample_count,
        )
    except Exception:
        pass

    col_index = {c: i for i, c in enumerate(col_names)} if col_names else {}

    # Build field info with AI suggestions
    fields = []
    ai_suggestions = {}
    uncovered_fields = []

    for col in columns:
        fname = col["field_name"]
        ftype = col["db_data_type"]
        is_pk = fname in pk_fields or col.get("is_primary_key", False)
        is_sys = fname.lower() in _SYSTEM_KEYWORDS
        is_numeric = _is_numeric_type(ftype)
        auto_inc = _is_auto_increment(ftype)
        is_ro = is_readonly_field(fname) if auto_ai else False

        # Sample value
        sample_val = None
        if col_index.get(fname) is not None and sample_rows:
            idx = col_index[fname]
            for sr in sample_rows:
                if sr[idx] is not None:
                    sample_val = str(sr[idx])
                    break

        field_info = {
            "field_name": fname,
            "db_data_type": ftype,
            "field_order_no": col.get("ordinal_position", 0),
            "is_primary_key": 1 if is_pk else 0,
            "is_editable": 0 if (is_pk or is_sys or auto_inc) else 1,
            "is_required": 1 if is_pk else 0,
            "is_system_field": 1 if is_sys else 0,
            "is_displayed": 1,
            "include_in_export": 0 if is_sys else 1,
            "include_in_import": 0 if (is_pk or is_sys) else 1,
            "sample_value": sample_val,
        }
        fields.append(field_info)

        if auto_ai:
            recs = []
            # Semantic name
            semantic = suggest_semantic_name(fname)
            if semantic:
                recs.append({
                    "property": "display_name",
                    "value": semantic,
                    "reason": f"字段名'{fname}'的常用中文释义",
                    "confidence": 0.95,
                })
            else:
                uncovered_fields.append({"field_name": fname, "db_data_type": ftype})

            # Readonly
            if is_pk or auto_inc or is_ro:
                reason_parts = []
                if is_pk:
                    reason_parts.append("主键字段")
                if auto_inc:
                    reason_parts.append("自增字段")
                if is_ro and not is_pk:
                    reason_parts.append("系统时间/审计字段")
                recs.append({
                    "property": "is_readonly",
                    "value": True,
                    "reason": "、".join(reason_parts) + "不应由用户编辑",
                    "confidence": 0.98 if is_pk else 0.92,
                })

            # System field
            if is_system_field(fname):
                recs.append({
                    "property": "is_system_field",
                    "value": True,
                    "reason": f"字段名'{fname}'匹配系统字段模式",
                    "confidence": 0.95,
                })

            # Enum values
            if not is_numeric and col_index.get(fname) is not None:
                analysis = _analyze_column_values(col_index[fname], sample_rows, False)
                if "enum_candidates" in analysis:
                    candidates = analysis["enum_candidates"]
                    unique_count = analysis.get("unique_count", len(candidates))
                    recs.append({
                        "property": "enum_values",
                        "value": candidates,
                        "reason": f"采样{len(sample_rows)}条数据中发现{unique_count}个唯一值",
                        "confidence": 0.85 if unique_count <= 10 else 0.70,
                    })

            if recs:
                ai_suggestions[fname] = recs

    # Try LLM for table name and uncovered fields
    table_display_name = None
    engine_used = "builtin_rules"

    if auto_ai:
        engine = AIEngine(db)
        if engine.engine_mode == "cloud" and engine.is_feature_enabled("field_suggest"):
            client = engine.get_llm_client()
            if client:
                try:
                    sample_info = {}
                    for f in uncovered_fields:
                        idx = col_index.get(f["field_name"])
                        if idx is not None:
                            is_num = _is_numeric_type(f["db_data_type"])
                            sample_info[f["field_name"]] = _analyze_column_values(idx, sample_rows, is_num)

                    # Include all fields for table name inference
                    all_fields_for_prompt = uncovered_fields if uncovered_fields else [
                        {"field_name": col["field_name"], "db_data_type": col["db_data_type"]}
                        for col in columns[:10]
                    ]
                    prompt = _build_llm_prompt(all_fields_for_prompt, sample_info, table_name)
                    resp = await client.chat([
                        {"role": "system", "content": "你是数据库字段语义分析助手。只返回JSON，不要其他文字。"},
                        {"role": "user", "content": prompt},
                    ])
                    content = resp.get("content", "").strip()

                    # Parse response
                    json_match = re.search(r'\{.*\}', content, re.DOTALL)
                    if json_match:
                        llm_result = json.loads(json_match.group())
                        table_display_name = llm_result.get("table_display_name")
                        llm_fields = llm_result.get("fields", [])
                        for item in llm_fields:
                            fname = item.get("field_name", "")
                            display_name = item.get("display_name", "")
                            reason = item.get("reason", "AI语义推断")
                            if not fname or not display_name:
                                continue
                            rec = {
                                "property": "display_name",
                                "value": display_name,
                                "reason": f"AI推荐：{reason}",
                                "confidence": 0.80,
                            }
                            if fname in ai_suggestions:
                                has_dn = any(r["property"] == "display_name" for r in ai_suggestions[fname])
                                if not has_dn:
                                    ai_suggestions[fname].insert(0, rec)
                            else:
                                ai_suggestions[fname] = [rec]
                        engine_used = "builtin_rules+llm"
                except Exception:
                    pass

    # If no LLM table name, try heuristic from table_name
    if not table_display_name:
        # Simple heuristic: use table name as-is
        table_display_name = table_name

    return {
        "table_name": table_name,
        "status": "success",
        "is_managed": is_managed,
        "table_display_name": table_display_name,
        "primary_key": ",".join(pk_fields),
        "fields": fields,
        "ai_suggestions": ai_suggestions,
        "engine": engine_used,
        "field_count": len(fields),
        "sample_count": len(sample_rows),
    }


# ── Endpoints ──

@router.post("/tables")
async def batch_manage_tables(
    body: BatchManageTablesRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin")),
):
    """Batch table onboarding with AI field suggestions."""
    start = time.time()
    ds = _get_ds(db, body.datasource_id)

    # Find already-managed tables for this datasource
    managed = db.query(TableConfig.table_name).filter(
        TableConfig.datasource_id == body.datasource_id,
        TableConfig.is_deleted == 0,
    ).all()
    managed_tables = {row[0] for row in managed}

    # Filter out already-managed tables
    tables_to_process = [tn for tn in body.table_names if tn not in managed_tables]
    if not tables_to_process:
        raise HTTPException(400, t("batch_manage.all_managed"))

    results = []
    for table_name in tables_to_process:
        result = await _process_single_table(
            ds, table_name, body.sample_count, db, body.auto_ai_suggest, managed_tables,
        )
        results.append(result)

    elapsed = int((time.time() - start) * 1000)
    return {
        "success": True,
        "data": {
            "total": len(tables_to_process),
            "results": results,
            "elapsed_ms": elapsed,
        },
    }


@router.post("/confirm")
def batch_confirm(
    body: BatchConfirmRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin")),
):
    """Batch save table configs and field configs."""
    ds = _get_ds(db, body.datasource_id)
    pwd = decrypt_password(ds.password_encrypted)
    use_db = ds.database_name
    use_schema = ds.schema_name

    created_tables = []
    errors = []

    for table_item in body.tables:
        try:
            # Check duplicate
            existing = db.query(TableConfig).filter(
                TableConfig.datasource_id == body.datasource_id,
                TableConfig.db_name == use_db,
                TableConfig.table_name == table_item.table_name,
                TableConfig.is_deleted == 0,
            ).first()
            if existing:
                errors.append({"table_name": table_item.table_name, "error": t("batch_manage.already_managed")})
                continue

            # Compute structure hash
            try:
                columns = list_columns(
                    db_type=ds.db_type, host=ds.host, port=ds.port,
                    user=ds.username, password=pwd,
                    table_name=table_item.table_name,
                    database=use_db, schema=use_schema,
                    charset=ds.charset or "utf8",
                    timeout=ds.connect_timeout_seconds or 10,
                )
                structure_hash = compute_structure_hash(columns)
            except Exception:
                structure_hash = None

            # Create table config
            tc = TableConfig(
                table_config_code=_gen_code(db),
                datasource_id=body.datasource_id,
                db_name=use_db,
                schema_name=use_schema,
                table_name=table_item.table_name,
                table_alias=table_item.display_name or table_item.table_name,
                config_version=1,
                structure_version_hash=structure_hash,
                primary_key_fields=table_item.primary_key,
                status="enabled",
                structure_check_status="normal",
                last_structure_check_at=_now_bjt(),
                last_sync_at=_now_bjt(),
                created_by=user.username,
                updated_by=user.username,
            )
            db.add(tc)
            db.flush()  # get tc.id

            # Create field configs
            pk_set = set(f.strip() for f in table_item.primary_key.split(","))
            for fi in table_item.fields:
                fc = FieldConfig(
                    table_config_id=tc.id,
                    field_name=fi.field_name,
                    field_alias=fi.field_alias or fi.field_name,
                    db_data_type=fi.db_data_type,
                    field_order_no=fi.field_order_no,
                    sample_value=fi.sample_value,
                    is_displayed=fi.is_displayed,
                    is_editable=fi.is_editable,
                    is_required=fi.is_required,
                    is_primary_key=fi.is_primary_key,
                    is_unique_key=0,
                    is_system_field=fi.is_system_field,
                    include_in_export=fi.include_in_export,
                    include_in_import=fi.include_in_import,
                    enum_options_json=fi.enum_options_json,
                    editable_roles=fi.editable_roles,
                    remark=fi.remark,
                    created_by=user.username,
                    updated_by=user.username,
                )
                db.add(fc)

            log_operation(
                db, "纳管表配置", "批量纳管", "success",
                target_id=tc.id, target_code=tc.table_config_code,
                target_name=tc.table_name,
                message=f"批量纳管表 {tc.table_alias or tc.table_name}",
                operator=user.username,
            )

            created_tables.append({
                "table_name": table_item.table_name,
                "table_config_id": tc.id,
                "display_name": tc.table_alias,
                "field_count": len(table_item.fields),
            })
        except Exception as e:
            errors.append({"table_name": table_item.table_name, "error": str(e)[:200]})

    db.commit()

    return {
        "success": True,
        "data": {
            "created": len(created_tables),
            "failed": len(errors),
            "tables": created_tables,
            "errors": errors,
        },
    }


# ── Batch Export ──

@router.post("/export")
def batch_export(
    body: BatchExportRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin", "operator")),
):
    """Batch export multiple managed tables as ZIP or multi-sheet xlsx."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, t("batch_manage.openpyxl_missing"))

    ds = _get_ds(db, body.datasource_id)
    pwd = decrypt_password(ds.password_encrypted)

    table_configs = []
    for tid in body.table_ids:
        tc = db.query(TableConfig).filter(
            TableConfig.id == tid, TableConfig.is_deleted == 0,
            TableConfig.datasource_id == body.datasource_id,
        ).first()
        if tc:
            table_configs.append(tc)

    if not table_configs:
        raise HTTPException(400, t("batch_manage.no_tables"))

    if body.format == "multi_sheet":
        # Single xlsx with multiple sheets
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for tc in table_configs:
            fields = (
                db.query(FieldConfig)
                .filter(
                    FieldConfig.table_config_id == tc.id,
                    FieldConfig.is_deleted == 0,
                    FieldConfig.include_in_export == 1,
                )
                .order_by(FieldConfig.field_order_no)
                .all()
            )
            sheet_name = (tc.table_alias or tc.table_name)[:31]  # Excel sheet name limit
            ws = wb.create_sheet(title=sheet_name)

            # Header row
            for ci, f in enumerate(fields, 1):
                ws.cell(row=1, column=ci, value=f.field_alias or f.field_name)

            # Data rows
            try:
                col_names, rows = fetch_sample_data(
                    db_type=ds.db_type, host=ds.host, port=ds.port,
                    user=ds.username, password=pwd,
                    table_name=tc.table_name,
                    database=tc.db_name or ds.database_name,
                    schema=tc.schema_name or ds.schema_name,
                    charset=ds.charset or "utf8",
                    timeout=ds.connect_timeout_seconds or 10,
                    limit=50000,
                )
                field_names = [f.field_name for f in fields]
                col_idx_map = {c: i for i, c in enumerate(col_names)}

                for ri, row_data in enumerate(rows, 2):
                    for ci, fname in enumerate(field_names, 1):
                        idx = col_idx_map.get(fname)
                        if idx is not None:
                            val = row_data[idx]
                            ws.cell(row=ri, column=ci, value=str(val) if val is not None else "")
            except Exception:
                ws.cell(row=2, column=1, value="数据获取失败")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"batch_export_{_now_bjt().strftime('%Y%m%d_%H%M%S')}.xlsx"
        log_operation(
            db, "数据维护", "批量导出", "success",
            message=f"批量导出 {len(table_configs)} 张表（多Sheet）",
            operator=user.username,
        )
        db.commit()

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    else:
        # ZIP format: each table as a separate xlsx
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for tc in table_configs:
                fields = (
                    db.query(FieldConfig)
                    .filter(
                        FieldConfig.table_config_id == tc.id,
                        FieldConfig.is_deleted == 0,
                        FieldConfig.include_in_export == 1,
                    )
                    .order_by(FieldConfig.field_order_no)
                    .all()
                )

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = (tc.table_alias or tc.table_name)[:31]

                # Header
                for ci, f in enumerate(fields, 1):
                    ws.cell(row=1, column=ci, value=f.field_alias or f.field_name)

                # Data
                try:
                    col_names, rows = fetch_sample_data(
                        db_type=ds.db_type, host=ds.host, port=ds.port,
                        user=ds.username, password=pwd,
                        table_name=tc.table_name,
                        database=tc.db_name or ds.database_name,
                        schema=tc.schema_name or ds.schema_name,
                        charset=ds.charset or "utf8",
                        timeout=ds.connect_timeout_seconds or 10,
                        limit=50000,
                    )
                    field_names = [f.field_name for f in fields]
                    col_idx_map = {c: i for i, c in enumerate(col_names)}

                    for ri, row_data in enumerate(rows, 2):
                        for ci, fname in enumerate(field_names, 1):
                            idx = col_idx_map.get(fname)
                            if idx is not None:
                                val = row_data[idx]
                                ws.cell(row=ri, column=ci, value=str(val) if val is not None else "")
                except Exception:
                    ws.cell(row=2, column=1, value="数据获取失败")

                xlsx_buf = io.BytesIO()
                wb.save(xlsx_buf)
                xlsx_buf.seek(0)

                safe_name = (tc.table_alias or tc.table_name).replace("/", "_").replace("\\", "_")
                zf.writestr(f"{safe_name}.xlsx", xlsx_buf.getvalue())

        zip_buf.seek(0)

        filename = f"batch_export_{_now_bjt().strftime('%Y%m%d_%H%M%S')}.zip"
        log_operation(
            db, "数据维护", "批量导出", "success",
            message=f"批量导出 {len(table_configs)} 张表（ZIP）",
            operator=user.username,
        )
        db.commit()

        return StreamingResponse(
            zip_buf,
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
