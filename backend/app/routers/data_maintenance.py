"""Data maintenance endpoints: browse, export, import, diff, writeback, delete rows, async export, batch export."""

from __future__ import annotations
import json
import os
import uuid
import tempfile
import threading
import zipfile
from datetime import datetime
from typing import Dict, List, Optional, Any

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db, DATA_DIR, SessionLocal
from app.models import (
    TableConfig, DatasourceConfig, FieldConfig,
    TemplateExportLog, ImportTaskLog, WritebackLog, TableBackupVersion,
    FieldChangeLog, ExportTask, _now_bjt,
)
from app.utils.crypto import decrypt_password
from app.utils.remote_db import _connect, fetch_sample_data, compute_structure_hash, list_columns
from app.utils.audit import log_operation
from app.utils.auth import get_current_user, require_role
from app.utils.permissions import get_permitted_datasource_ids
from app.utils.sql_security import sanitize_search_input, check_sql_injection
from app.utils.security_middleware import validate_upload_file, sanitize_dict
from app.models import UserAccount, ApprovalRequest, SystemSetting
from app.i18n import t

router = APIRouter(prefix="/api/data-maintenance", tags=["数据维护"])

def _needs_approval(db: Session, user) -> bool:
    """Check if approval workflow is enabled and user is not admin."""
    if not user or not hasattr(user, 'role') or user.role == "admin":
        return False
    row = db.query(SystemSetting).filter(
        SystemSetting.setting_key == "approval_enabled"
    ).first()
    return row is not None and row.setting_value == "true"


def _create_approval_request(
    db: Session, user, table_config_id: int, request_type: str,
    import_task_id: int = None, request_data_json: str = None,
):
    """Create an approval request and return the response."""
    from app.models import TableConfig, _now_bjt
    tc = db.query(TableConfig).filter(TableConfig.id == table_config_id).first()
    approval = ApprovalRequest(
        import_task_id=import_task_id,
        table_config_id=table_config_id,
        request_type=request_type,
        request_data_json=request_data_json,
        requested_by=_get_username(user),
        request_time=_now_bjt(),
        status="pending",
        structure_hash_at_request=tc.structure_version_hash if tc else None,
    )
    db.add(approval)
    db.flush()
    from app.utils.audit import log_operation
    log_operation(
        db, "审批流", "提交审批", "success",
        target_id=approval.id,
        target_name=tc.table_name if tc else None,
        message="用户 %s 提交 %s 审批" % (_get_username(user), request_type),
        operator=_get_username(user),
    )
    db.commit()
    return {
        "approval_required": True,
        "approval_id": approval.id,
        "status": "pending",
        "message": t("data_maintenance.submitted_for_approval"),
    }


def _get_username(user) -> str:
    """Get username from JWT user or fallback."""
    if user and hasattr(user, 'username'):
        return user.username
    return "system"


UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
EXPORT_DIR = os.path.join(DATA_DIR, "exports")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(EXPORT_DIR, exist_ok=True)


def _gen_batch(prefix: str) -> str:
    ts = _now_bjt().strftime("%Y%m%d%H%M%S")
    rand = uuid.uuid4().hex[:4].upper()
    return f"{prefix}_{ts}_{rand}"


def _get_tc(db: Session, tc_id: int) -> TableConfig:
    tc = db.query(TableConfig).filter(
        TableConfig.id == tc_id, TableConfig.is_deleted == 0, TableConfig.status == "enabled"
    ).first()
    if not tc:
        raise HTTPException(404, t("data_maintenance.table_not_found"))
    return tc


def _get_ds(db: Session, ds_id: int) -> DatasourceConfig:
    ds = db.query(DatasourceConfig).filter(
        DatasourceConfig.id == ds_id, DatasourceConfig.is_deleted == 0
    ).first()
    if not ds:
        raise HTTPException(404, t("data_maintenance.datasource_not_found"))
    return ds


def _get_fields(db: Session, tc_id: int) -> List[FieldConfig]:
    return (
        db.query(FieldConfig)
        .filter(FieldConfig.table_config_id == tc_id, FieldConfig.is_deleted == 0)
        .order_by(FieldConfig.field_order_no)
        .all()
    )


def _qualified_table(db_type: str, table_name: str, schema: Optional[str]) -> str:
    if db_type in ("postgresql", "kingbase"):
        sch = schema or "public"
        return f'"{sch}"."{table_name}"'
    elif db_type == "sqlserver":
        sch = schema or "dbo"
        return f"[{sch}].[{table_name}]"
    elif db_type == "oracle":
        return f'"{table_name.upper()}"'
    elif db_type == "dm":
        return f'"{table_name.upper()}"'
    return f"`{table_name}`"


def _quote_col(db_type: str, col: str) -> str:
    """Quote a column name for the given DB type."""
    if db_type == "sqlserver":
        return f"[{col}]"
    elif db_type in ("mysql", "sqlite"):
        return f"`{col}`"
    elif db_type in ("oracle", "dm"):
        return f'"{col.upper()}"'
    return f'"{col}"'


def _placeholder(db_type: str) -> str:
    """Always return %s. Use _exec() to execute SQL which converts placeholders for each DB."""
    return "%s"


def _exec(cur, db_type: str, sql: str, params: list) -> None:
    """Execute SQL with params, converting placeholder style for Oracle.
    Assumes SQL uses %s placeholders (like MySQL/PG). Converts:
    - Oracle: %s -> :1, :2, ...
    - SQL Server/DM: %s -> ?
    """
    if db_type in ("oracle",):
        idx = [0]
        def _repl(m):
            idx[0] += 1
            return f":{idx[0]}"
        import re
        converted = re.sub(r'%s', lambda m: _repl(m), sql)
        cur.execute(converted, params)
    elif db_type in ("sqlserver", "dm", "sqlite"):
        converted = sql.replace("%s", "?")
        cur.execute(converted, params)
    else:
        cur.execute(sql, params)


def _cast_to_text(db_type: str, col_expr: str) -> str:
    """Wrap a column expression with CAST to text for LIKE comparisons."""
    if db_type == "sqlserver":
        return f"CAST({col_expr} AS NVARCHAR(MAX))"
    elif db_type in ("postgresql", "kingbase"):
        return f"CAST({col_expr} AS TEXT)"
    elif db_type in ("oracle", "dm"):
        return f"TO_CHAR({col_expr})"
    elif db_type == "sqlite":
        return f"CAST({col_expr} AS TEXT)"
    return f"CAST({col_expr} AS CHAR)"


def _create_backup_table(cur, db_type: str, source_qt: str, backup_table_name: str, schema: Optional[str]) -> None:
    """Create a backup table as a copy of the source table. DB-dialect aware."""
    if db_type == "mysql":
        cur.execute(f"CREATE TABLE `{backup_table_name}` AS SELECT * FROM {source_qt}")
    elif db_type in ("postgresql", "kingbase"):
        sch = schema or "public"
        cur.execute(f'CREATE TABLE "{sch}"."{backup_table_name}" AS SELECT * FROM {source_qt}')
    elif db_type == "sqlserver":
        sch = schema or "dbo"
        cur.execute(f"SELECT * INTO [{sch}].[{backup_table_name}] FROM {source_qt}")
    elif db_type == "oracle":
        cur.execute(f'CREATE TABLE "{backup_table_name.upper()}" AS SELECT * FROM {source_qt}')
    elif db_type == "dm":
        cur.execute(f'CREATE TABLE "{backup_table_name.upper()}" AS SELECT * FROM {source_qt}')
    elif db_type == "sqlite":
        cur.execute(f"CREATE TABLE `{backup_table_name}` AS SELECT * FROM {source_qt}")


def _drop_table_if_exists(cur, db_type: str, table_qt: str) -> None:
    """Drop a table if it exists. DB-dialect aware."""
    if db_type == "sqlserver":
        cur.execute(f"IF OBJECT_ID('{table_qt}') IS NOT NULL DROP TABLE {table_qt}")
    elif db_type in ("oracle", "dm"):
        # Oracle/DM: use PL/SQL block or just try DROP
        try:
            cur.execute(f"DROP TABLE {table_qt}")
        except Exception:
            pass  # table doesn't exist
    else:
        cur.execute(f"DROP TABLE IF EXISTS {table_qt}")


# ─────────────────────────────────────────────
# P2-1: 数据浏览
# ─────────────────────────────────────────────

@router.get("/tables")
def list_maintenance_tables(
    keyword: Optional[str] = None,
    datasource_id: Optional[int] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    user: UserAccount = Depends(get_current_user),
):
    """可维护表列表，支持按数据源和关键词筛选。"""
    q = db.query(TableConfig).filter(
        TableConfig.is_deleted == 0, TableConfig.status == "enabled"
    )
    # v2.2: datasource-level permission filtering
    permitted_ids = get_permitted_datasource_ids(db, user)
    if permitted_ids is not None:
        if not permitted_ids:
            return {"total": 0, "items": []}
        q = q.filter(TableConfig.datasource_id.in_(permitted_ids))
    if datasource_id:
        q = q.filter(TableConfig.datasource_id == datasource_id)
    if keyword:
        q = q.filter(
            (TableConfig.table_name.contains(keyword)) |
            (TableConfig.table_alias.contains(keyword))
        )
    total = q.count()
    rows = q.order_by(TableConfig.id.desc()).offset((page - 1) * page_size).limit(page_size).all()

    result = []
    for row in rows:
        ds = db.query(DatasourceConfig).filter(DatasourceConfig.id == row.datasource_id).first()
        field_count = db.query(FieldConfig).filter(
            FieldConfig.table_config_id == row.id, FieldConfig.is_deleted == 0
        ).count()
        result.append({
            "id": row.id,
            "table_config_code": row.table_config_code,
            "datasource_id": row.datasource_id,
            "datasource_name": ds.datasource_name if ds else None,
            "db_type": ds.db_type if ds else None,
            "db_name": row.db_name,
            "schema_name": row.schema_name,
            "table_name": row.table_name,
            "table_alias": row.table_alias,
            "config_version": row.config_version,
            "structure_check_status": row.structure_check_status,
            "field_count": field_count,
            "allow_insert_rows": row.allow_insert_rows if user.role in ("admin", "operator") else 0,
            "allow_delete_rows": row.allow_delete_rows if user.role in ("admin", "operator") else 0,
            "updated_by": row.updated_by,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
        })
    return {"total": total, "items": result}


@router.get("/{table_config_id}/data")
def browse_table_data(
    table_config_id: int,
    keyword: Optional[str] = None,
    field_filters: Optional[str] = None,  # JSON: {"field_name": "value", ...}
    structured_filters: Optional[str] = None,  # JSON: [{"field":"x","operator":"eq","value":"y"}, ...]
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    user: UserAccount = Depends(get_current_user),
):
    """分页读取业务表数据，所有字段值按文本返回。"""
    tc = _get_tc(db, table_config_id)
    ds = _get_ds(db, tc.datasource_id)
    pwd = decrypt_password(ds.password_encrypted)
    fields = _get_fields(db, table_config_id)
    display_fields = [f for f in fields if f.is_displayed]
    if not display_fields:
        return {"columns": [], "rows": [], "total": 0, "page": page, "page_size": page_size}

    col_names = [f.field_name for f in display_fields]

    conn = _connect(
        ds.db_type, ds.host, ds.port, ds.username, pwd,
        tc.db_name, tc.schema_name, ds.charset, ds.connect_timeout_seconds or 10,
    )
    try:
        cur = conn.cursor()
        qt = _qualified_table(ds.db_type, tc.table_name, tc.schema_name)
        ph = _placeholder(ds.db_type)

        # Build WHERE clause
        where_parts: List[str] = []
        params: list = []

        if keyword:
            # SQL injection check on search keyword
            if check_sql_injection(keyword):
                raise HTTPException(400, "搜索关键字包含不安全字符")
            kw_parts = []
            for f in display_fields:
                kw_parts.append(f"{_cast_to_text(ds.db_type, _quote_col(ds.db_type, f.field_name))} LIKE {ph}")
                params.append(f"%{keyword}%")
            if kw_parts:
                where_parts.append(f"({' OR '.join(kw_parts)})")

        if field_filters:
            try:
                ff = json.loads(field_filters)
                for fname, fval in ff.items():
                    if fval and any(f.field_name == fname for f in display_fields):
                        # SQL injection check on filter values
                        if check_sql_injection(str(fval)):
                            raise HTTPException(400, f"筛选值包含不安全字符: {fname}")
                        where_parts.append(f"{_cast_to_text(ds.db_type, _quote_col(ds.db_type, fname))} LIKE {ph}")
                        params.append(f"%{fval}%")
            except (json.JSONDecodeError, TypeError):
                pass

        # v3.0: Structured filters from AI NL Query
        if structured_filters:
            try:
                sf_list = json.loads(structured_filters)
                valid_field_names = {f.field_name for f in display_fields}
                _OP_MAP = {"eq", "neq", "gt", "gte", "lt", "lte", "like", "not_like", "is_null", "is_not_null", "in", "not_in", "between"}
                for sf in sf_list:
                    fname = sf.get("field", "")
                    op = sf.get("operator", "")
                    val = sf.get("value")
                    if fname not in valid_field_names or op not in _OP_MAP:
                        continue
                    if check_sql_injection(str(val) if val is not None else ""):
                        continue
                    col_expr = _quote_col(ds.db_type, fname)
                    # For non-date types, use CAST; for date/time comparisons, use direct column
                    field_cfg = next((f for f in display_fields if f.field_name == fname), None)
                    dtype = (field_cfg.db_data_type or "").lower() if field_cfg else ""
                    is_date = any(dt in dtype for dt in ("date", "time", "timestamp"))
                    is_numeric = any(dt in dtype for dt in ("int", "decimal", "numeric", "float", "double", "real", "bigint", "smallint"))

                    if op == "eq":
                        if is_numeric or is_date:
                            where_parts.append(f"{col_expr} = {ph}")
                        else:
                            where_parts.append(f"{_cast_to_text(ds.db_type, col_expr)} = {ph}")
                        params.append(str(val))
                    elif op == "neq":
                        if is_numeric or is_date:
                            where_parts.append(f"{col_expr} != {ph}")
                        else:
                            where_parts.append(f"{_cast_to_text(ds.db_type, col_expr)} != {ph}")
                        params.append(str(val))
                    elif op == "gt":
                        where_parts.append(f"{col_expr} > {ph}")
                        params.append(str(val))
                    elif op == "gte":
                        where_parts.append(f"{col_expr} >= {ph}")
                        params.append(str(val))
                    elif op == "lt":
                        where_parts.append(f"{col_expr} < {ph}")
                        params.append(str(val))
                    elif op == "lte":
                        where_parts.append(f"{col_expr} <= {ph}")
                        params.append(str(val))
                    elif op == "like":
                        where_parts.append(f"{_cast_to_text(ds.db_type, col_expr)} LIKE {ph}")
                        params.append(f"%{val}%")
                    elif op == "not_like":
                        where_parts.append(f"{_cast_to_text(ds.db_type, col_expr)} NOT LIKE {ph}")
                        params.append(f"%{val}%")
                    elif op == "is_null":
                        where_parts.append(f"{col_expr} IS NULL")
                    elif op == "is_not_null":
                        where_parts.append(f"{col_expr} IS NOT NULL")
                    elif op == "in" and isinstance(val, list):
                        phs = ", ".join([ph] * len(val))
                        where_parts.append(f"{_cast_to_text(ds.db_type, col_expr)} IN ({phs})")
                        params.extend([str(v) for v in val])
                    elif op == "not_in" and isinstance(val, list):
                        phs = ", ".join([ph] * len(val))
                        where_parts.append(f"{_cast_to_text(ds.db_type, col_expr)} NOT IN ({phs})")
                        params.extend([str(v) for v in val])
                    elif op == "between" and isinstance(val, list) and len(val) == 2:
                        where_parts.append(f"{col_expr} BETWEEN {ph} AND {ph}")
                        params.extend([str(val[0]), str(val[1])])
            except (json.JSONDecodeError, TypeError):
                pass

        where_sql = ""
        if where_parts:
            where_sql = " WHERE " + " AND ".join(where_parts)

        # Count
        _exec(cur, ds.db_type, f"SELECT COUNT(*) FROM {qt}{where_sql}", params)
        total = cur.fetchone()[0]

        # Select columns
        cols_sql = ", ".join(_quote_col(ds.db_type, c) for c in col_names)

        offset_val = (page - 1) * page_size
        if ds.db_type == "sqlserver":
            pk_fields = tc.primary_key_fields.split(",")
            order_col = _quote_col(ds.db_type, pk_fields[0].strip())
            data_sql = f"SELECT {cols_sql} FROM {qt}{where_sql} ORDER BY {order_col} OFFSET {ph} ROWS FETCH NEXT {ph} ROWS ONLY"
            _exec(cur, ds.db_type, data_sql, params + [offset_val, page_size])
        elif ds.db_type in ("oracle", "dm"):
            # Oracle/达梦 使用 OFFSET-FETCH (12c+) 或 ROWNUM 嵌套
            pk_fields = tc.primary_key_fields.split(",")
            order_col = _quote_col(ds.db_type, pk_fields[0].strip())
            # 使用子查询 + ROWNUM 兼容老版本 Oracle 和达梦
            inner_where = where_sql
            data_sql = f"SELECT * FROM (SELECT a.*, ROWNUM rn FROM (SELECT {cols_sql} FROM {qt}{inner_where} ORDER BY {order_col}) a WHERE ROWNUM <= {offset_val + page_size}) WHERE rn > {offset_val}"
            _exec(cur, ds.db_type, data_sql, params)
        else:
            data_sql = f"SELECT {cols_sql} FROM {qt}{where_sql} LIMIT {ph} OFFSET {ph}"
            _exec(cur, ds.db_type, data_sql, params + [page_size, offset_val])

        raw_rows = cur.fetchall()
        rows = []
        for raw in raw_rows:
            row_dict = {}
            for i, cn in enumerate(col_names):
                row_dict[cn] = str(raw[i]) if raw[i] is not None else None
            rows.append(row_dict)

        # Build column meta (v2.4: include editable_roles for field-level permission)
        columns_meta = []
        pk_set = set(f.strip() for f in tc.primary_key_fields.split(","))
        user_role = user.role if user else "readonly"
        for f in display_fields:
            # v2.4: field-level permission — if editable_roles is set, only those roles can edit
            effective_editable = f.is_editable
            if effective_editable and f.editable_roles:
                allowed_roles = [r.strip() for r in f.editable_roles.split(",") if r.strip()]
                if allowed_roles and user_role not in allowed_roles:
                    effective_editable = 0

            columns_meta.append({
                "field_name": f.field_name,
                "field_alias": f.field_alias or f.field_name,
                "db_data_type": f.db_data_type,
                "is_primary_key": f.is_primary_key,
                "is_editable": effective_editable,
                "is_system_field": f.is_system_field,
                "editable_roles": f.editable_roles,
            })

        return {
            "columns": columns_meta,
            "rows": sanitize_dict(rows),
            "total": total,
            "page": page,
            "page_size": page_size,
            "allow_delete_rows": tc.allow_delete_rows if user.role in ("admin", "operator") else 0,
        }
    except Exception as e:
        raise HTTPException(500, t("data_maintenance.query_failed", error=str(e)))
    finally:
        conn.close()


# ─────────────────────────────────────────────
# P2-3: 模板导出 (v2.0: 空白行主键可编辑 + _meta 区分)
# ─────────────────────────────────────────────

@router.post("/{table_config_id}/export")
def export_template(
    table_config_id: int,
    export_type: str = Query("all", regex="^(current|all)$"),
    keyword: Optional[str] = None,
    field_filters: Optional[str] = None,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin", "operator")),
):
    """导出含隐藏元信息的 Excel 模板。"""
    import openpyxl
    from openpyxl.utils import get_column_letter

    tc = _get_tc(db, table_config_id)
    ds = _get_ds(db, tc.datasource_id)
    pwd = decrypt_password(ds.password_encrypted)
    fields = _get_fields(db, table_config_id)
    export_fields = [f for f in fields if f.include_in_export]
    if not export_fields:
        raise HTTPException(400, t("data_maintenance.no_export_fields"))

    col_names = [f.field_name for f in export_fields]
    pk_set = set(f.strip() for f in tc.primary_key_fields.split(","))
    ph = _placeholder(ds.db_type)

    # Query data from remote DB
    conn = _connect(
        ds.db_type, ds.host, ds.port, ds.username, pwd,
        tc.db_name, tc.schema_name, ds.charset, ds.connect_timeout_seconds or 10,
    )
    try:
        cur = conn.cursor()
        qt = _qualified_table(ds.db_type, tc.table_name, tc.schema_name)

        where_parts: List[str] = []
        params: list = []
        if export_type == "current" and (keyword or field_filters):
            if keyword:
                kw_parts = []
                for f in export_fields:
                    kw_parts.append(f"{_cast_to_text(ds.db_type, _quote_col(ds.db_type, f.field_name))} LIKE {ph}")
                    params.append(f"%{keyword}%")
                where_parts.append(f"({' OR '.join(kw_parts)})")
            if field_filters:
                try:
                    ff = json.loads(field_filters)
                    for fname, fval in ff.items():
                        if fval and any(f.field_name == fname for f in export_fields):
                            where_parts.append(f"{_cast_to_text(ds.db_type, _quote_col(ds.db_type, fname))} LIKE {ph}")
                            params.append(f"%{fval}%")
                except (json.JSONDecodeError, TypeError):
                    pass

        where_sql = ""
        if where_parts:
            where_sql = " WHERE " + " AND ".join(where_parts)

        cols_sql = ", ".join(_quote_col(ds.db_type, c) for c in col_names)
        data_sql = f"SELECT {cols_sql} FROM {qt}{where_sql}"
        _exec(cur, ds.db_type, data_sql, params)
        raw_rows = cur.fetchall()
    except Exception as e:
        raise HTTPException(500, t("data_maintenance.query_failed", error=str(e)))
    finally:
        conn.close()

    # Generate Excel with openpyxl
    from openpyxl.styles import Protection as CellProtection, PatternFill, Font as XlFont
    from openpyxl.worksheet.protection import SheetProtection

    # v3.5: protection password (internal, not exposed to user)
    _SHEET_PROTECTION_PASSWORD = "DOW_tpl_v35_sec"

    batch_no = _gen_batch("EXP")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据"

    RESERVED_BLANK_ROWS = 50  # 预留空白行数
    locked_cell = CellProtection(locked=True)
    unlocked_cell = CellProtection(locked=False)

    # v3.5: Visual style fills
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = XlFont(bold=True, color="FFFFFF", size=11)
    readonly_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    editable_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    blank_zone_fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")

    # Build editable field set for quick lookup (v2.4: respect editable_roles)
    user_role = user.role if user else "readonly"
    editable_field_names = set()
    for f in export_fields:
        if f.is_editable:
            if f.editable_roles:
                allowed_roles = [r.strip() for r in f.editable_roles.split(",") if r.strip()]
                if allowed_roles and user_role not in allowed_roles:
                    continue
            editable_field_names.add(f.field_name)

    data_row_count = len(raw_rows)

    # Row 1: Header (field aliases) — always locked, v3.5: blue bg + white bold font
    for i, f in enumerate(export_fields, 1):
        cell = ws.cell(row=1, column=i, value=f.field_alias or f.field_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.protection = locked_cell

    # Row 2+: Data rows — v3.5: visual color coding
    for row_idx, raw in enumerate(raw_rows, 2):
        for col_idx, (val, ef) in enumerate(zip(raw, export_fields), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
            if ef.field_name in pk_set:
                # 已有数据行的主键列 — 锁定 + 灰色
                cell.protection = locked_cell
                cell.fill = readonly_fill
            elif ef.field_name in editable_field_names:
                cell.protection = unlocked_cell
                cell.fill = editable_fill
            else:
                cell.protection = locked_cell
                cell.fill = readonly_fill

    # Reserved blank rows (for new-row support) — v2.0: 主键列也解锁, v3.5: 浅黄色背景
    blank_start = 2 + data_row_count
    for row_idx in range(blank_start, blank_start + RESERVED_BLANK_ROWS):
        for col_idx, ef in enumerate(export_fields, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            cell.fill = blank_zone_fill
            if ef.field_name in pk_set:
                # v2.0: 空白行的主键列 — 解锁，允许用户填写新主键
                cell.protection = unlocked_cell
            elif ef.field_name in editable_field_names:
                cell.protection = unlocked_cell
            else:
                cell.protection = locked_cell

    # v3.10: Add "_操作" column (last column) for delete marking
    from openpyxl.comments import Comment as XlComment
    op_col = len(export_fields) + 1
    op_header = ws.cell(row=1, column=op_col, value="_操作")
    op_header.font = header_font
    op_header.fill = header_fill
    op_header.protection = locked_cell
    op_header.comment = XlComment("填 DELETE 标记删除", "系统")
    # Data rows — unlocked so user can mark DELETE
    for row_idx in range(2, 2 + data_row_count):
        cell = ws.cell(row=row_idx, column=op_col, value="")
        cell.protection = unlocked_cell
        cell.fill = editable_fill
    # Blank rows
    for row_idx in range(blank_start, blank_start + RESERVED_BLANK_ROWS):
        cell = ws.cell(row=row_idx, column=op_col, value="")
        cell.protection = unlocked_cell
        cell.fill = blank_zone_fill

    # Enable worksheet protection — v3.5: with password, stricter settings
    ws.protection = SheetProtection(
        sheet=True,
        password=_SHEET_PROTECTION_PASSWORD,
        formatColumns=False,
        formatRows=False,
        formatCells=False,
        insertRows=False,
        deleteRows=True,
        deleteColumns=True,
        insertColumns=True,
        sort=False,
        autoFilter=False,
    )

    # Hidden meta sheet — v2.0: 增加 data_row_count 和 blank_row_start 标记
    meta_ws = wb.create_sheet("_meta")
    meta_info = {
        "datasource_id": tc.datasource_id,
        "table_config_id": tc.id,
        "config_version": tc.config_version,
        "export_time": _now_bjt().isoformat(),
        "export_batch_no": batch_no,
        "field_codes": [f.field_name for f in export_fields],
        "field_aliases": [f.field_alias or f.field_name for f in export_fields],
        "primary_key_fields": list(pk_set),
        "structure_hash": tc.structure_version_hash,
        "data_row_count": data_row_count,
        "blank_row_start": blank_start,
        "reserved_blank_rows": RESERVED_BLANK_ROWS,
        "allow_insert_rows": tc.allow_insert_rows,
        "has_operation_column": True,
    }
    meta_ws.cell(row=1, column=1, value=json.dumps(meta_info, ensure_ascii=False))
    meta_ws.sheet_state = "hidden"

    # Auto-width
    for i, f in enumerate(export_fields, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max(12, len(f.field_alias or f.field_name) * 2 + 4)
    # v3.10: auto-width for operation column
    ws.column_dimensions[get_column_letter(op_col)].width = 12

    file_name = f"{tc.table_alias or tc.table_name}_{batch_no}.xlsx"
    file_path = os.path.join(EXPORT_DIR, file_name)
    wb.save(file_path)

    # Record export log
    log = TemplateExportLog(
        export_batch_no=batch_no,
        table_config_id=tc.id,
        datasource_id=tc.datasource_id,
        export_type=export_type,
        row_count=len(raw_rows),
        field_count=len(export_fields),
        template_version=tc.config_version,
        file_name=file_name,
        file_path=file_path,
        export_filters_json=json.dumps({"keyword": keyword, "field_filters": field_filters}, ensure_ascii=False) if keyword or field_filters else None,
        operator_user=_get_username(user),
    )
    db.add(log)
    log_operation(db, "数据维护", "导出模板", "success",
                  target_id=tc.id, target_name=tc.table_name,
                  message=f"导出模板 {file_name}，{len(raw_rows)} 行",
                  operator=_get_username(user))
    db.commit()

    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=file_name,
        headers={"X-Export-Batch-No": batch_no},
    )


@router.get("/{table_config_id}/export-info")
def get_export_info(table_config_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    """获取导出前的预估信息。"""
    tc = _get_tc(db, table_config_id)
    ds = _get_ds(db, tc.datasource_id)
    pwd = decrypt_password(ds.password_encrypted)
    fields = _get_fields(db, table_config_id)
    export_fields = [f for f in fields if f.include_in_export]

    conn = _connect(
        ds.db_type, ds.host, ds.port, ds.username, pwd,
        tc.db_name, tc.schema_name, ds.charset, ds.connect_timeout_seconds or 10,
    )
    try:
        cur = conn.cursor()
        qt = _qualified_table(ds.db_type, tc.table_name, tc.schema_name)
        cur.execute(f"SELECT COUNT(*) FROM {qt}")
        total = cur.fetchone()[0]
    except Exception:
        total = -1
    finally:
        conn.close()

    return {
        "table_config_id": tc.id,
        "table_alias": tc.table_alias,
        "table_name": tc.table_name,
        "config_version": tc.config_version,
        "field_count": len(export_fields),
        "estimated_rows": total,
    }


# ─────────────────────────────────────────────
# P2-5: 模板导入 (v2.0: 支持新增行)
# ─────────────────────────────────────────────

@router.post("/{table_config_id}/import")
async def import_template(
    table_config_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin", "operator")),
):
    """上传平台模板，解析校验，生成差异数据。不直接写库。"""
    import openpyxl

    tc = _get_tc(db, table_config_id)
    ds = _get_ds(db, tc.datasource_id)
    pwd = decrypt_password(ds.password_encrypted)
    fields = _get_fields(db, table_config_id)

    # Validate uploaded file
    file_name = file.filename or "upload.xlsx"
    content = await file.read()
    
    upload_error = validate_upload_file(file_name, content=content)
    if upload_error:
        raise HTTPException(400, upload_error)
    
    # Save uploaded file with sanitized name
    safe_name = os.path.basename(file_name)  # prevent path traversal
    save_path = os.path.join(UPLOAD_DIR, f"{uuid.uuid4().hex}_{safe_name}")
    with open(save_path, "wb") as f:
        f.write(content)

    # Parse Excel
    try:
        wb = openpyxl.load_workbook(save_path, data_only=True)
    except Exception as e:
        raise HTTPException(400, t("data_maintenance.file_parse_failed", error=str(e)))

    # ── 1. Validate meta sheet ──
    if "_meta" not in wb.sheetnames:
        raise HTTPException(400, t("data_maintenance.not_platform_template"))

    meta_ws = wb["_meta"]
    meta_raw = meta_ws.cell(row=1, column=1).value
    if not meta_raw:
        raise HTTPException(400, t("data_maintenance.meta_empty"))

    try:
        meta = json.loads(meta_raw)
    except json.JSONDecodeError:
        raise HTTPException(400, t("data_maintenance.meta_format_error"))

    # ── 2. Validate template legitimacy ──
    errors: List[dict] = []
    warnings: List[dict] = []

    if meta.get("table_config_id") != table_config_id:
        raise HTTPException(400, t("data_maintenance.table_id_mismatch", expected=table_config_id, actual=meta.get('table_config_id')))

    if meta.get("datasource_id") != tc.datasource_id:
        raise HTTPException(400, t("data_maintenance.datasource_id_mismatch"))

    # ── 3. Version match ──
    meta_version = meta.get("config_version", 0)
    if tc.strict_template_version and meta_version != tc.config_version:
        raise HTTPException(400, t("data_maintenance.config_version_mismatch", template_ver=meta_version, current_ver=tc.config_version))

    # ── 4. Read data sheet ──
    data_ws = wb["数据"] if "数据" in wb.sheetnames else wb.worksheets[0]

    header_row = [cell.value for cell in data_ws[1]]
    header_row = [h for h in header_row if h is not None]

    pk_fields = set(meta.get("primary_key_fields", []))

    # v3.10: Detect "_操作" (operation) column for DELETE marking
    _OPERATION_COL_NAME = "_操作"
    _DELETE_MARKERS = {"DELETE", "delete", "删除"}
    has_operation_col = _OPERATION_COL_NAME in header_row or meta.get("has_operation_column", False)
    operation_col_idx = None
    if _OPERATION_COL_NAME in header_row:
        operation_col_idx = header_row.index(_OPERATION_COL_NAME)
        # Remove operation column from header for field matching
        header_row = [h for h in header_row if h != _OPERATION_COL_NAME]

    # Build field map
    import_fields = [f for f in fields if f.include_in_import or f.is_primary_key]
    export_fields = [f for f in fields if f.include_in_export]
    field_alias_to_name = {f.field_alias or f.field_name: f.field_name for f in export_fields}
    field_name_map = {f.field_name: f for f in fields}

    # ── 4.1 列数校验 ──
    expected_aliases = meta.get("field_aliases") or [f.field_alias or f.field_name for f in export_fields]
    expected_col_count = len(expected_aliases)
    actual_col_count = len(header_row)
    if actual_col_count != expected_col_count:
        raise HTTPException(
            400,
            t("data_maintenance.col_count_mismatch", expected=expected_col_count, actual=actual_col_count),
        )

    # ── 4.2 列名逐列校验 ──
    mismatched_cols: List[str] = []
    for idx, (expected, actual) in enumerate(zip(expected_aliases, header_row)):
        if expected != actual:
            mismatched_cols.append(t("data_maintenance.col_name_mismatch", col=idx+1, expected=expected, actual=actual))
    if mismatched_cols:
        raise HTTPException(
            400,
            '; '.join(mismatched_cols),
        )

    # ── 5. Field completeness check ──
    # v3.10: build mapped_cols using the cleaned header_row (without _操作 column)
    # The actual Excel column indices need to account for the removed operation column
    mapped_cols: Dict[int, str] = {}
    _real_header = [cell.value for cell in data_ws[1]]
    _real_header = [h for h in _real_header if h is not None]
    for i, h in enumerate(_real_header):
        if h == _OPERATION_COL_NAME:
            continue  # skip operation column
        if h in field_alias_to_name:
            mapped_cols[i] = field_alias_to_name[h]
        elif h in field_name_map:
            mapped_cols[i] = h

    mapped_field_names = set(mapped_cols.values())
    for f in import_fields:
        if f.field_name not in mapped_field_names and f.field_name in set(f2.field_name for f2 in export_fields):
            if f.is_primary_key:
                errors.append({"row": 0, "field": f.field_name, "type": "field_missing", "message": t("data_maintenance.pk_field_missing", field=f.field_name)})
            else:
                warnings.append({"row": 0, "field": f.field_name, "type": "field_missing", "message": t("data_maintenance.field_missing", field=f.field_name)})

    if any(e["type"] == "field_missing" for e in errors):
        batch_no = _gen_batch("IMP")
        log = ImportTaskLog(
            import_batch_no=batch_no,
            table_config_id=tc.id,
            datasource_id=tc.datasource_id,
            related_export_batch_no=meta.get("export_batch_no"),
            import_file_name=file_name,
            import_file_path=save_path,
            template_version=meta_version,
            total_row_count=0,
            passed_row_count=0,
            warning_row_count=0,
            failed_row_count=len(errors),
            diff_row_count=0,
            new_row_count=0,
            validation_status="failed",
            validation_message="主键字段缺失",
            error_detail_json=json.dumps(errors, ensure_ascii=False),
            import_status="validated",
            operator_user=_get_username(user),
        )
        db.add(log)
        db.commit()
        return {
            "task_id": log.id,
            "import_batch_no": batch_no,
            "validation_status": "failed",
            "total": 0, "passed": 0, "failed": len(errors), "warnings": len(warnings),
            "diff_count": 0, "new_count": 0,
            "errors": errors,
            "warnings_list": warnings,
        }

    # ── 6. Read rows & validate ──
    pk_col_indices = [i for i, fn in mapped_cols.items() if fn in pk_fields]
    data_rows: List[dict] = []
    seen_pks: Dict[str, int] = {}

    # v2.0: meta 中的 data_row_count 用于区分已有行和新增行范围（仅辅助）
    meta_data_row_count = meta.get("data_row_count", 0)

    # v3.10 fix: resolve the real Excel column index for the operation column
    _real_op_col_idx = None
    if has_operation_col:
        _all_headers = [cell.value for cell in data_ws[1]]
        for _hi, _hv in enumerate(_all_headers):
            if _hv == _OPERATION_COL_NAME:
                _real_op_col_idx = _hi  # 0-based index in Excel row
                break

    for row_idx in range(2, data_ws.max_row + 1):
        # Read all columns from the real header (including _操作)
        _all_cells_count = len(_real_header)
        row_cells_all = [data_ws.cell(row=row_idx, column=i + 1).value for i in range(_all_cells_count)]
        # Build row_cells excluding the operation column for field mapping
        row_cells = [row_cells_all[i] for i in range(len(row_cells_all)) if i != _real_op_col_idx]

        # v3.10 fix: Read the operation column value for DELETE detection
        _row_operation = None
        if _real_op_col_idx is not None and _real_op_col_idx < len(row_cells_all):
            _op_val = row_cells_all[_real_op_col_idx]
            if _op_val is not None:
                _row_operation = str(_op_val).strip()

        # Skip entirely empty rows (check all cells including operation)
        if all(c is None or str(c).strip() == "" for c in row_cells) and not _row_operation:
            continue

        row_data: Dict[str, Optional[str]] = {}
        row_errors: List[dict] = []
        row_warnings: List[dict] = []

        for col_i, fname in mapped_cols.items():
            val = row_cells[col_i] if col_i < len(row_cells) else None
            str_val = str(val).strip() if val is not None else None
            row_data[fname] = str_val
            fc = field_name_map.get(fname)
            if not fc:
                continue

            # Required check
            if fc.is_required and (str_val is None or str_val == ""):
                row_errors.append({
                    "row": row_idx, "field": fname,
                    "type": "required", "value": str_val,
                    "message": t("data_maintenance.row_field_required", row=row_idx, field=fc.field_alias or fname),
                })

            # Length check
            if fc.max_length and str_val and len(str_val) > fc.max_length:
                row_errors.append({
                    "row": row_idx, "field": fname,
                    "type": "length", "value": str_val,
                    "message": t("data_maintenance.row_field_too_long", row=row_idx, field=fc.field_alias or fname, max_len=fc.max_length),
                })

            # Data type check — v3.5: enhanced with date validation
            if str_val and fc.db_data_type:
                dtype = fc.db_data_type.lower()
                if any(dt in dtype for dt in ("int", "bigint", "smallint", "tinyint")):
                    try:
                        int(str_val)
                    except ValueError:
                        row_errors.append({
                            "row": row_idx, "field": fname,
                            "type": "data_type", "value": str_val,
                            "message": t("data_maintenance.row_field_expect_int", row=row_idx, field=fc.field_alias or fname),
                        })
                elif any(dt in dtype for dt in ("decimal", "numeric", "float", "double", "real")):
                    try:
                        float(str_val)
                    except ValueError:
                        row_errors.append({
                            "row": row_idx, "field": fname,
                            "type": "data_type", "value": str_val,
                            "message": t("data_maintenance.row_field_expect_number", row=row_idx, field=fc.field_alias or fname),
                        })
                elif any(dt in dtype for dt in ("date", "time", "timestamp", "datetime")):
                    # v3.5: date/datetime format validation
                    from dateutil import parser as date_parser
                    try:
                        date_parser.parse(str_val)
                    except (ValueError, OverflowError):
                        row_errors.append({
                            "row": row_idx, "field": fname,
                            "type": "data_type", "value": str_val,
                            "message": t("data_maintenance.row_field_expect_date", row=row_idx, field=fc.field_alias or fname, value=str_val),
                        })

            # Enum check
            if fc.enum_options_json and str_val:
                try:
                    options = json.loads(fc.enum_options_json)
                    if isinstance(options, list) and str_val not in options:
                        row_errors.append({
                            "row": row_idx, "field": fname,
                            "type": "enum", "value": str_val,
                            "message": t("data_maintenance.row_field_not_in_enum", row=row_idx, field=fc.field_alias or fname),
                        })
                except json.JSONDecodeError:
                    pass

        # PK check
        pk_vals = tuple(row_data.get(mapped_cols[i], "") for i in pk_col_indices)
        pk_key = "|".join(str(v) for v in pk_vals)
        if any(v is None or v == "" for v in pk_vals):
            row_errors.append({
                "row": row_idx, "field": ",".join(pk_fields),
                "type": "pk_empty", "value": pk_key,
                "message": t("data_maintenance.row_pk_empty", row=row_idx),
            })
        elif pk_key in seen_pks:
            row_errors.append({
                "row": row_idx, "field": ",".join(pk_fields),
                "type": "duplicate", "value": pk_key,
                "message": t("data_maintenance.row_pk_duplicate", row=row_idx, dup_row=seen_pks[pk_key]),
            })
        else:
            seen_pks[pk_key] = row_idx

        # v3.10 fix: detect DELETE-marked rows
        is_delete_row = _row_operation in _DELETE_MARKERS if _row_operation else False

        data_rows.append({
            "row_num": row_idx,
            "data": row_data,
            "errors": row_errors,
            "warnings": row_warnings,
            "pk_key": pk_key,
            "is_delete": is_delete_row,
        })
        errors.extend(row_errors)
        warnings.extend(row_warnings)

    # ── 7. Generate diff (original vs new) — v2.0 支持新增行, v3.10 支持删除行 ──
    diff_rows: List[dict] = []
    new_rows: List[dict] = []
    delete_rows: List[dict] = []
    passed_count = 0

    if not all(r["errors"] for r in data_rows):
        # Fetch current data from DB for comparison
        conn = _connect(
            ds.db_type, ds.host, ds.port, ds.username, pwd,
            tc.db_name, tc.schema_name, ds.charset, ds.connect_timeout_seconds or 10,
        )
        try:
            cur = conn.cursor()
            qt = _qualified_table(ds.db_type, tc.table_name, tc.schema_name)
            all_field_names = [f.field_name for f in export_fields]

            cols_sql = ", ".join(_quote_col(ds.db_type, c) for c in all_field_names)

            cur.execute(f"SELECT {cols_sql} FROM {qt}")
            db_rows = cur.fetchall()

            pk_field_indices = [all_field_names.index(p) for p in pk_fields if p in all_field_names]
            db_pk_map: Dict[str, dict] = {}
            for db_row in db_rows:
                pk_val = "|".join(str(db_row[i]) if db_row[i] is not None else "" for i in pk_field_indices)
                row_dict = {}
                for j, fn in enumerate(all_field_names):
                    row_dict[fn] = str(db_row[j]) if db_row[j] is not None else None
                db_pk_map[pk_val] = row_dict
        except Exception as e:
            raise HTTPException(500, t("data_maintenance.query_original_failed", error=str(e)))
        finally:
            conn.close()

        db_all_pk_set = set(db_pk_map.keys())

        # v3.5: 主键不可变校验 — 原数据行的主键不允许修改
        # 通过 meta 中的 data_row_count 和 blank_row_start 判断哪些是原数据行
        meta_blank_start = meta.get("blank_row_start", 2 + meta_data_row_count)
        for row in data_rows:
            if row["errors"]:
                continue
            pk_key = row["pk_key"]
            row_num = row["row_num"]
            # 如果是原数据区域的行（非空白区新增行），但 PK 在 DB 中找不到，说明 PK 被修改了
            if row_num < meta_blank_start and pk_key not in db_all_pk_set:
                row["errors"].append({
                    "row": row_num, "field": ",".join(pk_fields),
                    "type": "pk_modified", "value": pk_key,
                    "message": t("data_maintenance.row_pk_modified", row=row_num, old_val="(原始值)", new_val=pk_key),
                })
                errors.append(row["errors"][-1])

        # v2.0: 区分已有行和新增行; v3.10: 支持删除行
        editable_fields = [f for f in fields if f.is_editable and f.include_in_import]

        for row in data_rows:
            if row["errors"]:
                continue
            pk_key = row["pk_key"]

            # v3.10 fix: handle DELETE-marked rows
            if row.get("is_delete"):
                if pk_key in db_all_pk_set:
                    if tc.allow_delete_rows:
                        original = db_pk_map[pk_key]
                        delete_rows.append({
                            "row_num": row["row_num"],
                            "data": row["data"],
                            "pk_key": pk_key,
                            "change_type": "delete",
                        })
                        # Record all fields as delete diff
                        for ef in export_fields:
                            fn = ef.field_name
                            old_val = original.get(fn)
                            if old_val is not None:
                                diff_rows.append({
                                    "row_num": row["row_num"],
                                    "pk_key": pk_key,
                                    "field_name": fn,
                                    "field_alias": ef.field_alias or fn,
                                    "old_value": old_val,
                                    "new_value": None,
                                    "status": "deleted",
                                    "change_type": "delete",
                                })
                        passed_count += 1
                    else:
                        row["errors"].append({
                            "row": row["row_num"], "field": "_操作",
                            "type": "delete_not_allowed", "value": pk_key,
                            "message": f"第 {row['row_num']} 行标记删除，但该表未启用删除功能",
                        })
                        errors.append(row["errors"][-1])
                else:
                    row["errors"].append({
                        "row": row["row_num"], "field": ",".join(pk_fields),
                        "type": "pk_not_found", "value": pk_key,
                        "message": f"第 {row['row_num']} 行标记删除，但主键 {pk_key} 在数据库中不存在",
                    })
                    errors.append(row["errors"][-1])
                continue

            if pk_key not in db_all_pk_set:
                # v2.0: 新增行 — 如果允许新增则标记为 insert，否则报错
                if tc.allow_insert_rows:
                    # 记录为新增行
                    new_rows.append({
                        "row_num": row["row_num"],
                        "data": row["data"],
                        "pk_key": pk_key,
                        "change_type": "insert",
                    })
                    # 将所有字段记入 diff
                    for ef in export_fields:
                        fn = ef.field_name
                        new_val = row["data"].get(fn)
                        if new_val is not None and new_val != "":
                            diff_rows.append({
                                "row_num": row["row_num"],
                                "pk_key": pk_key,
                                "field_name": fn,
                                "field_alias": ef.field_alias or fn,
                                "old_value": None,
                                "new_value": new_val,
                                "status": "new",
                                "change_type": "insert",
                            })
                    passed_count += 1
                else:
                    row["errors"].append({
                        "row": row["row_num"], "field": ",".join(pk_fields),
                        "type": "pk_not_found", "value": pk_key,
                        "message": t("data_maintenance.row_pk_not_in_db_no_insert", row=row['row_num'], pk=pk_key),
                    })
                    errors.append(row["errors"][-1])
                continue

            # 已有行 — 对比差异
            original = db_pk_map[pk_key]
            has_diff = False
            for ef in editable_fields:
                fn = ef.field_name
                new_val = row["data"].get(fn)
                old_val = original.get(fn)
                if new_val != old_val:
                    has_diff = True
                    diff_rows.append({
                        "row_num": row["row_num"],
                        "pk_key": pk_key,
                        "field_name": fn,
                        "field_alias": ef.field_alias or fn,
                        "old_value": old_val,
                        "new_value": new_val,
                        "status": "changed",
                        "change_type": "update",
                    })
            passed_count += 1

    # ── 7.5 AI Smart Validation (v3.0) ──
    ai_warnings = []
    try:
        from app.ai.ai_engine import AIEngine
        ai_engine = AIEngine(db)
        if ai_engine.is_enabled and ai_engine.is_feature_enabled("data_validate"):
            from app.routers.ai_validate import (
                _get_validate_config, _sample_historical_data,
                _run_outlier_checks, _run_format_checks,
                _run_duplicate_checks, _run_cross_field_checks,
            )
            validate_config = _get_validate_config(db)
            skip_fields_set = set(validate_config.get("skip_fields", []))
            sample_size = validate_config.get("history_sample_size", 1000)
            hist_data = _sample_historical_data(ds, tc, fields, sample_size)

            if hist_data:
                # Prepare import data with row numbers
                ai_import_data = []
                for row in data_rows:
                    if not row["errors"]:
                        d = dict(row["data"])
                        d["_row_num"] = row["row_num"]
                        ai_import_data.append(d)

                if ai_import_data:
                    ai_warnings.extend(_run_outlier_checks(ai_import_data, hist_data, fields, validate_config, skip_fields_set))
                    ai_warnings.extend(_run_format_checks(ai_import_data, hist_data, fields, skip_fields_set))
                    ai_warnings.extend(_run_duplicate_checks(ai_import_data, hist_data, fields, skip_fields_set))
                    ai_warnings.extend(_run_cross_field_checks(ai_import_data, fields, skip_fields_set))
    except Exception:
        pass  # AI validation failure should not block import

    # ── 8. Save import task log ──
    batch_no = _gen_batch("IMP")
    total_rows = len(data_rows)
    actual_failed = len([r for r in data_rows if r["errors"]])
    actual_passed = total_rows - actual_failed

    validation_status = "success" if actual_failed == 0 else ("failed" if actual_passed == 0 else "partial")

    log = ImportTaskLog(
        import_batch_no=batch_no,
        table_config_id=tc.id,
        datasource_id=tc.datasource_id,
        related_export_batch_no=meta.get("export_batch_no"),
        import_file_name=file_name,
        import_file_path=save_path,
        template_version=meta_version,
        total_row_count=total_rows,
        passed_row_count=actual_passed,
        warning_row_count=len(warnings),
        failed_row_count=actual_failed,
        diff_row_count=len(diff_rows),
        new_row_count=len(new_rows),
        validation_status=validation_status,
        validation_message=f"总计 {total_rows} 行，通过 {actual_passed}（含新增 {len(new_rows)}，删除 {len(delete_rows)}），失败 {actual_failed}，差异 {len(diff_rows)} 处",
        error_detail_json=json.dumps(errors, ensure_ascii=False) if errors else None,
        import_status="validated",
        operator_user=_get_username(user),
    )
    db.add(log)
    db.flush()
    log_operation(db, "数据维护", "导入模板", validation_status,
                  target_id=tc.id, target_name=tc.table_name,
                  message=f"导入模板 {file_name}，{total_rows} 行，通过 {actual_passed}（新增 {len(new_rows)}），失败 {actual_failed}",
                  operator=_get_username(user))

    # Store diff data in a temp JSON file — v2.0: 包含 new_rows, v3.10: 包含 delete_rows
    diff_file = os.path.join(UPLOAD_DIR, f"diff_{log.id}.json")
    diff_data = {
        "diff_rows": diff_rows,
        "new_rows": new_rows,
        "delete_rows": delete_rows,
        "import_data": [{"row_num": r["row_num"], "data": r["data"], "pk_key": r["pk_key"]}
                        for r in data_rows if not r["errors"]],
    }
    with open(diff_file, "w", encoding="utf-8") as f:
        json.dump(diff_data, f, ensure_ascii=False)

    db.commit()

    return {
        "task_id": log.id,
        "import_batch_no": batch_no,
        "validation_status": validation_status,
        "total": total_rows,
        "passed": actual_passed,
        "failed": actual_failed,
        "warnings": len(warnings),
        "diff_count": len(diff_rows),
        "new_count": len(new_rows),
        "delete_count": len(delete_rows),
        "errors": errors[:100],
        "warnings_list": warnings[:50],
        "ai_warnings": ai_warnings[:100],
        "ai_warnings_count": len(ai_warnings),
    }


# ─────────────────────────────────────────────
# P2-8: 差异预览 (v2.0: 区分更新行和新增行)
# ─────────────────────────────────────────────

@router.get("/import-tasks/{task_id}/diff")
def get_import_diff(task_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    """返回原值/新值对比数据。"""
    task = db.query(ImportTaskLog).filter(ImportTaskLog.id == task_id).first()
    if not task:
        raise HTTPException(404, t("data_maintenance.import_task_not_found"))

    tc = db.query(TableConfig).filter(TableConfig.id == task.table_config_id).first()

    diff_file = os.path.join(UPLOAD_DIR, f"diff_{task_id}.json")
    if not os.path.isfile(diff_file):
        raise HTTPException(404, t("data_maintenance.diff_not_found"))

    with open(diff_file, "r", encoding="utf-8") as f:
        diff_data = json.load(f)

    return {
        "task_id": task_id,
        "import_batch_no": task.import_batch_no,
        "table_config_id": task.table_config_id,
        "table_name": tc.table_name if tc else None,
        "table_alias": tc.table_alias if tc else None,
        "config_version": task.template_version,
        "operator_user": task.operator_user,
        "import_time": task.created_at.isoformat() if task.created_at else None,
        "total_rows": task.total_row_count,
        "passed_rows": task.passed_row_count,
        "failed_rows": task.failed_row_count,
        "diff_count": task.diff_row_count,
        "new_count": task.new_row_count,
        "diff_rows": diff_data.get("diff_rows", []),
        "new_rows": diff_data.get("new_rows", []),
        "validation_status": task.validation_status,
    }


@router.get("/import-tasks/{task_id}/diff-report")
def get_diff_report(task_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    """Generate and download a diff comparison Excel report."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    task = db.query(ImportTaskLog).filter(ImportTaskLog.id == task_id).first()
    if not task:
        raise HTTPException(404, t("data_maintenance.import_task_not_found"))

    tc = db.query(TableConfig).filter(TableConfig.id == task.table_config_id).first()

    diff_file = os.path.join(UPLOAD_DIR, "diff_%d.json" % task_id)
    if not os.path.isfile(diff_file):
        raise HTTPException(404, t("data_maintenance.diff_not_found"))

    with open(diff_file, "r", encoding="utf-8") as f:
        diff_data = json.load(f)

    diff_rows = diff_data.get("diff_rows", [])
    if not diff_rows:
        raise HTTPException(400, t("data_maintenance.no_diff_data"))

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "差异对比报告"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    update_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # yellow
    insert_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # green
    delete_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")  # red

    change_type_map = {
        "update": ("更新", update_fill),
        "insert": ("新增", insert_fill),
        "delete": ("删除", delete_fill),
    }

    # Headers
    headers = ["行号", "主键值", "字段名", "原值", "新值", "变更类型"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, dr in enumerate(diff_rows, 2):
        change_type = dr.get("change_type", "update")
        type_text, row_fill = change_type_map.get(change_type, ("更新", update_fill))

        values = [
            dr.get("row_num", ""),
            dr.get("pk_key", ""),
            dr.get("field_alias", dr.get("field_name", "")),
            dr.get("old_value") if dr.get("old_value") is not None else "",
            dr.get("new_value") if dr.get("new_value") is not None else "",
            type_text,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill

    # Column widths
    col_widths = [8, 20, 20, 25, 25, 12]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Save to temp file
    report_name = "diff_report_%s_%d.xlsx" % (_gen_batch("DIFF"), task_id)
    report_path = os.path.join(EXPORT_DIR, report_name)
    wb.save(report_path)

    return FileResponse(
        report_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=report_name,
    )


@router.post("/import-tasks/{task_id}/retry")
def retry_import_validation(task_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(require_role("admin", "operator"))):
    """Retry validation using the original uploaded file."""
    task = db.query(ImportTaskLog).filter(ImportTaskLog.id == task_id).first()
    if not task:
        raise HTTPException(404, t("data_maintenance.import_task_not_found"))

    # Only allow retry on failed validations
    if task.validation_status not in ("failed", "partial"):
        raise HTTPException(400, t("data_maintenance.revalidate_status_error"))

    # Check original file exists
    if not task.import_file_path or not os.path.isfile(task.import_file_path):
        raise HTTPException(404, t("data_maintenance.upload_file_not_found"))

    # Re-read the original file and re-run the import validation
    # We'll simulate re-upload by reading the file and calling the import logic
    import openpyxl

    tc = db.query(TableConfig).filter(
        TableConfig.id == task.table_config_id,
        TableConfig.is_deleted == 0,
        TableConfig.status == "enabled",
    ).first()
    if not tc:
        raise HTTPException(404, t("data_maintenance.table_not_found"))

    ds = db.query(DatasourceConfig).filter(
        DatasourceConfig.id == tc.datasource_id,
        DatasourceConfig.is_deleted == 0,
    ).first()
    if not ds:
        raise HTTPException(404, t("data_maintenance.datasource_not_found"))

    pwd = decrypt_password(ds.password_encrypted)
    fields = _get_fields(db, tc.id)

    # Parse Excel
    try:
        wb = openpyxl.load_workbook(task.import_file_path, data_only=True)
    except Exception as e:
        raise HTTPException(400, t("data_maintenance.file_parse_failed", error=str(e)))

    if "_meta" not in wb.sheetnames:
        raise HTTPException(400, t("data_maintenance.not_platform_template"))

    meta_ws = wb["_meta"]
    meta_raw = meta_ws.cell(row=1, column=1).value
    if not meta_raw:
        raise HTTPException(400, t("data_maintenance.meta_empty"))

    try:
        meta = json.loads(meta_raw)
    except json.JSONDecodeError:
        raise HTTPException(400, t("data_maintenance.meta_format_error"))

    # Validate
    errors = []
    warnings = []

    meta_version = meta.get("config_version", 0)
    if tc.strict_template_version and meta_version != tc.config_version:
        raise HTTPException(400, t("data_maintenance.config_version_mismatch", template_ver=meta_version, current_ver=tc.config_version))

    data_ws = wb["数据"] if "数据" in wb.sheetnames else wb.worksheets[0]
    header_row = [cell.value for cell in data_ws[1]]
    header_row = [h for h in header_row if h is not None]

    pk_fields = set(meta.get("primary_key_fields", []))
    import_fields = [f for f in fields if f.include_in_import or f.is_primary_key]
    export_fields = [f for f in fields if f.include_in_export]
    field_alias_to_name = {f.field_alias or f.field_name: f.field_name for f in export_fields}
    field_name_map = {f.field_name: f for f in fields}

    mapped_cols = {}
    for i, h in enumerate(header_row):
        if h in field_alias_to_name:
            mapped_cols[i] = field_alias_to_name[h]
        elif h in field_name_map:
            mapped_cols[i] = h

    pk_col_indices = [i for i, fn in mapped_cols.items() if fn in pk_fields]
    data_rows = []
    seen_pks = {}

    for row_idx in range(2, data_ws.max_row + 1):
        row_cells = [data_ws.cell(row=row_idx, column=i + 1).value for i in range(len(header_row))]
        if all(c is None or str(c).strip() == "" for c in row_cells):
            continue

        row_data = {}
        row_errors = []
        for col_i, fname in mapped_cols.items():
            val = row_cells[col_i] if col_i < len(row_cells) else None
            str_val = str(val).strip() if val is not None else None
            row_data[fname] = str_val
            fc = field_name_map.get(fname)
            if not fc:
                continue
            if fc.is_required and (str_val is None or str_val == ""):
                row_errors.append({"row": row_idx, "field": fname, "type": "required", "value": str_val,
                                   "message": t("data_maintenance.row_field_required", row=row_idx, field=fc.field_alias or fname)})
            if fc.max_length and str_val and len(str_val) > fc.max_length:
                row_errors.append({"row": row_idx, "field": fname, "type": "length", "value": str_val,
                                   "message": t("data_maintenance.row_field_too_long", row=row_idx, field=fc.field_alias or fname, max_len=fc.max_length)})

        pk_vals = tuple(row_data.get(mapped_cols[i], "") for i in pk_col_indices)
        pk_key = "|".join(str(v) for v in pk_vals)
        if any(v is None or v == "" for v in pk_vals):
            row_errors.append({"row": row_idx, "field": ",".join(pk_fields), "type": "pk_empty", "value": pk_key,
                               "message": t("data_maintenance.row_pk_empty", row=row_idx)})
        elif pk_key in seen_pks:
            row_errors.append({"row": row_idx, "field": ",".join(pk_fields), "type": "duplicate", "value": pk_key,
                               "message": t("data_maintenance.row_pk_duplicate", row=row_idx, dup_row=seen_pks[pk_key])})
        else:
            seen_pks[pk_key] = row_idx

        data_rows.append({"row_num": row_idx, "data": row_data, "errors": row_errors, "warnings": [], "pk_key": pk_key})
        errors.extend(row_errors)

    # Generate diff
    diff_rows = []
    new_rows = []

    if not all(r["errors"] for r in data_rows):
        conn = _connect(ds.db_type, ds.host, ds.port, ds.username, pwd,
                        tc.db_name, tc.schema_name, ds.charset, ds.connect_timeout_seconds or 10)
        try:
            cur = conn.cursor()
            qt = _qualified_table(ds.db_type, tc.table_name, tc.schema_name)
            all_field_names = [f.field_name for f in export_fields]
            cols_sql = ", ".join(_quote_col(ds.db_type, c) for c in all_field_names)
            cur.execute("SELECT %s FROM %s" % (cols_sql, qt))
            db_rows = cur.fetchall()
            pk_field_indices = [all_field_names.index(p) for p in pk_fields if p in all_field_names]
            db_pk_map = {}
            for db_row in db_rows:
                pk_val = "|".join(str(db_row[i]) if db_row[i] is not None else "" for i in pk_field_indices)
                row_dict = {}
                for j, fn in enumerate(all_field_names):
                    row_dict[fn] = str(db_row[j]) if db_row[j] is not None else None
                db_pk_map[pk_val] = row_dict
        except Exception as e:
            raise HTTPException(500, t("data_maintenance.query_original_failed", error=str(e)))
        finally:
            conn.close()

        db_all_pk_set = set(db_pk_map.keys())
        editable_fields = [f for f in fields if f.is_editable and f.include_in_import]

        for row in data_rows:
            if row["errors"]:
                continue
            pk_key = row["pk_key"]
            if pk_key not in db_all_pk_set:
                if tc.allow_insert_rows:
                    new_rows.append({"row_num": row["row_num"], "data": row["data"], "pk_key": pk_key, "change_type": "insert"})
                    for ef in export_fields:
                        fn = ef.field_name
                        new_val = row["data"].get(fn)
                        if new_val is not None and new_val != "":
                            diff_rows.append({"row_num": row["row_num"], "pk_key": pk_key, "field_name": fn,
                                             "field_alias": ef.field_alias or fn, "old_value": None,
                                             "new_value": new_val, "status": "new", "change_type": "insert"})
                else:
                    row["errors"].append({"row": row["row_num"], "field": ",".join(pk_fields),
                                          "type": "pk_not_found", "value": pk_key,
                                          "message": t("data_maintenance.row_pk_not_in_db", row=row["row_num"])})
                    errors.append(row["errors"][-1])
                continue
            original = db_pk_map[pk_key]
            for ef in editable_fields:
                fn = ef.field_name
                new_val = row["data"].get(fn)
                old_val = original.get(fn)
                if new_val != old_val:
                    diff_rows.append({"row_num": row["row_num"], "pk_key": pk_key, "field_name": fn,
                                     "field_alias": ef.field_alias or fn, "old_value": old_val,
                                     "new_value": new_val, "status": "changed", "change_type": "update"})

    # Save new import task log
    batch_no = _gen_batch("IMP")
    total_rows = len(data_rows)
    actual_failed = len([r for r in data_rows if r["errors"]])
    actual_passed = total_rows - actual_failed
    validation_status = "success" if actual_failed == 0 else ("failed" if actual_passed == 0 else "partial")

    new_log = ImportTaskLog(
        import_batch_no=batch_no,
        table_config_id=tc.id,
        datasource_id=tc.datasource_id,
        related_export_batch_no=meta.get("export_batch_no"),
        import_file_name=task.import_file_name,
        import_file_path=task.import_file_path,
        template_version=meta_version,
        total_row_count=total_rows,
        passed_row_count=actual_passed,
        warning_row_count=len(warnings),
        failed_row_count=actual_failed,
        diff_row_count=len(diff_rows),
        new_row_count=len(new_rows),
        validation_status=validation_status,
        validation_message="重新校验: 总计 %d 行，通过 %d，失败 %d，差异 %d 处" % (total_rows, actual_passed, actual_failed, len(diff_rows)),
        error_detail_json=json.dumps(errors, ensure_ascii=False) if errors else None,
        import_status="validated",
        operator_user=_get_username(user),
    )
    db.add(new_log)
    db.flush()

    # Store diff data
    diff_file = os.path.join(UPLOAD_DIR, "diff_%d.json" % new_log.id)
    diff_data_out = {
        "diff_rows": diff_rows,
        "new_rows": new_rows,
        "import_data": [{"row_num": r["row_num"], "data": r["data"], "pk_key": r["pk_key"]}
                        for r in data_rows if not r["errors"]],
    }
    with open(diff_file, "w", encoding="utf-8") as f:
        json.dump(diff_data_out, f, ensure_ascii=False)

    log_operation(db, "数据维护", "重新校验", validation_status,
                  target_id=tc.id, target_name=tc.table_name,
                  message="重新校验导入 %s（原任务 #%d），%d 行，通过 %d，失败 %d" % (batch_no, task_id, total_rows, actual_passed, actual_failed),
                  operator=_get_username(user))
    db.commit()

    return {
        "task_id": new_log.id,
        "import_batch_no": batch_no,
        "validation_status": validation_status,
        "total": total_rows,
        "passed": actual_passed,
        "failed": actual_failed,
        "warnings": len(warnings),
        "diff_count": len(diff_rows),
        "new_count": len(new_rows),
        "errors": errors[:100],
        "original_task_id": task_id,
    }


@router.get("/import-tasks/{task_id}")
def get_import_task(task_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(get_current_user)):
    """获取导入任务详情。"""
    task = db.query(ImportTaskLog).filter(ImportTaskLog.id == task_id).first()
    if not task:
        raise HTTPException(404, t("data_maintenance.import_task_not_found"))

    errors = []
    if task.error_detail_json:
        try:
            errors = json.loads(task.error_detail_json)
        except json.JSONDecodeError:
            pass

    return {
        "task_id": task.id,
        "import_batch_no": task.import_batch_no,
        "table_config_id": task.table_config_id,
        "datasource_id": task.datasource_id,
        "import_file_name": task.import_file_name,
        "template_version": task.template_version,
        "total_row_count": task.total_row_count,
        "passed_row_count": task.passed_row_count,
        "warning_row_count": task.warning_row_count,
        "failed_row_count": task.failed_row_count,
        "diff_row_count": task.diff_row_count,
        "new_row_count": task.new_row_count,
        "validation_status": task.validation_status,
        "validation_message": task.validation_message,
        "import_status": task.import_status,
        "operator_user": task.operator_user,
        "created_at": task.created_at.isoformat() if task.created_at else None,
        "errors": errors,
    }


# ─────────────────────────────────────────────
# P2-10: 安全回写 (v2.0: UPDATE + INSERT + FieldChangeLog)
# ─────────────────────────────────────────────

@router.post("/import-tasks/{task_id}/writeback")
def writeback(task_id: int, db: Session = Depends(get_db), user: UserAccount = Depends(require_role("admin", "operator"))):
    """执行回写：写前全表备份 → UPDATE/INSERT → 记录逐字段变更日志。"""
    task = db.query(ImportTaskLog).filter(ImportTaskLog.id == task_id).first()
    if not task:
        raise HTTPException(404, t("data_maintenance.import_task_not_found"))
    if task.import_status not in ("validated",):
        raise HTTPException(400, t("data_maintenance.import_status_not_allow_writeback", status=task.import_status))
    if task.validation_status == "failed":
        raise HTTPException(400, t("data_maintenance.all_validation_failed"))

    # v2.2: approval workflow
    if _needs_approval(db, user):
        return _create_approval_request(
            db, user,
            table_config_id=task.table_config_id,
            request_type="writeback",
            import_task_id=task_id,
        )

    tc = _get_tc(db, task.table_config_id)
    ds = _get_ds(db, tc.datasource_id)
    pwd = decrypt_password(ds.password_encrypted)
    fields = _get_fields(db, tc.id)
    field_name_map = {f.field_name: f for f in fields}
    pk_fields_list = [p.strip() for p in tc.primary_key_fields.split(",")]

    # Load diff data
    diff_file = os.path.join(UPLOAD_DIR, f"diff_{task_id}.json")
    if not os.path.isfile(diff_file):
        raise HTTPException(404, t("data_maintenance.diff_not_found"))
    with open(diff_file, "r", encoding="utf-8") as f:
        diff_data = json.load(f)

    import_rows = diff_data.get("import_data", [])
    new_rows_data = diff_data.get("new_rows", [])
    delete_rows_data = diff_data.get("delete_rows", [])
    diff_rows_data = diff_data.get("diff_rows", [])
    if not import_rows and not new_rows_data and not delete_rows_data:
        raise HTTPException(400, t("data_maintenance.no_writeback_data"))

    # 分离更新行、新增行和删除行
    new_pk_set = set(nr["pk_key"] for nr in new_rows_data)
    delete_pk_set = set(dr["pk_key"] for dr in delete_rows_data)
    update_rows = [r for r in import_rows if r["pk_key"] not in new_pk_set and r["pk_key"] not in delete_pk_set]

    wb_batch = _gen_batch("WB")
    bk_batch = _gen_batch("BK")
    started_at = _now_bjt()

    conn = _connect(
        ds.db_type, ds.host, ds.port, ds.username, pwd,
        tc.db_name, tc.schema_name, ds.charset, ds.connect_timeout_seconds or 10,
    )
    try:
        cur = conn.cursor()
        qt = _qualified_table(ds.db_type, tc.table_name, tc.schema_name)
        ph = _placeholder(ds.db_type)

        # ── Step 1: Full table backup ──
        ts = _now_bjt().strftime("%Y%m%d_%H%M%S")
        rand_suffix = uuid.uuid4().hex[:4].upper()
        backup_table_name = f"{tc.table_name}_bak_{ts}_{rand_suffix}"

        _create_backup_table(cur, ds.db_type, qt, backup_table_name, tc.schema_name)

        # Count backup rows
        bk_qt = _qualified_table(ds.db_type, backup_table_name, tc.schema_name)
        cur.execute(f"SELECT COUNT(*) FROM {bk_qt}")
        backup_count = cur.fetchone()[0]
        conn.commit()

        # Record backup version
        backup_rec = TableBackupVersion(
            backup_version_no=bk_batch,
            table_config_id=tc.id,
            datasource_id=tc.datasource_id,
            backup_table_name=backup_table_name,
            source_table_name=tc.table_name,
            source_db_name=tc.db_name,
            source_schema_name=tc.schema_name,
            trigger_type="triggered_by_writeback",
            related_writeback_batch_no=wb_batch,
            record_count=backup_count,
            storage_status="valid",
            can_rollback=1,
            backup_started_at=started_at,
            backup_finished_at=_now_bjt(),
            operator_user=_get_username(user),
        )
        db.add(backup_rec)
        db.flush()

        # Clean old backups
        old_backups = (
            db.query(TableBackupVersion)
            .filter(
                TableBackupVersion.table_config_id == tc.id,
                TableBackupVersion.storage_status == "valid",
            )
            .order_by(TableBackupVersion.id.desc())
            .all()
        )
        if len(old_backups) > tc.backup_keep_count:
            for old_bk in old_backups[tc.backup_keep_count:]:
                try:
                    old_bk_qt = _qualified_table(ds.db_type, old_bk.backup_table_name, old_bk.source_schema_name)
                    _drop_table_if_exists(cur, ds.db_type, old_bk_qt)
                    conn.commit()
                except Exception:
                    pass
                old_bk.storage_status = "expired"
                old_bk.can_rollback = 0

        # ── Step 2: Execute UPDATEs ──
        success_count = 0
        fail_count = 0
        update_count = 0
        insert_count = 0
        delete_count = 0
        failed_details: List[dict] = []
        change_logs: List[dict] = []
        editable_fields = [f for f in fields if f.is_editable and f.include_in_import]

        # Build diff lookup: pk_key+field_name -> {old_value, new_value}
        diff_lookup: Dict[str, List[dict]] = {}
        for dr in diff_rows_data:
            key = dr["pk_key"]
            if key not in diff_lookup:
                diff_lookup[key] = []
            diff_lookup[key].append(dr)

        for irow in update_rows:
            row_data = irow["data"]
            pk_key = irow["pk_key"]

            set_parts = []
            set_params = []
            for ef in editable_fields:
                fn = ef.field_name
                if fn in row_data:
                    new_val = row_data[fn]
                    set_parts.append(f"{_quote_col(ds.db_type, fn)} = {ph}")
                    set_params.append(new_val)

            if not set_parts:
                continue

            # Build WHERE from PK
            where_parts = []
            where_params = []
            pk_vals = pk_key.split("|")
            for i, pkf in enumerate(pk_fields_list):
                pk_val = pk_vals[i] if i < len(pk_vals) else ""
                where_parts.append(f"{_cast_to_text(ds.db_type, _quote_col(ds.db_type, pkf))} = {ph}")
                where_params.append(pk_val)

            update_sql = f"UPDATE {qt} SET {', '.join(set_parts)} WHERE {' AND '.join(where_parts)}"
            try:
                _exec(cur, ds.db_type, update_sql, set_params + where_params)
                success_count += 1
                update_count += 1

                # Record field-level changes
                diffs_for_row = diff_lookup.get(pk_key, [])
                for d in diffs_for_row:
                    if d.get("change_type") == "update":
                        change_logs.append({
                            "row_pk_value": pk_key,
                            "field_name": d["field_name"],
                            "old_value": d.get("old_value"),
                            "new_value": d.get("new_value"),
                            "change_type": "update",
                        })
            except Exception as e:
                fail_count += 1
                failed_details.append({
                    "row_num": irow["row_num"],
                    "pk_key": pk_key,
                    "error": str(e),
                })

        # ── Step 3: Execute INSERTs (v2.0) ──
        if new_rows_data:
            # 获取所有导出字段（用于 INSERT 列列表）
            all_export_fields = [f for f in fields if f.include_in_export]
            insert_col_names = [f.field_name for f in all_export_fields]
            insert_cols_sql = ", ".join(_quote_col(ds.db_type, c) for c in insert_col_names)
            placeholders = ", ".join([ph] * len(insert_col_names))

            for nr in new_rows_data:
                row_data = nr["data"]
                pk_key = nr["pk_key"]
                vals = []
                for fn in insert_col_names:
                    v = row_data.get(fn)
                    if v is not None and v != "":
                        vals.append(v)
                    else:
                        vals.append(None)

                insert_sql = f"INSERT INTO {qt} ({insert_cols_sql}) VALUES ({placeholders})"
                try:
                    _exec(cur, ds.db_type, insert_sql, vals)
                    success_count += 1
                    insert_count += 1

                    # Record field-level changes for INSERT
                    for fn, val in zip(insert_col_names, vals):
                        if val is not None:
                            change_logs.append({
                                "row_pk_value": pk_key,
                                "field_name": fn,
                                "old_value": None,
                                "new_value": str(val),
                                "change_type": "insert",
                            })
                except Exception as e:
                    fail_count += 1
                    failed_details.append({
                        "row_num": nr["row_num"],
                        "pk_key": pk_key,
                        "error": str(e),
                    })

        # ── Step 4: Execute DELETEs (v3.10) ──
        if delete_rows_data:
            all_export_fields = [f for f in fields if f.include_in_export]
            all_field_names_del = [f.field_name for f in all_export_fields]
            cols_sql_del = ", ".join(_quote_col(ds.db_type, c) for c in all_field_names_del)

            for dr in delete_rows_data:
                pk_key = dr["pk_key"]
                pk_vals = pk_key.split("|")

                # Build WHERE from PK
                where_parts_del = []
                where_params_del = []
                for i, pkf in enumerate(pk_fields_list):
                    pv = pk_vals[i] if i < len(pk_vals) else ""
                    where_parts_del.append(f"{_cast_to_text(ds.db_type, _quote_col(ds.db_type, pkf))} = {ph}")
                    where_params_del.append(pv)

                where_sql_del = " AND ".join(where_parts_del)

                try:
                    # Read row before delete for change log
                    _exec(cur, ds.db_type, f"SELECT {cols_sql_del} FROM {qt} WHERE {where_sql_del}", where_params_del)
                    row_before = cur.fetchone()

                    # Execute DELETE
                    _exec(cur, ds.db_type, f"DELETE FROM {qt} WHERE {where_sql_del}", where_params_del)
                    success_count += 1
                    delete_count += 1

                    # Record field-level change logs
                    if row_before:
                        for j, fn in enumerate(all_field_names_del):
                            old_val = str(row_before[j]) if row_before[j] is not None else None
                            if old_val is not None:
                                change_logs.append({
                                    "row_pk_value": pk_key,
                                    "field_name": fn,
                                    "old_value": old_val,
                                    "new_value": None,
                                    "change_type": "delete",
                                })
                except Exception as e:
                    fail_count += 1
                    failed_details.append({
                        "row_num": dr.get("row_num", 0),
                        "pk_key": pk_key,
                        "error": str(e),
                    })

        conn.commit()

        finished_at = _now_bjt()
        wb_status = "success" if fail_count == 0 else ("failed" if success_count == 0 else "partial")

        # Record writeback log
        wb_log = WritebackLog(
            writeback_batch_no=wb_batch,
            import_task_id=task_id,
            table_config_id=tc.id,
            datasource_id=tc.datasource_id,
            backup_version_no=bk_batch,
            total_row_count=len(update_rows) + len(new_rows_data) + len(delete_rows_data),
            success_row_count=success_count,
            failed_row_count=fail_count,
            skipped_row_count=0,
            inserted_row_count=insert_count,
            updated_row_count=update_count,
            deleted_row_count=delete_count,
            writeback_status=wb_status,
            writeback_message=f"更新 {update_count}，新增 {insert_count}，删除 {delete_count}，失败 {fail_count}",
            failed_detail_json=json.dumps(failed_details, ensure_ascii=False) if failed_details else None,
            operator_user=_get_username(user),
            started_at=started_at,
            finished_at=finished_at,
        )
        db.add(wb_log)
        db.flush()

        # ── Step 4: Record field-level change logs (v2.0 Task 3) ──
        for cl in change_logs:
            db.add(FieldChangeLog(
                writeback_log_id=wb_log.id,
                row_pk_value=cl["row_pk_value"],
                field_name=cl["field_name"],
                old_value=cl["old_value"],
                new_value=cl["new_value"],
                change_type=cl["change_type"],
            ))

        log_operation(db, "数据维护", "执行回写", wb_status,
                      target_id=tc.id, target_name=tc.table_name,
                      message=f"回写 {wb_batch}，更新 {update_count}，新增 {insert_count}，删除 {delete_count}，失败 {fail_count}，备份 {bk_batch}",
                      operator=_get_username(user))

        # v2.3: Notification
        from app.utils.notifications import notify_user_by_username
        ntype = "success" if wb_status == "success" else ("error" if wb_status == "failed" else "warning")
        notify_user_by_username(
            db, _get_username(user),
            "回写%s" % ("成功" if wb_status == "success" else "完成"),
            "表「%s」回写 %s，更新 %d，新增 %d，删除 %d，失败 %d" % (
                tc.table_alias or tc.table_name, wb_batch, update_count, insert_count, delete_count, fail_count),
            ntype=ntype,
            related_url="/log-center",
        )

        # Update import task status
        task.import_status = "confirmed"
        task.updated_at = _now_bjt()

        db.commit()

        return {
            "writeback_batch_no": wb_batch,
            "backup_version_no": bk_batch,
            "status": wb_status,
            "total": len(update_rows) + len(new_rows_data) + len(delete_rows_data),
            "success": success_count,
            "failed": fail_count,
            "updated": update_count,
            "inserted": insert_count,
            "deleted": delete_count,
            "backup_table": backup_table_name,
            "backup_record_count": backup_count,
            "operator_user": _get_username(user),
            "started_at": started_at.isoformat(),
            "finished_at": finished_at.isoformat(),
            "failed_details": failed_details[:50],
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, t("data_maintenance.writeback_failed", error=str(e)))
    finally:
        conn.close()


# ─────────────────────────────────────────────
# v2.1: 在线编辑模式 — 行内编辑 + 单行新增
# ─────────────────────────────────────────────

class InlineChange(BaseModel):
    pk_values: Dict[str, str]
    updates: Dict[str, Optional[str]]


class InlineUpdateRequest(BaseModel):
    changes: List[InlineChange]


class InlineInsertRequest(BaseModel):
    row_data: Dict[str, Optional[str]]


@router.put("/{table_config_id}/inline-update")
def inline_update(
    table_config_id: int,
    body: InlineUpdateRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin", "operator")),
):
    """行内编辑：写前备份 → 逐行 UPDATE → 记录 writeback_log + field_change_log。"""
    # v2.2: approval workflow
    if _needs_approval(db, user):
        changes_data = [{"pk_values": c.pk_values, "updates": c.updates} for c in body.changes]
        return _create_approval_request(
            db, user,
            table_config_id=table_config_id,
            request_type="inline_update",
            request_data_json=json.dumps({"changes": changes_data}, ensure_ascii=False),
        )

    tc = _get_tc(db, table_config_id)
    ds = _get_ds(db, tc.datasource_id)
    pwd = decrypt_password(ds.password_encrypted)
    fields = _get_fields(db, tc.id)
    field_name_map = {f.field_name: f for f in fields}
    pk_fields_list = [p.strip() for p in tc.primary_key_fields.split(",")]

    if not body.changes:
        raise HTTPException(400, t("data_maintenance.no_changes"))

    wb_batch = _gen_batch("INL")
    bk_batch = _gen_batch("BK")
    started_at = _now_bjt()

    conn = _connect(
        ds.db_type, ds.host, ds.port, ds.username, pwd,
        tc.db_name, tc.schema_name, ds.charset, ds.connect_timeout_seconds or 10,
    )
    try:
        cur = conn.cursor()
        qt = _qualified_table(ds.db_type, tc.table_name, tc.schema_name)
        ph = _placeholder(ds.db_type)

        # ── Step 1: Full table backup ──
        ts = _now_bjt().strftime("%Y%m%d_%H%M%S")
        rand_suffix = uuid.uuid4().hex[:4].upper()
        backup_table_name = f"{tc.table_name}_bak_{ts}_{rand_suffix}"
        _create_backup_table(cur, ds.db_type, qt, backup_table_name, tc.schema_name)

        bk_qt = _qualified_table(ds.db_type, backup_table_name, tc.schema_name)
        cur.execute(f"SELECT COUNT(*) FROM {bk_qt}")
        backup_count = cur.fetchone()[0]
        conn.commit()

        backup_rec = TableBackupVersion(
            backup_version_no=bk_batch,
            table_config_id=tc.id,
            datasource_id=tc.datasource_id,
            backup_table_name=backup_table_name,
            source_table_name=tc.table_name,
            source_db_name=tc.db_name,
            source_schema_name=tc.schema_name,
            trigger_type="triggered_by_inline_update",
            related_writeback_batch_no=wb_batch,
            record_count=backup_count,
            storage_status="valid",
            can_rollback=1,
            backup_started_at=started_at,
            backup_finished_at=_now_bjt(),
            operator_user=_get_username(user),
        )
        db.add(backup_rec)
        db.flush()

        # ── Step 2: Read old values & execute UPDATEs ──
        success_count = 0
        fail_count = 0
        failed_details = []  # type: List[dict]
        change_logs = []  # type: List[dict]

        for change in body.changes:
            pk_values = change.pk_values
            updates = change.updates

            # Filter to only editable fields (v2.4: respect editable_roles)
            user_role = user.role if user else "readonly"
            valid_updates = {}  # type: Dict[str, Optional[str]]
            for fn, new_val in updates.items():
                fc = field_name_map.get(fn)
                if fc and fc.is_editable and not fc.is_primary_key and not fc.is_system_field:
                    # v2.4: check editable_roles
                    if fc.editable_roles:
                        allowed_roles = [r.strip() for r in fc.editable_roles.split(",") if r.strip()]
                        if allowed_roles and user_role not in allowed_roles:
                            continue
                    valid_updates[fn] = new_val

            if not valid_updates:
                continue

            # Build WHERE from pk_values
            where_parts = []  # type: List[str]
            where_params = []  # type: list
            for pkf in pk_fields_list:
                pv = pk_values.get(pkf, "")
                where_parts.append(f"{_cast_to_text(ds.db_type, _quote_col(ds.db_type, pkf))} = %s")
                where_params.append(pv)

            where_sql = " AND ".join(where_parts)
            pk_key = "|".join(pk_values.get(pkf, "") for pkf in pk_fields_list)

            # Fetch old values
            old_col_names = list(valid_updates.keys())
            old_cols_sql = ", ".join(_quote_col(ds.db_type, c) for c in old_col_names)
            try:
                _exec(cur, ds.db_type, f"SELECT {old_cols_sql} FROM {qt} WHERE {where_sql}", where_params)
                old_row = cur.fetchone()
            except Exception:
                old_row = None

            # Build SET clause
            set_parts = []  # type: List[str]
            set_params = []  # type: list
            for fn, new_val in valid_updates.items():
                set_parts.append(f"{_quote_col(ds.db_type, fn)} = %s")
                set_params.append(new_val)

            update_sql = f"UPDATE {qt} SET {', '.join(set_parts)} WHERE {where_sql}"
            try:
                _exec(cur, ds.db_type, update_sql, set_params + where_params)
                success_count += 1

                # Record field-level changes
                for idx_col, fn in enumerate(old_col_names):
                    old_val = str(old_row[idx_col]) if old_row and old_row[idx_col] is not None else None
                    new_val = valid_updates[fn]
                    if old_val != new_val:
                        change_logs.append({
                            "row_pk_value": pk_key,
                            "field_name": fn,
                            "old_value": old_val,
                            "new_value": new_val,
                            "change_type": "update",
                        })
            except Exception as e:
                fail_count += 1
                failed_details.append({"pk_key": pk_key, "error": str(e)})

        conn.commit()
        finished_at = _now_bjt()
        wb_status = "success" if fail_count == 0 else ("failed" if success_count == 0 else "partial")

        # Record writeback log
        wb_log = WritebackLog(
            writeback_batch_no=wb_batch,
            import_task_id=0,
            table_config_id=tc.id,
            datasource_id=tc.datasource_id,
            backup_version_no=bk_batch,
            total_row_count=len(body.changes),
            success_row_count=success_count,
            failed_row_count=fail_count,
            skipped_row_count=0,
            inserted_row_count=0,
            updated_row_count=success_count,
            deleted_row_count=0,
            writeback_status=wb_status,
            writeback_message=f"在线编辑更新 {success_count} 行，失败 {fail_count} 行",
            failed_detail_json=json.dumps(failed_details, ensure_ascii=False) if failed_details else None,
            operator_user=_get_username(user),
            started_at=started_at,
            finished_at=finished_at,
        )
        db.add(wb_log)
        db.flush()

        for cl in change_logs:
            db.add(FieldChangeLog(
                writeback_log_id=wb_log.id,
                row_pk_value=cl["row_pk_value"],
                field_name=cl["field_name"],
                old_value=cl["old_value"],
                new_value=cl["new_value"],
                change_type=cl["change_type"],
            ))

        log_operation(db, "数据维护", "在线编辑", wb_status,
                      target_id=tc.id, target_name=tc.table_name,
                      message=f"在线编辑 {wb_batch}，更新 {success_count}，失败 {fail_count}，备份 {bk_batch}",
                      operator=_get_username(user))
        db.commit()

        try:
            from app.utils.notifications import notify_user_by_username
            notify_user_by_username(
                db, _get_username(user),
                title="在线编辑完成",
                message_text=f"在线编辑【{tc.table_alias or tc.table_name}】完成，更新 {success_count} 行" + (f"，失败 {fail_count} 行" if fail_count else ""),
                notif_type="success" if fail_count == 0 else "warning",
            )
        except Exception:
            pass

        return {
            "writeback_batch_no": wb_batch,
            "backup_version_no": bk_batch,
            "status": wb_status,
            "total": len(body.changes),
            "success": success_count,
            "failed": fail_count,
            "updated": success_count,
            "backup_table": backup_table_name,
            "backup_record_count": backup_count,
            "change_count": len(change_logs),
            "failed_details": failed_details[:50],
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, t("data_maintenance.online_edit_failed", error=str(e)))
    finally:
        conn.close()


@router.post("/{table_config_id}/inline-insert")
def inline_insert(
    table_config_id: int,
    body: InlineInsertRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin", "operator")),
):
    """单行新增：写前备份 → INSERT → 记录日志。"""
    # v2.2: approval workflow
    if _needs_approval(db, user):
        return _create_approval_request(
            db, user,
            table_config_id=table_config_id,
            request_type="inline_insert",
            request_data_json=json.dumps({"row_data": body.row_data}, ensure_ascii=False),
        )

    tc = _get_tc(db, table_config_id)
    if not tc.allow_insert_rows:
        raise HTTPException(403, t("data_maintenance.insert_not_enabled"))

    ds = _get_ds(db, tc.datasource_id)
    pwd = decrypt_password(ds.password_encrypted)
    fields = _get_fields(db, tc.id)
    pk_fields_list = [p.strip() for p in tc.primary_key_fields.split(",")]

    if not body.row_data:
        raise HTTPException(400, t("data_maintenance.no_data"))

    # Validate PK fields are provided
    for pkf in pk_fields_list:
        pv = body.row_data.get(pkf)
        if not pv or str(pv).strip() == "":
            raise HTTPException(400, t("data_maintenance.pk_field_required", field=pkf))

    wb_batch = _gen_batch("INS")
    bk_batch = _gen_batch("BK")
    started_at = _now_bjt()

    conn = _connect(
        ds.db_type, ds.host, ds.port, ds.username, pwd,
        tc.db_name, tc.schema_name, ds.charset, ds.connect_timeout_seconds or 10,
    )
    try:
        cur = conn.cursor()
        qt = _qualified_table(ds.db_type, tc.table_name, tc.schema_name)

        # Backup
        ts = _now_bjt().strftime("%Y%m%d_%H%M%S")
        rand_suffix = uuid.uuid4().hex[:4].upper()
        backup_table_name = f"{tc.table_name}_bak_{ts}_{rand_suffix}"
        _create_backup_table(cur, ds.db_type, qt, backup_table_name, tc.schema_name)

        bk_qt = _qualified_table(ds.db_type, backup_table_name, tc.schema_name)
        cur.execute(f"SELECT COUNT(*) FROM {bk_qt}")
        backup_count = cur.fetchone()[0]
        conn.commit()

        backup_rec = TableBackupVersion(
            backup_version_no=bk_batch,
            table_config_id=tc.id,
            datasource_id=tc.datasource_id,
            backup_table_name=backup_table_name,
            source_table_name=tc.table_name,
            source_db_name=tc.db_name,
            source_schema_name=tc.schema_name,
            trigger_type="triggered_by_inline_insert",
            related_writeback_batch_no=wb_batch,
            record_count=backup_count,
            storage_status="valid",
            can_rollback=1,
            backup_started_at=started_at,
            backup_finished_at=_now_bjt(),
            operator_user=_get_username(user),
        )
        db.add(backup_rec)
        db.flush()

        # Build INSERT
        export_fields = [f for f in fields if f.include_in_export]
        insert_col_names = [f.field_name for f in export_fields]
        insert_cols_sql = ", ".join(_quote_col(ds.db_type, c) for c in insert_col_names)
        placeholders = ", ".join(["%s"] * len(insert_col_names))

        vals = []  # type: list
        for fn in insert_col_names:
            v = body.row_data.get(fn)
            if v is not None and str(v).strip() != "":
                vals.append(str(v).strip())
            else:
                vals.append(None)

        pk_key = "|".join(body.row_data.get(pkf, "") for pkf in pk_fields_list)

        insert_sql = f"INSERT INTO {qt} ({insert_cols_sql}) VALUES ({placeholders})"
        try:
            _exec(cur, ds.db_type, insert_sql, vals)
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise HTTPException(500, t("data_maintenance.insert_row_failed", error=str(e)))

        finished_at = _now_bjt()

        # Record writeback log
        wb_log = WritebackLog(
            writeback_batch_no=wb_batch,
            import_task_id=0,
            table_config_id=tc.id,
            datasource_id=tc.datasource_id,
            backup_version_no=bk_batch,
            total_row_count=1,
            success_row_count=1,
            failed_row_count=0,
            skipped_row_count=0,
            inserted_row_count=1,
            updated_row_count=0,
            deleted_row_count=0,
            writeback_status="success",
            writeback_message=f"在线新增 1 行",
            operator_user=_get_username(user),
            started_at=started_at,
            finished_at=finished_at,
        )
        db.add(wb_log)
        db.flush()

        # Record field-level change logs
        for fn, val in zip(insert_col_names, vals):
            if val is not None:
                db.add(FieldChangeLog(
                    writeback_log_id=wb_log.id,
                    row_pk_value=pk_key,
                    field_name=fn,
                    old_value=None,
                    new_value=str(val),
                    change_type="insert",
                ))

        log_operation(db, "数据维护", "在线新增行", "success",
                      target_id=tc.id, target_name=tc.table_name,
                      message=f"在线新增行 {wb_batch}，备份 {bk_batch}",
                      operator=_get_username(user))
        db.commit()

        try:
            from app.utils.notifications import notify_user_by_username
            notify_user_by_username(
                db, _get_username(user),
                title="新增行完成",
                message_text=f"在线新增【{tc.table_alias or tc.table_name}】1 行数据",
                notif_type="success",
            )
        except Exception:
            pass

        return {
            "writeback_batch_no": wb_batch,
            "backup_version_no": bk_batch,
            "status": "success",
            "pk_key": pk_key,
            "backup_table": backup_table_name,
            "backup_record_count": backup_count,
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, t("data_maintenance.insert_row_failed", error=str(e)))
    finally:
        conn.close()


# ─────────────────────────────────────────────
# v2.0 Task 2: 数据浏览页面直接删除行
# ─────────────────────────────────────────────

class DeleteRowsRequest(BaseModel):
    pk_values: List[str]  # 主键值列表，复合主键用 | 分隔


@router.delete("/{table_config_id}/rows")
def delete_rows(
    table_config_id: int,
    body: DeleteRowsRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin", "operator")),
):
    """按主键批量删除数据行（写前备份 + 逐字段变更日志）。"""
    # v2.2: approval workflow
    if _needs_approval(db, user):
        return _create_approval_request(
            db, user,
            table_config_id=table_config_id,
            request_type="delete",
            request_data_json=json.dumps({"pk_values": body.pk_values}, ensure_ascii=False),
        )

    tc = _get_tc(db, table_config_id)
    if not tc.allow_delete_rows:
        raise HTTPException(403, t("data_maintenance.delete_not_enabled"))

    ds = _get_ds(db, tc.datasource_id)
    pwd = decrypt_password(ds.password_encrypted)
    fields = _get_fields(db, tc.id)
    pk_fields_list = [p.strip() for p in tc.primary_key_fields.split(",")]
    export_fields = [f for f in fields if f.include_in_export]
    all_field_names = [f.field_name for f in export_fields]

    if not body.pk_values:
        raise HTTPException(400, t("data_maintenance.no_rows_selected"))

    conn = _connect(
        ds.db_type, ds.host, ds.port, ds.username, pwd,
        tc.db_name, tc.schema_name, ds.charset, ds.connect_timeout_seconds or 10,
    )
    try:
        cur = conn.cursor()
        qt = _qualified_table(ds.db_type, tc.table_name, tc.schema_name)
        ph = _placeholder(ds.db_type)

        # ── Step 1: Backup ──
        wb_batch = _gen_batch("DEL")
        bk_batch = _gen_batch("BK")
        started_at = _now_bjt()

        ts = _now_bjt().strftime("%Y%m%d_%H%M%S")
        rand_suffix = uuid.uuid4().hex[:4].upper()
        backup_table_name = f"{tc.table_name}_bak_{ts}_{rand_suffix}"

        _create_backup_table(cur, ds.db_type, qt, backup_table_name, tc.schema_name)

        bk_qt = _qualified_table(ds.db_type, backup_table_name, tc.schema_name)
        cur.execute(f"SELECT COUNT(*) FROM {bk_qt}")
        backup_count = cur.fetchone()[0]
        conn.commit()

        backup_rec = TableBackupVersion(
            backup_version_no=bk_batch,
            table_config_id=tc.id,
            datasource_id=tc.datasource_id,
            backup_table_name=backup_table_name,
            source_table_name=tc.table_name,
            source_db_name=tc.db_name,
            source_schema_name=tc.schema_name,
            trigger_type="triggered_by_delete",
            related_writeback_batch_no=wb_batch,
            record_count=backup_count,
            storage_status="valid",
            can_rollback=1,
            backup_started_at=started_at,
            backup_finished_at=_now_bjt(),
            operator_user=_get_username(user),
        )
        db.add(backup_rec)
        db.flush()

        # ── Step 2: Fetch rows to be deleted (for change log) ──
        cols_sql = ", ".join(_quote_col(ds.db_type, c) for c in all_field_names)

        success_count = 0
        fail_count = 0
        failed_details: List[dict] = []
        change_logs: List[dict] = []

        for pk_val_str in body.pk_values:
            pk_vals = pk_val_str.split("|")

            # Fetch the row first for change log
            where_parts = []
            where_params = []
            for i, pkf in enumerate(pk_fields_list):
                pv = pk_vals[i] if i < len(pk_vals) else ""
                where_parts.append(f"{_cast_to_text(ds.db_type, _quote_col(ds.db_type, pkf))} = {ph}")
                where_params.append(pv)

            where_sql = " AND ".join(where_parts)

            try:
                # Read row before delete
                _exec(cur, ds.db_type, f"SELECT {cols_sql} FROM {qt} WHERE {where_sql}", where_params)
                row_before = cur.fetchone()

                # DELETE
                _exec(cur, ds.db_type, f"DELETE FROM {qt} WHERE {where_sql}", where_params)
                success_count += 1

                # Record change log
                if row_before:
                    for j, fn in enumerate(all_field_names):
                        old_val = str(row_before[j]) if row_before[j] is not None else None
                        if old_val is not None:
                            change_logs.append({
                                "row_pk_value": pk_val_str,
                                "field_name": fn,
                                "old_value": old_val,
                                "new_value": None,
                                "change_type": "delete",
                            })
            except Exception as e:
                fail_count += 1
                failed_details.append({"pk_key": pk_val_str, "error": str(e)})

        conn.commit()
        finished_at = _now_bjt()
        wb_status = "success" if fail_count == 0 else ("failed" if success_count == 0 else "partial")

        # Record writeback log (reuse WritebackLog for delete operations)
        wb_log = WritebackLog(
            writeback_batch_no=wb_batch,
            import_task_id=0,  # 无关联导入任务
            table_config_id=tc.id,
            datasource_id=tc.datasource_id,
            backup_version_no=bk_batch,
            total_row_count=len(body.pk_values),
            success_row_count=success_count,
            failed_row_count=fail_count,
            skipped_row_count=0,
            inserted_row_count=0,
            updated_row_count=0,
            deleted_row_count=success_count,
            writeback_status=wb_status,
            writeback_message=f"删除 {success_count} 行，失败 {fail_count} 行",
            failed_detail_json=json.dumps(failed_details, ensure_ascii=False) if failed_details else None,
            operator_user=_get_username(user),
            started_at=started_at,
            finished_at=finished_at,
        )
        db.add(wb_log)
        db.flush()

        # Record field-level change logs
        for cl in change_logs:
            db.add(FieldChangeLog(
                writeback_log_id=wb_log.id,
                row_pk_value=cl["row_pk_value"],
                field_name=cl["field_name"],
                old_value=cl["old_value"],
                new_value=cl["new_value"],
                change_type=cl["change_type"],
            ))

        log_operation(db, "数据维护", "删除数据行", wb_status,
                      target_id=tc.id, target_name=tc.table_name,
                      message=f"删除 {success_count} 行，失败 {fail_count}，备份 {bk_batch}",
                      operator=_get_username(user))
        db.commit()

        try:
            from app.utils.notifications import notify_user_by_username
            notify_user_by_username(
                db, _get_username(user),
                title="删除行完成",
                message_text=f"删除【{tc.table_alias or tc.table_name}】{success_count} 行数据" + (f"，失败 {fail_count} 行" if fail_count else ""),
                notif_type="success" if fail_count == 0 else "warning",
            )
        except Exception:
            pass

        return {
            "status": wb_status,
            "deleted": success_count,
            "failed": fail_count,
            "backup_version_no": bk_batch,
            "backup_table": backup_table_name,
            "failed_details": failed_details[:50],
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, t("data_maintenance.delete_failed", error=str(e)))
    finally:
        conn.close()


# ─────────────────────────────────────────────
# v2.1.2: 批量新增行（弹窗批量粘贴）
# ─────────────────────────────────────────────

class BatchInsertRequest(BaseModel):
    rows: List[Dict[str, Optional[str]]]


@router.post("/{table_config_id}/batch-insert")
def batch_insert(
    table_config_id: int,
    body: BatchInsertRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin", "operator")),
):
    """批量新增行：写前备份 → 逐行 INSERT → 记录日志。"""
    # v2.2: approval workflow
    if _needs_approval(db, user):
        return _create_approval_request(
            db, user,
            table_config_id=table_config_id,
            request_type="batch_insert",
            request_data_json=json.dumps({"rows": body.rows}, ensure_ascii=False),
        )

    tc = _get_tc(db, table_config_id)
    if not tc.allow_insert_rows:
        raise HTTPException(403, t("data_maintenance.insert_not_enabled"))

    ds = _get_ds(db, tc.datasource_id)
    pwd = decrypt_password(ds.password_encrypted)
    fields = _get_fields(db, tc.id)
    pk_fields_list = [p.strip() for p in tc.primary_key_fields.split(",")]

    if not body.rows:
        raise HTTPException(400, t("data_maintenance.no_data"))

    # Filter out completely empty rows
    valid_rows = []
    for row_data in body.rows:
        if any(v is not None and str(v).strip() != "" for v in row_data.values()):
            valid_rows.append(row_data)

    if not valid_rows:
        raise HTTPException(400, t("data_maintenance.all_rows_empty"))

    # Validate PK fields for each row
    for idx, row_data in enumerate(valid_rows):
        for pkf in pk_fields_list:
            pv = row_data.get(pkf)
            if not pv or str(pv).strip() == "":
                fc = next((f for f in fields if f.field_name == pkf), None)
                pk_alias = fc.field_alias if fc and fc.field_alias else pkf
                raise HTTPException(400, t("data_maintenance.batch_pk_required", row=idx + 1, field=pk_alias))

    wb_batch = _gen_batch("BINS")
    bk_batch = _gen_batch("BK")
    started_at = _now_bjt()

    conn = _connect(
        ds.db_type, ds.host, ds.port, ds.username, pwd,
        tc.db_name, tc.schema_name, ds.charset, ds.connect_timeout_seconds or 10,
    )
    try:
        cur = conn.cursor()
        qt = _qualified_table(ds.db_type, tc.table_name, tc.schema_name)

        # ── Step 1: Backup ──
        ts = _now_bjt().strftime("%Y%m%d_%H%M%S")
        rand_suffix = uuid.uuid4().hex[:4].upper()
        backup_table_name = "%s_bak_%s_%s" % (tc.table_name, ts, rand_suffix)
        _create_backup_table(cur, ds.db_type, qt, backup_table_name, tc.schema_name)

        bk_qt = _qualified_table(ds.db_type, backup_table_name, tc.schema_name)
        cur.execute("SELECT COUNT(*) FROM %s" % bk_qt)
        backup_count = cur.fetchone()[0]
        conn.commit()

        backup_rec = TableBackupVersion(
            backup_version_no=bk_batch,
            table_config_id=tc.id,
            datasource_id=tc.datasource_id,
            backup_table_name=backup_table_name,
            source_table_name=tc.table_name,
            source_db_name=tc.db_name,
            source_schema_name=tc.schema_name,
            trigger_type="triggered_by_batch_insert",
            related_writeback_batch_no=wb_batch,
            record_count=backup_count,
            storage_status="valid",
            can_rollback=1,
            backup_started_at=started_at,
            backup_finished_at=_now_bjt(),
            operator_user=_get_username(user),
        )
        db.add(backup_rec)
        db.flush()

        # ── Step 2: Execute INSERTs ──
        export_fields = [f for f in fields if f.include_in_export]
        insert_col_names = [f.field_name for f in export_fields]
        insert_cols_sql = ", ".join(_quote_col(ds.db_type, c) for c in insert_col_names)
        placeholders = ", ".join(["%s"] * len(insert_col_names))
        insert_sql = "INSERT INTO %s (%s) VALUES (%s)" % (qt, insert_cols_sql, placeholders)

        success_count = 0
        fail_count = 0
        failed_details = []  # type: List[dict]
        change_logs = []  # type: List[dict]

        for row_idx, row_data in enumerate(valid_rows):
            vals = []  # type: list
            for fn in insert_col_names:
                v = row_data.get(fn)
                if v is not None and str(v).strip() != "":
                    vals.append(str(v).strip())
                else:
                    vals.append(None)

            pk_key = "|".join(str(row_data.get(pkf, "") or "") for pkf in pk_fields_list)

            try:
                _exec(cur, ds.db_type, insert_sql, vals)
                success_count += 1

                for fn, val in zip(insert_col_names, vals):
                    if val is not None:
                        change_logs.append({
                            "row_pk_value": pk_key,
                            "field_name": fn,
                            "old_value": None,
                            "new_value": str(val),
                            "change_type": "insert",
                        })
            except Exception as e:
                fail_count += 1
                failed_details.append({
                    "row_num": row_idx + 1,
                    "pk_key": pk_key,
                    "error": str(e),
                })

        conn.commit()
        finished_at = _now_bjt()
        wb_status = "success" if fail_count == 0 else ("failed" if success_count == 0 else "partial")

        # Record writeback log
        wb_log = WritebackLog(
            writeback_batch_no=wb_batch,
            import_task_id=0,
            table_config_id=tc.id,
            datasource_id=tc.datasource_id,
            backup_version_no=bk_batch,
            total_row_count=len(valid_rows),
            success_row_count=success_count,
            failed_row_count=fail_count,
            skipped_row_count=0,
            inserted_row_count=success_count,
            updated_row_count=0,
            deleted_row_count=0,
            writeback_status=wb_status,
            writeback_message="批量新增 %d 行，失败 %d 行" % (success_count, fail_count),
            failed_detail_json=json.dumps(failed_details, ensure_ascii=False) if failed_details else None,
            operator_user=_get_username(user),
            started_at=started_at,
            finished_at=finished_at,
        )
        db.add(wb_log)
        db.flush()

        for cl in change_logs:
            db.add(FieldChangeLog(
                writeback_log_id=wb_log.id,
                row_pk_value=cl["row_pk_value"],
                field_name=cl["field_name"],
                old_value=cl["old_value"],
                new_value=cl["new_value"],
                change_type=cl["change_type"],
            ))

        log_operation(db, "数据维护", "批量新增行", wb_status,
                      target_id=tc.id, target_name=tc.table_name,
                      message="批量新增 %s，成功 %d，失败 %d，备份 %s" % (wb_batch, success_count, fail_count, bk_batch),
                      operator=_get_username(user))
        db.commit()

        try:
            from app.utils.notifications import notify_user_by_username
            notify_user_by_username(
                db, _get_username(user),
                title="批量新增完成",
                message_text=f"批量新增【{tc.table_alias or tc.table_name}】{success_count} 行" + (f"，失败 {fail_count} 行" if fail_count else ""),
                notif_type="success" if fail_count == 0 else "warning",
            )
        except Exception:
            pass

        return {
            "writeback_batch_no": wb_batch,
            "backup_version_no": bk_batch,
            "status": wb_status,
            "total": len(valid_rows),
            "success": success_count,
            "failed": fail_count,
            "backup_table": backup_table_name,
            "backup_record_count": backup_count,
            "failed_details": failed_details[:50],
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, t("data_maintenance.batch_insert_failed", error=str(e)))
    finally:
        conn.close()


# ─────────────────────────────────────────────
# v2.3: 异步导出（大表）
# ─────────────────────────────────────────────

def _run_async_export(task_id: str, table_config_id: int, export_type: str,
                      keyword: Optional[str], field_filters: Optional[str],
                      operator_user: str):
    """Background thread: execute export and update task status."""
    import openpyxl
    from openpyxl.utils import get_column_letter
    from app.database import SessionLocal

    db = SessionLocal()
    try:
        task = db.query(ExportTask).filter(ExportTask.task_id == task_id).first()
        if not task:
            return

        tc = db.query(TableConfig).filter(TableConfig.id == table_config_id).first()
        ds = db.query(DatasourceConfig).filter(DatasourceConfig.id == tc.datasource_id).first()
        pwd = decrypt_password(ds.password_encrypted)
        fields = db.query(FieldConfig).filter(
            FieldConfig.table_config_id == table_config_id,
            FieldConfig.is_deleted == 0,
        ).order_by(FieldConfig.field_order_no).all()
        export_fields = [f for f in fields if f.include_in_export]
        if not export_fields:
            task.status = "failed"
            task.error_message = "没有可导出的字段"
            task.finished_at = _now_bjt()
            db.commit()
            return

        col_names = [f.field_name for f in export_fields]
        pk_set = set(f.strip() for f in tc.primary_key_fields.split(","))
        ph = _placeholder(ds.db_type)

        conn = _connect(
            ds.db_type, ds.host, ds.port, ds.username, pwd,
            tc.db_name, tc.schema_name, ds.charset, ds.connect_timeout_seconds or 10,
        )
        try:
            cur = conn.cursor()
            qt = _qualified_table(ds.db_type, tc.table_name, tc.schema_name)

            where_parts = []  # type: List[str]
            params = []  # type: list
            if export_type == "current" and (keyword or field_filters):
                if keyword:
                    kw_parts = []
                    for f in export_fields:
                        kw_parts.append(
                            "%s LIKE %s" % (_cast_to_text(ds.db_type, _quote_col(ds.db_type, f.field_name)), ph)
                        )
                        params.append("%%%s%%" % keyword)
                    where_parts.append("(%s)" % " OR ".join(kw_parts))
                if field_filters:
                    try:
                        ff = json.loads(field_filters)
                        for fname, fval in ff.items():
                            if fval and any(f.field_name == fname for f in export_fields):
                                where_parts.append(
                                    "%s LIKE %s" % (_cast_to_text(ds.db_type, _quote_col(ds.db_type, fname)), ph)
                                )
                                params.append("%%%s%%" % fval)
                    except (json.JSONDecodeError, TypeError):
                        pass

            where_sql = ""
            if where_parts:
                where_sql = " WHERE " + " AND ".join(where_parts)

            cols_sql = ", ".join(_quote_col(ds.db_type, c) for c in col_names)
            data_sql = "SELECT %s FROM %s%s" % (cols_sql, qt, where_sql)
            _exec(cur, ds.db_type, data_sql, params)
            raw_rows = cur.fetchall()
        except Exception as e:
            task.status = "failed"
            task.error_message = "查询数据失败: %s" % str(e)
            task.finished_at = _now_bjt()
            db.commit()
            return
        finally:
            conn.close()

        # Generate Excel
        from openpyxl.styles import Protection as CellProtection, PatternFill, Font as XlFont
        from openpyxl.worksheet.protection import SheetProtection

        # v3.5: protection password (internal)
        _SHEET_PROTECTION_PASSWORD = "DOW_tpl_v35_sec"

        batch_no = _gen_batch("EXP")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "数据"
        locked_cell = CellProtection(locked=True)
        unlocked_cell = CellProtection(locked=False)
        editable_field_names = set(f.field_name for f in export_fields if f.is_editable)
        data_row_count = len(raw_rows)
        RESERVED_BLANK_ROWS = 50

        # v3.5: Visual style fills
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = XlFont(bold=True, color="FFFFFF", size=11)
        readonly_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        editable_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        blank_zone_fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")

        for i, f in enumerate(export_fields, 1):
            cell = ws.cell(row=1, column=i, value=f.field_alias or f.field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.protection = locked_cell

        for row_idx, raw in enumerate(raw_rows, 2):
            for col_idx, (val, ef) in enumerate(zip(raw, export_fields), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
                if ef.field_name in pk_set:
                    cell.protection = locked_cell
                    cell.fill = readonly_fill
                elif ef.field_name in editable_field_names:
                    cell.protection = unlocked_cell
                    cell.fill = editable_fill
                else:
                    cell.protection = locked_cell
                    cell.fill = readonly_fill

        blank_start = 2 + data_row_count
        for row_idx in range(blank_start, blank_start + RESERVED_BLANK_ROWS):
            for col_idx, ef in enumerate(export_fields, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value="")
                cell.fill = blank_zone_fill
                if ef.field_name in pk_set:
                    cell.protection = unlocked_cell
                elif ef.field_name in editable_field_names:
                    cell.protection = unlocked_cell
                else:
                    cell.protection = locked_cell

        ws.protection = SheetProtection(
            sheet=True, password=_SHEET_PROTECTION_PASSWORD,
            formatColumns=False, formatRows=False, formatCells=False,
            insertRows=False, deleteRows=True, deleteColumns=True, insertColumns=True,
            sort=False, autoFilter=False,
        )

        meta_ws = wb.create_sheet("_meta")
        meta_info = {
            "datasource_id": tc.datasource_id,
            "table_config_id": tc.id,
            "config_version": tc.config_version,
            "export_time": _now_bjt().isoformat(),
            "export_batch_no": batch_no,
            "field_codes": [f.field_name for f in export_fields],
            "field_aliases": [f.field_alias or f.field_name for f in export_fields],
            "primary_key_fields": list(pk_set),
            "structure_hash": tc.structure_version_hash,
            "data_row_count": data_row_count,
            "blank_row_start": blank_start,
            "reserved_blank_rows": RESERVED_BLANK_ROWS,
            "allow_insert_rows": tc.allow_insert_rows,
        }
        meta_ws.cell(row=1, column=1, value=json.dumps(meta_info, ensure_ascii=False))
        meta_ws.sheet_state = "hidden"

        for i, f in enumerate(export_fields, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = max(12, len(f.field_alias or f.field_name) * 2 + 4)

        file_name = "%s_%s.xlsx" % (tc.table_alias or tc.table_name, batch_no)
        file_path = os.path.join(EXPORT_DIR, file_name)
        wb.save(file_path)

        # Update task
        task.status = "completed"
        task.row_count = data_row_count
        task.file_name = file_name
        task.file_path = file_path
        task.finished_at = _now_bjt()

        # Record export log
        log = TemplateExportLog(
            export_batch_no=batch_no,
            table_config_id=tc.id,
            datasource_id=tc.datasource_id,
            export_type=export_type,
            row_count=data_row_count,
            field_count=len(export_fields),
            template_version=tc.config_version,
            file_name=file_name,
            file_path=file_path,
            export_filters_json=json.dumps(
                {"keyword": keyword, "field_filters": field_filters}, ensure_ascii=False
            ) if keyword or field_filters else None,
            operator_user=operator_user,
        )
        db.add(log)
        log_operation(db, "数据维护", "异步导出模板", "success",
                      target_id=tc.id, target_name=tc.table_name,
                      message="异步导出模板 %s，%d 行" % (file_name, data_row_count),
                      operator=operator_user)

        # Send notification to operator
        from app.utils.notifications import notify_user_by_username
        notify_user_by_username(
            db, operator_user,
            "导出完成",
            "表「%s」的导出已完成，共 %d 行数据" % (tc.table_alias or tc.table_name, data_row_count),
            ntype="success",
            related_url="/log-center",
        )

        db.commit()

    except Exception as e:
        try:
            task_obj = db.query(ExportTask).filter(ExportTask.task_id == task_id).first()
            if task_obj:
                task_obj.status = "failed"
                task_obj.error_message = str(e)[:1000]
                task_obj.finished_at = _now_bjt()
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


@router.post("/{table_config_id}/async-export")
def async_export(
    table_config_id: int,
    export_type: str = Query("all", regex="^(current|all)$"),
    keyword: Optional[str] = None,
    field_filters: Optional[str] = None,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(get_current_user),
):
    """Create async export task for large tables."""
    tc = _get_tc(db, table_config_id)
    ds = _get_ds(db, tc.datasource_id)

    task_id = uuid.uuid4().hex[:16].upper()
    task = ExportTask(
        task_id=task_id,
        table_config_id=tc.id,
        datasource_id=tc.datasource_id,
        export_type=export_type,
        export_filters_json=json.dumps(
            {"keyword": keyword, "field_filters": field_filters}, ensure_ascii=False
        ) if keyword or field_filters else None,
        status="processing",
        operator_user=_get_username(user),
    )
    db.add(task)
    db.commit()

    # Start background thread
    t = threading.Thread(
        target=_run_async_export,
        args=(task_id, table_config_id, export_type, keyword, field_filters, _get_username(user)),
        daemon=True,
    )
    t.start()

    return {
        "task_id": task_id,
        "status": "processing",
        "message": t("data_maintenance.export_task_created"),
    }


@router.get("/export-tasks")
def list_export_tasks(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    user: UserAccount = Depends(get_current_user),
):
    """List async export tasks."""
    q = db.query(ExportTask)
    if user.role != "admin":
        q = q.filter(ExportTask.operator_user == user.username)

    total = q.count()
    rows = q.order_by(ExportTask.id.desc()).offset(
        (page - 1) * page_size
    ).limit(page_size).all()

    items = []
    for r in rows:
        tc = db.query(TableConfig).filter(TableConfig.id == r.table_config_id).first()
        items.append({
            "id": r.id,
            "task_id": r.task_id,
            "table_config_id": r.table_config_id,
            "table_name": tc.table_name if tc else None,
            "table_alias": tc.table_alias if tc else None,
            "export_type": r.export_type,
            "status": r.status,
            "row_count": r.row_count,
            "file_name": r.file_name,
            "error_message": r.error_message,
            "operator_user": r.operator_user,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "finished_at": r.finished_at.isoformat() if r.finished_at else None,
        })

    return {"total": total, "items": items}


@router.get("/export-tasks/{task_id}/download")
def download_export_task(
    task_id: str,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(get_current_user),
):
    """Download completed async export file."""
    task = db.query(ExportTask).filter(ExportTask.task_id == task_id).first()
    if not task:
        raise HTTPException(404, t("data_maintenance.export_task_not_found"))
    if task.status != "completed":
        raise HTTPException(400, t("data_maintenance.export_not_ready", status=task.status))
    if not task.file_path or not os.path.isfile(task.file_path):
        raise HTTPException(404, t("data_maintenance.export_file_not_found"))

    return FileResponse(
        task.file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=task.file_name or "export.xlsx",
    )


# ─────────────────────────────────────────────
# v2.3: 批量表导出（多表 zip）
# ─────────────────────────────────────────────

class BatchExportRequest(BaseModel):
    table_config_ids: List[int]


@router.post("/batch-export")
def batch_export(
    body: BatchExportRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(get_current_user),
):
    """Export multiple tables and package as zip."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    if not body.table_config_ids:
        raise HTTPException(400, t("data_maintenance.no_tables_selected"))
    if len(body.table_config_ids) > 20:
        raise HTTPException(400, t("data_maintenance.too_many_tables"))

    zip_batch = _gen_batch("BEXP")
    zip_name = "batch_export_%s.zip" % zip_batch
    zip_path = os.path.join(EXPORT_DIR, zip_name)

    exported_files = []  # type: List[str]
    errors = []  # type: List[dict]

    for tc_id in body.table_config_ids:
        try:
            tc = db.query(TableConfig).filter(
                TableConfig.id == tc_id,
                TableConfig.is_deleted == 0,
                TableConfig.status == "enabled",
            ).first()
            if not tc:
                errors.append({"table_config_id": tc_id, "error": t("data_maintenance.table_not_found")})
                continue

            ds = db.query(DatasourceConfig).filter(
                DatasourceConfig.id == tc.datasource_id,
                DatasourceConfig.is_deleted == 0,
            ).first()
            if not ds:
                errors.append({"table_config_id": tc_id, "error": t("data_maintenance.datasource_not_found")})
                continue

            pwd = decrypt_password(ds.password_encrypted)
            fields = (
                db.query(FieldConfig)
                .filter(FieldConfig.table_config_id == tc_id, FieldConfig.is_deleted == 0)
                .order_by(FieldConfig.field_order_no)
                .all()
            )
            export_fields = [f for f in fields if f.include_in_export]
            if not export_fields:
                errors.append({"table_config_id": tc_id, "error": t("data_maintenance.no_export_fields")})
                continue

            col_names = [f.field_name for f in export_fields]

            conn = _connect(
                ds.db_type, ds.host, ds.port, ds.username, pwd,
                tc.db_name, tc.schema_name, ds.charset, ds.connect_timeout_seconds or 10,
            )
            try:
                cur = conn.cursor()
                qt = _qualified_table(ds.db_type, tc.table_name, tc.schema_name)
                cols_sql = ", ".join(_quote_col(ds.db_type, c) for c in col_names)
                cur.execute("SELECT %s FROM %s" % (cols_sql, qt))
                raw_rows = cur.fetchall()
            except Exception as e:
                errors.append({"table_config_id": tc_id, "error": t("data_maintenance.query_failed", error=str(e))})
                continue
            finally:
                conn.close()

            # Build Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "数据"

            for i, f in enumerate(export_fields, 1):
                cell = ws.cell(row=1, column=i, value=f.field_alias or f.field_name)
                cell.font = openpyxl.styles.Font(bold=True)

            for row_idx, raw in enumerate(raw_rows, 2):
                for col_idx, val in enumerate(raw, 1):
                    ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")

            for i, f in enumerate(export_fields, 1):
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = max(12, len(f.field_alias or f.field_name) * 2 + 4)

            single_batch = _gen_batch("EXP")
            single_name = "%s_%s.xlsx" % (tc.table_alias or tc.table_name, single_batch)
            single_path = os.path.join(EXPORT_DIR, single_name)
            wb.save(single_path)
            exported_files.append(single_path)

            # Record log
            log = TemplateExportLog(
                export_batch_no=single_batch,
                table_config_id=tc.id,
                datasource_id=tc.datasource_id,
                export_type="all",
                row_count=len(raw_rows),
                field_count=len(export_fields),
                template_version=tc.config_version,
                file_name=single_name,
                file_path=single_path,
                operator_user=_get_username(user),
                remark="批量导出 %s" % zip_batch,
            )
            db.add(log)

        except Exception as e:
            errors.append({"table_config_id": tc_id, "error": str(e)})

    if not exported_files:
        db.commit()
        raise HTTPException(400, t("data_maintenance.all_export_failed", error=json.dumps(errors, ensure_ascii=False)))

    # Package zip
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for fp in exported_files:
            zf.write(fp, os.path.basename(fp))

    log_operation(db, "数据维护", "批量导出", "success",
                  message="批量导出 %d 张表，%d 个错误" % (len(exported_files), len(errors)),
                  operator=_get_username(user))
    db.commit()

    return FileResponse(
        zip_path,
        media_type="application/zip",
        filename=zip_name,
    )


# ─────────────────────────────────────────────
# v3.1: 数据对比报告（Excel + PDF）
# ─────────────────────────────────────────────

class CompareReportRequest(BaseModel):
    format: str = "excel"  # excel / pdf
    import_task_id: int


@router.post("/{table_config_id}/compare-report")
def generate_compare_report(
    table_config_id: int,
    body: CompareReportRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(get_current_user),
):
    """Generate a standalone comparison report file (Excel or PDF)."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    task = db.query(ImportTaskLog).filter(ImportTaskLog.id == body.import_task_id).first()
    if not task:
        raise HTTPException(404, t("data_maintenance.import_task_not_found"))

    tc = db.query(TableConfig).filter(TableConfig.id == table_config_id).first()
    if not tc:
        raise HTTPException(404, t("data_maintenance.table_not_found"))

    diff_file = os.path.join(UPLOAD_DIR, "diff_%d.json" % body.import_task_id)
    if not os.path.isfile(diff_file):
        raise HTTPException(404, t("data_maintenance.diff_not_found"))

    with open(diff_file, "r", encoding="utf-8") as f:
        diff_data = json.load(f)

    diff_rows = diff_data.get("diff_rows", [])
    if not diff_rows:
        raise HTTPException(400, t("data_maintenance.no_diff_data"))

    # Count stats
    update_count = sum(1 for d in diff_rows if d.get("change_type") == "update")
    insert_count = sum(1 for d in diff_rows if d.get("change_type") == "insert")
    delete_count = sum(1 for d in diff_rows if d.get("change_type") == "delete")

    if body.format == "pdf":
        return _generate_pdf_report(
            tc, task, diff_rows, update_count, insert_count, delete_count, user
        )

    # ── Excel Report ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据对比报告"

    # Styles
    title_font = Font(bold=True, size=14, color="1F4E79")
    subtitle_font = Font(bold=True, size=11, color="333333")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    update_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # blue
    insert_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")  # green
    delete_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")  # red
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    change_type_map = {
        "update": ("更新", update_fill),
        "insert": ("新增", insert_fill),
        "delete": ("删除", delete_fill),
    }

    # ── Report Header (rows 1-5) ──
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="📊 数据对比报告")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    info_rows = [
        ("表名", tc.table_alias or tc.table_name),
        ("操作人", task.operator_user or _get_username(user)),
        ("生成时间", _now_bjt().strftime("%Y-%m-%d %H:%M:%S")),
        ("变更统计", f"更新 {update_count} 处 | 新增 {insert_count} 处 | 删除 {delete_count} 处 | 合计 {len(diff_rows)} 处"),
    ]
    for i, (label, value) in enumerate(info_rows, 2):
        ws.cell(row=i, column=1, value=label).font = subtitle_font
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        ws.cell(row=i, column=2, value=value)

    # ── Data Headers (row 7) ──
    data_start = 7
    headers = ["行号", "主键值", "字段名", "原值", "新值", "变更类型"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=data_start, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # ── Data Rows ──
    for row_idx, dr in enumerate(diff_rows, data_start + 1):
        change_type = dr.get("change_type", "update")
        type_text, row_fill = change_type_map.get(change_type, ("更新", update_fill))

        values = [
            dr.get("row_num", ""),
            dr.get("pk_key", ""),
            dr.get("field_alias", dr.get("field_name", "")),
            dr.get("old_value") if dr.get("old_value") is not None else "",
            dr.get("new_value") if dr.get("new_value") is not None else "",
            type_text,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = thin_border
            if col_idx == 6:
                cell.alignment = center_align

    # Column widths
    col_widths = [8, 22, 22, 28, 28, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Save
    report_name = "compare_report_%s_%d.xlsx" % (
        _now_bjt().strftime("%Y%m%d%H%M%S"), body.import_task_id
    )
    report_path = os.path.join(EXPORT_DIR, report_name)
    wb.save(report_path)

    log_operation(db, "数据维护", "导出对比报告", "success",
                  target_id=tc.id, target_name=tc.table_name,
                  message=f"导出对比报告 (Excel) {report_name}，{len(diff_rows)} 处变更",
                  operator=_get_username(user))
    db.commit()

    return FileResponse(
        report_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=report_name,
    )


def _generate_pdf_report(tc, task, diff_rows, update_count, insert_count, delete_count, user):
    """Generate PDF comparison report using reportlab."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        raise HTTPException(400, "PDF generation requires reportlab. Install it with: pip install reportlab")

    # Try to register a CJK font for Chinese support
    try:
        pdfmetrics.registerFont(TTFont('SimHei', '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc'))
        cjk_font = 'SimHei'
    except Exception:
        try:
            pdfmetrics.registerFont(TTFont('SimHei', '/usr/share/fonts/wqy-zenhei/wqy-zenhei.ttc'))
            cjk_font = 'SimHei'
        except Exception:
            cjk_font = 'Helvetica'  # fallback, won't render Chinese well

    report_name = "compare_report_%s_%d.pdf" % (
        _now_bjt().strftime("%Y%m%d%H%M%S"), task.id
    )
    report_path = os.path.join(EXPORT_DIR, report_name)

    doc = SimpleDocTemplate(report_path, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=cjk_font, fontSize=16, alignment=1)
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontName=cjk_font, fontSize=10)

    # Title
    elements.append(Paragraph("数据对比报告", title_style))
    elements.append(Spacer(1, 6*mm))

    # Info
    table_name = tc.table_alias or tc.table_name
    operator = task.operator_user or _get_username(user)
    gen_time = _now_bjt().strftime("%Y-%m-%d %H:%M:%S")
    info_text = (
        f"表名: {table_name}　　操作人: {operator}　　生成时间: {gen_time}<br/>"
        f"变更统计: 更新 {update_count} 处 | 新增 {insert_count} 处 | 删除 {delete_count} 处 | 合计 {len(diff_rows)} 处"
    )
    elements.append(Paragraph(info_text, info_style))
    elements.append(Spacer(1, 4*mm))

    # Table data
    headers = ["行号", "主键值", "字段名", "原值", "新值", "变更类型"]
    table_data = [headers]

    change_type_labels = {"update": "更新", "insert": "新增", "delete": "删除"}

    for dr in diff_rows[:500]:  # Limit to 500 rows for PDF
        change_type = dr.get("change_type", "update")
        table_data.append([
            str(dr.get("row_num", "")),
            str(dr.get("pk_key", ""))[:30],
            str(dr.get("field_alias", dr.get("field_name", "")))[:20],
            str(dr.get("old_value", ""))[:40] if dr.get("old_value") is not None else "",
            str(dr.get("new_value", ""))[:40] if dr.get("new_value") is not None else "",
            change_type_labels.get(change_type, change_type),
        ])

    col_widths = [35, 90, 80, 120, 120, 50]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Style the table
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), cjk_font),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (5, 1), (5, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
    ]

    # Color code rows by change type
    color_map = {
        "update": colors.HexColor("#D6E4F0"),
        "insert": colors.HexColor("#D5F5E3"),
        "delete": colors.HexColor("#FADBD8"),
    }
    for i, dr in enumerate(diff_rows[:500], 1):
        ct = dr.get("change_type", "update")
        if ct in color_map:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), color_map[ct]))

    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    if len(diff_rows) > 500:
        elements.append(Spacer(1, 4*mm))
        elements.append(Paragraph(f"（仅显示前 500 条，共 {len(diff_rows)} 条变更）", info_style))

    doc.build(elements)

    log_operation(SessionLocal(), "数据维护", "导出对比报告", "success",
                  target_id=tc.id, target_name=tc.table_name,
                  message=f"导出对比报告 (PDF) {report_name}，{len(diff_rows)} 处变更",
                  operator=_get_username(user))

    return FileResponse(
        report_path,
        media_type="application/pdf",
        filename=report_name,
    )
