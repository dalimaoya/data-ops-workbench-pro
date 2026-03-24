"""AI Data Prediction — AI 数据预填"""

import io
import json
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel

from app.database import get_db
from app.models import TableConfig, DatasourceConfig, FieldConfig, AIConfig, UserAccount
from app.utils.auth import get_current_user
from app.utils.crypto import decrypt_password
from app.utils.remote_db import _connect
from app.utils.audit import log_operation

router = APIRouter(prefix="/api/ai-predict", tags=["ai-predict"])


class PredictRequest(BaseModel):
    table_id: int
    time_field: Optional[str] = None  # field used as time dimension
    rows_for_context: int = 50  # how many recent rows to send to AI


@router.post("/generate")
async def generate_prediction(
    body: PredictRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(get_current_user),
):
    """Use AI to predict next period data based on historical records."""
    tc = db.query(TableConfig).filter(TableConfig.id == body.table_id, TableConfig.is_deleted == 0).first()
    if not tc:
        raise HTTPException(404, "表配置不存在")
    ds = db.query(DatasourceConfig).filter(DatasourceConfig.id == tc.datasource_id, DatasourceConfig.is_deleted == 0).first()
    if not ds:
        raise HTTPException(404, "数据源不存在")

    fields = db.query(FieldConfig).filter(
        FieldConfig.table_config_id == body.table_id, FieldConfig.is_deleted == 0, FieldConfig.is_displayed == 1
    ).order_by(FieldConfig.display_order).all()
    if not fields:
        raise HTTPException(400, "无可用字段")

    col_names = [f.field_name for f in fields]
    display_names = [f.display_name or f.field_name for f in fields]

    # Determine time field
    time_field = body.time_field
    if not time_field:
        for f in fields:
            fn = f.field_name.lower()
            if any(kw in fn for kw in ["year", "date", "time", "month", "quarter", "period", "日期", "年份", "月份"]):
                time_field = f.field_name
                break

    # Fetch recent data
    pwd = decrypt_password(ds.password_encrypted)
    conn = _connect(ds.db_type, ds.host, ds.port, ds.username, pwd,
                    tc.db_name, tc.schema_name, ds.charset, ds.connect_timeout_seconds or 10)
    try:
        cur = conn.cursor()
        col_str = ",".join(col_names)
        order = f"ORDER BY {time_field} DESC" if time_field else ""
        sql = f"SELECT {col_str} FROM {tc.table_name} {order}"
        cur.execute(sql)
        rows = cur.fetchmany(body.rows_for_context)
        rows.reverse()  # chronological order
    finally:
        conn.close()

    if not rows:
        raise HTTPException(400, "表中无数据，无法预测")

    # Format data for AI
    data_lines = []
    for row in rows:
        line = {display_names[i]: (str(v) if v is not None else "") for i, v in enumerate(row)}
        data_lines.append(line)

    prompt = f"""你是一个数据分析专家。以下是表 "{tc.table_alias or tc.table_name}" 的历史数据（按时间排列），共 {len(data_lines)} 行。

字段说明：{', '.join(f'{col_names[i]}({display_names[i]})' for i in range(len(col_names)))}

历史数据（JSON 数组）：
{json.dumps(data_lines[-20:], ensure_ascii=False, indent=1)}

请根据数据的变化趋势，预测下一期的数据。返回一个 JSON 对象，key 为字段名（使用英文字段名），value 为预测值。
对于数值型字段，给出合理的预测数值。
对于文本字段（如地区、类别等），保持与最近一期相同。
对于时间字段，推算下一期的时间值。

只返回 JSON 对象，不要其他内容。"""

    # Call AI
    ai_cfg = db.query(AIConfig).first()
    if not ai_cfg:
        raise HTTPException(400, "请先配置 AI 模型")

    # Determine which AI config to use
    api_url = ai_cfg.cloud_api_url or ai_cfg.local_api_url or ai_cfg.api_url
    api_key = None
    model_name = ai_cfg.cloud_model_name or ai_cfg.local_model_name or ai_cfg.model_name
    api_protocol = ai_cfg.cloud_api_protocol or ai_cfg.local_api_protocol or ai_cfg.api_protocol or "openai"

    for key_field in [ai_cfg.cloud_api_key_encrypted, ai_cfg.local_api_key_encrypted, ai_cfg.api_key_encrypted]:
        if key_field:
            try:
                api_key = decrypt_password(key_field)
                break
            except Exception:
                continue

    if not api_url or not api_key:
        raise HTTPException(400, "AI 模型未正确配置")

    import httpx
    headers = {"Content-Type": "application/json"}
    if api_protocol == "claude":
        headers["x-api-key"] = api_key
        headers["anthropic-version"] = "2023-06-01"
        payload = {
            "model": model_name,
            "max_tokens": 2000,
            "messages": [{"role": "user", "content": prompt}],
        }
    else:
        headers["Authorization"] = f"Bearer {api_key}"
        payload = {
            "model": model_name,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.3,
            "max_tokens": 2000,
        }

    # Ensure URL ends with correct path
    url = api_url.rstrip("/")
    if not url.endswith("/chat/completions") and api_protocol != "claude":
        if not url.endswith("/v1"):
            url += "/v1"
        url += "/chat/completions"
    elif api_protocol == "claude" and not url.endswith("/messages"):
        url += "/v1/messages"

    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(url, json=payload, headers=headers)
        if resp.status_code != 200:
            raise HTTPException(502, f"AI 调用失败: {resp.text[:200]}")
        result = resp.json()

    # Extract AI response
    if api_protocol == "claude":
        ai_text = result.get("content", [{}])[0].get("text", "")
    else:
        ai_text = result.get("choices", [{}])[0].get("message", {}).get("content", "")

    # Parse JSON from response
    predicted = {}
    try:
        # Try to find JSON in the response
        import re
        json_match = re.search(r'\{[^{}]+\}', ai_text, re.DOTALL)
        if json_match:
            predicted = json.loads(json_match.group())
    except (json.JSONDecodeError, TypeError):
        predicted = {"_raw": ai_text}

    log_operation(db, "AI预填", "数据预测", "success",
                  target_id=body.table_id, target_name=tc.table_name,
                  message=f"AI 预测了 {len(predicted)} 个字段",
                  operator=user.username)

    return {"predicted": predicted, "fields": col_names, "display_names": display_names}


@router.post("/download-template")
async def download_predict_template(
    body: PredictRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(get_current_user),
):
    """Generate prediction and download as Excel template."""
    # Re-use prediction logic
    result = await generate_prediction(body, db, user)
    predicted = result["predicted"]
    fields = result["fields"]
    display_names = result["display_names"]

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI预填模板"

    # Header
    for ci, dn in enumerate(display_names, 1):
        ws.cell(row=1, column=ci, value=dn)

    # Predicted row
    for ci, fn in enumerate(fields, 1):
        val = predicted.get(fn, "")
        ws.cell(row=2, column=ci, value=str(val) if val else "")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ai_predict_template.xlsx"},
    )
