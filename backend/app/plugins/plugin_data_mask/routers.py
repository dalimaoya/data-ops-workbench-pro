"""Data Masking Export — 脱敏规则配置 + 脱敏导出 API"""

import re
import io
import json
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel

from app.database import get_db
from app.models import TableConfig, DatasourceConfig, FieldConfig, UserAccount
from app.utils.auth import get_current_user, require_role
from app.utils.crypto import decrypt_password
from app.utils.remote_db import _connect
from app.utils.audit import log_operation

router = APIRouter(prefix="/api/data-maintenance", tags=["data-mask"])

# ─── Masking rule types ───
MASK_RULES = {
    "phone": {
        "pattern": r"(\d{3})\d{4}(\d{4})",
        "replace": r"\g<1>****\g<2>",
        "desc_zh": "手机号脱敏",
        "desc_en": "Phone masking",
    },
    "id_card": {
        "pattern": r"(\d{3})\d{11,12}(\d{4})",
        "replace": r"\g<1>***********\g<2>",
        "desc_zh": "身份证脱敏",
        "desc_en": "ID card masking",
    },
    "name": {
        "desc_zh": "姓名脱敏",
        "desc_en": "Name masking",
    },
    "bank_card": {
        "pattern": r"\d+(\d{4})",
        "replace": r"************\g<1>",
        "desc_zh": "银行卡脱敏",
        "desc_en": "Bank card masking",
    },
    "custom": {
        "desc_zh": "自定义正则脱敏",
        "desc_en": "Custom regex masking",
    },
}


def apply_mask(value: Any, rule_type: str, custom_pattern: str = None, custom_replace: str = None) -> str:
    """Apply masking rule to a value."""
    if value is None:
        return ""
    s = str(value)
    if not s.strip():
        return s

    if rule_type == "phone":
        return re.sub(MASK_RULES["phone"]["pattern"], MASK_RULES["phone"]["replace"], s)
    elif rule_type == "id_card":
        return re.sub(MASK_RULES["id_card"]["pattern"], MASK_RULES["id_card"]["replace"], s)
    elif rule_type == "name":
        if len(s) <= 1:
            return s
        return s[0] + "*" * (len(s) - 1)
    elif rule_type == "bank_card":
        if len(s) <= 4:
            return s
        return "*" * (len(s) - 4) + s[-4:]
    elif rule_type == "custom" and custom_pattern:
        try:
            return re.sub(custom_pattern, custom_replace or "***", s)
        except re.error:
            return s
    return s


class MaskRuleConfig(BaseModel):
    field_name: str
    rule_type: str  # phone/id_card/name/bank_card/custom
    custom_pattern: Optional[str] = None
    custom_replace: Optional[str] = None


class MaskRuleUpdate(BaseModel):
    rules: List[MaskRuleConfig]


@router.get("/mask-rules")
def list_mask_rules():
    """Return available masking rule types."""
    return {"rules": {k: {"desc_zh": v["desc_zh"], "desc_en": v["desc_en"]} for k, v in MASK_RULES.items()}}


@router.get("/{table_id}/mask-config")
def get_mask_config(
    table_id: int,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(get_current_user),
):
    """Get masking config for a table (stored in field_config.remark JSON)."""
    fields = db.query(FieldConfig).filter(
        FieldConfig.table_config_id == table_id, FieldConfig.is_deleted == 0
    ).all()
    result = []
    for f in fields:
        mask_rule = None
        if f.remark:
            try:
                meta = json.loads(f.remark)
                mask_rule = meta.get("mask_rule")
            except (json.JSONDecodeError, TypeError):
                pass
        result.append({
            "field_name": f.field_name,
            "display_name": f.display_name or f.field_name,
            "mask_rule": mask_rule,
        })
    return {"fields": result}


@router.put("/{table_id}/mask-config")
def update_mask_config(
    table_id: int,
    body: MaskRuleUpdate,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin")),
):
    """Update masking rules for fields in a table."""
    for rule in body.rules:
        field = db.query(FieldConfig).filter(
            FieldConfig.table_config_id == table_id,
            FieldConfig.field_name == rule.field_name,
            FieldConfig.is_deleted == 0,
        ).first()
        if not field:
            continue
        meta = {}
        if field.remark:
            try:
                meta = json.loads(field.remark)
            except (json.JSONDecodeError, TypeError):
                meta = {"original_remark": field.remark}
        meta["mask_rule"] = {
            "type": rule.rule_type,
            "custom_pattern": rule.custom_pattern,
            "custom_replace": rule.custom_replace,
        }
        field.remark = json.dumps(meta, ensure_ascii=False)
    db.commit()
    log_operation(db, "数据脱敏", "更新脱敏规则", "success",
                  target_id=table_id, message=f"更新了 {len(body.rules)} 个字段的脱敏规则",
                  operator=user.username)
    return {"success": True}


@router.post("/{table_id}/export-masked")
def export_masked(
    table_id: int,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(get_current_user),
):
    """Export table data with masking rules applied."""
    tc = db.query(TableConfig).filter(TableConfig.id == table_id, TableConfig.is_deleted == 0).first()
    if not tc:
        raise HTTPException(404, "表配置不存在")
    ds = db.query(DatasourceConfig).filter(DatasourceConfig.id == tc.datasource_id, DatasourceConfig.is_deleted == 0).first()
    if not ds:
        raise HTTPException(404, "数据源不存在")

    fields = db.query(FieldConfig).filter(
        FieldConfig.table_config_id == table_id, FieldConfig.is_deleted == 0, FieldConfig.is_displayed == 1
    ).order_by(FieldConfig.display_order).all()
    if not fields:
        raise HTTPException(400, "无可导出字段")

    # Load mask rules
    mask_map: Dict[str, dict] = {}
    for f in fields:
        if f.remark:
            try:
                meta = json.loads(f.remark)
                mr = meta.get("mask_rule")
                if mr:
                    mask_map[f.field_name] = mr
            except (json.JSONDecodeError, TypeError):
                pass

    pwd = decrypt_password(ds.password_encrypted)
    conn = _connect(ds.db_type, ds.host, ds.port, ds.username, pwd,
                    tc.db_name, tc.schema_name, ds.charset, ds.connect_timeout_seconds or 10)
    try:
        cur = conn.cursor()
        col_names = [f.field_name for f in fields]
        sql = f"SELECT {','.join(col_names)} FROM {tc.table_name}"
        cur.execute(sql)
        rows = cur.fetchall()
    finally:
        conn.close()

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tc.table_alias or tc.table_name

    # Header
    for ci, f in enumerate(fields, 1):
        ws.cell(row=1, column=ci, value=f.display_name or f.field_name)

    # Data with masking
    for ri, row in enumerate(rows, 2):
        for ci, f in enumerate(fields):
            val = row[ci]
            if f.field_name in mask_map:
                mr = mask_map[f.field_name]
                val = apply_mask(val, mr.get("type", ""), mr.get("custom_pattern"), mr.get("custom_replace"))
            ws.cell(row=ri, column=ci + 1, value=str(val) if val is not None else "")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    log_operation(db, "数据脱敏", "脱敏导出", "success",
                  target_id=table_id, target_name=tc.table_name,
                  message=f"脱敏导出 {len(rows)} 行",
                  operator=user.username)

    filename = f"{tc.table_alias or tc.table_name}_masked.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
