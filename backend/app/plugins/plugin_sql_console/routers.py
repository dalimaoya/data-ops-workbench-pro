"""SQL Console — 轻量 SQL 控制台 (只允许 SELECT)"""

import csv
import io
import re
import json
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel

from app.database import get_db
from app.models import DatasourceConfig, UserAccount
from app.utils.auth import get_current_user, require_role
from app.utils.crypto import decrypt_password
from app.utils.remote_db import _connect
from app.utils.audit import log_operation

logger = logging.getLogger("sql-console")
_BJT = timezone(timedelta(hours=8))

router = APIRouter(prefix="/api/sql-console", tags=["sql-console"])

# Dangerous keywords to block
BLOCKED_KEYWORDS = re.compile(
    r'\b(INSERT|UPDATE|DELETE|DROP|ALTER|CREATE|TRUNCATE|GRANT|REVOKE|EXEC|EXECUTE|MERGE|REPLACE|CALL|INTO)\b',
    re.IGNORECASE,
)

# Allow only SELECT (and WITH for CTEs)
ALLOWED_START = re.compile(r'^\s*(SELECT|WITH|EXPLAIN|SHOW|DESCRIBE|DESC)\b', re.IGNORECASE)


def _validate_sql(sql: str):
    """Validate SQL is safe (SELECT only)."""
    cleaned = sql.strip().rstrip(";").strip()
    if not cleaned:
        raise HTTPException(400, "SQL 不能为空")
    if not ALLOWED_START.match(cleaned):
        raise HTTPException(400, "只允许 SELECT / EXPLAIN / SHOW 查询")
    if BLOCKED_KEYWORDS.search(cleaned):
        raise HTTPException(400, "检测到危险关键字，只允许 SELECT 查询")
    # Check for multiple statements
    # Simple heuristic: split by ; and check
    parts = [p.strip() for p in cleaned.split(";") if p.strip()]
    if len(parts) > 1:
        raise HTTPException(400, "不允许执行多条 SQL 语句")
    return cleaned


class ExecuteRequest(BaseModel):
    datasource_id: int
    sql: str
    max_rows: int = 1000
    schema_name: Optional[str] = None


@router.post("/execute")
def execute_sql(
    body: ExecuteRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin")),
):
    """Execute a SELECT query on the specified datasource."""
    sql = _validate_sql(body.sql)

    ds = db.query(DatasourceConfig).filter(
        DatasourceConfig.id == body.datasource_id, DatasourceConfig.is_deleted == 0
    ).first()
    if not ds:
        raise HTTPException(404, "数据源不存在")

    pwd = decrypt_password(ds.password_encrypted)
    conn = _connect(ds.db_type, ds.host, ds.port, ds.username, pwd,
                    ds.database_name, body.schema_name, ds.charset, ds.connect_timeout_seconds or 10)
    try:
        cur = conn.cursor()
        cur.execute(sql)

        # Get column names
        columns = []
        if cur.description:
            columns = [desc[0] for desc in cur.description]

        rows = cur.fetchmany(body.max_rows)
        # Convert to list of lists (string values)
        data = []
        for row in rows:
            data.append([str(v) if v is not None else "" for v in row])

        total_available = len(data)
        if len(data) >= body.max_rows:
            # There might be more rows
            try:
                extra = cur.fetchone()
                if extra:
                    total_available = body.max_rows + 1  # Indicate truncation
            except Exception:
                pass

    except Exception as e:
        raise HTTPException(400, f"SQL 执行失败: {str(e)}")
    finally:
        conn.close()

    log_operation(db, "SQL控制台", "执行查询", "success",
                  message=f"查询返回 {len(data)} 行, SQL: {sql[:100]}",
                  operator=user.username)

    return {
        "columns": columns,
        "rows": data,
        "row_count": len(data),
        "truncated": total_available > body.max_rows,
    }


class ExportRequest(BaseModel):
    datasource_id: int
    sql: str
    max_rows: int = 50000
    schema_name: Optional[str] = None


@router.post("/export")
def export_query_result(
    body: ExportRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin")),
):
    """Export query result as Excel."""
    sql = _validate_sql(body.sql)

    ds = db.query(DatasourceConfig).filter(
        DatasourceConfig.id == body.datasource_id, DatasourceConfig.is_deleted == 0
    ).first()
    if not ds:
        raise HTTPException(404, "数据源不存在")

    pwd = decrypt_password(ds.password_encrypted)
    conn = _connect(ds.db_type, ds.host, ds.port, ds.username, pwd,
                    ds.database_name, body.schema_name, ds.charset, ds.connect_timeout_seconds or 10)
    try:
        cur = conn.cursor()
        cur.execute(sql)
        columns = [desc[0] for desc in cur.description] if cur.description else []
        rows = cur.fetchmany(body.max_rows)
    except Exception as e:
        raise HTTPException(400, f"SQL 执行失败: {str(e)}")
    finally:
        conn.close()

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Query Result"

    for ci, col in enumerate(columns, 1):
        ws.cell(row=1, column=ci, value=col)

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=str(val) if val is not None else "")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    log_operation(db, "SQL控制台", "导出查询结果", "success",
                  message=f"导出 {len(rows)} 行",
                  operator=user.username)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=query_result.xlsx"},
    )


# ── AI availability check ──

@router.get("/ai-available")
def check_ai_available(
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin")),
):
    """Check if AI SQL generation is available."""
    try:
        from app.ai.ai_engine import AIEngine
        engine = AIEngine(db)
        if engine.is_enabled and engine.engine_mode != "builtin":
            client = engine.get_llm_client()
            return {"available": client is not None}
    except Exception:
        pass
    return {"available": False}


# ── CSV Export ──

@router.post("/export-csv")
def export_csv(
    body: ExportRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin")),
):
    """Export query result as CSV (UTF-8 with BOM)."""
    sql = _validate_sql(body.sql)

    ds = db.query(DatasourceConfig).filter(
        DatasourceConfig.id == body.datasource_id, DatasourceConfig.is_deleted == 0
    ).first()
    if not ds:
        raise HTTPException(404, "数据源不存在")

    pwd = decrypt_password(ds.password_encrypted)
    conn = _connect(ds.db_type, ds.host, ds.port, ds.username, pwd,
                    ds.database_name, body.schema_name, ds.charset, ds.connect_timeout_seconds or 10)
    try:
        cur = conn.cursor()
        cur.execute(sql)
        columns = [desc[0] for desc in cur.description] if cur.description else []
        rows = cur.fetchmany(body.max_rows)
    except Exception as e:
        raise HTTPException(400, f"SQL 执行失败: {str(e)}")
    finally:
        conn.close()

    buf = io.StringIO()
    buf.write('\ufeff')  # UTF-8 BOM for Excel compatibility
    writer = csv.writer(buf)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([str(v) if v is not None else "" for v in row])

    log_operation(db, "SQL控制台", "导出CSV", "success",
                  message=f"导出 {len(rows)} 行",
                  operator=user.username)

    content = buf.getvalue().encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=query_result.csv"},
    )


# ── AI SQL Generate ──

class AIGenerateRequest(BaseModel):
    datasource_id: int
    table_name: str
    query_text: str
    schema_name: Optional[str] = None


@router.post("/ai-generate")
async def ai_generate_sql(
    body: AIGenerateRequest,
    db: Session = Depends(get_db),
    user: UserAccount = Depends(require_role("admin")),
):
    """Use AI to generate a SELECT SQL from natural language description."""
    from app.ai.ai_engine import AIEngine

    engine = AIEngine(db)
    if not engine.is_enabled or engine.engine_mode == "builtin":
        raise HTTPException(400, "AI 功能未启用或未配置 LLM")

    client = engine.get_llm_client()
    if not client:
        raise HTTPException(400, "LLM 客户端不可用，请检查 AI 配置")

    ds = db.query(DatasourceConfig).filter(
        DatasourceConfig.id == body.datasource_id, DatasourceConfig.is_deleted == 0
    ).first()
    if not ds:
        raise HTTPException(404, "数据源不存在")

    # Get table structure for context
    pwd = decrypt_password(ds.password_encrypted)
    conn = _connect(ds.db_type, ds.host, ds.port, ds.username, pwd,
                    ds.database_name, body.schema_name, ds.charset, ds.connect_timeout_seconds or 10)
    try:
        cur = conn.cursor()
        field_info = []
        if ds.db_type == "mysql":
            cur.execute(f"DESCRIBE `{body.table_name}`")
            for row in cur.fetchall():
                field_info.append({"name": row[0], "type": row[1]})
        elif ds.db_type == "postgresql":
            schema = body.schema_name or ds.schema_name or "public"
            cur.execute(
                "SELECT column_name, data_type FROM information_schema.columns "
                "WHERE table_schema = %s AND table_name = %s ORDER BY ordinal_position",
                (schema, body.table_name),
            )
            for row in cur.fetchall():
                field_info.append({"name": row[0], "type": row[1]})
        elif ds.db_type == "sqlserver":
            schema = body.schema_name or ds.schema_name or "dbo"
            cur.execute(
                "SELECT COLUMN_NAME, DATA_TYPE FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ? ORDER BY ORDINAL_POSITION",
                (schema, body.table_name),
            )
            for row in cur.fetchall():
                field_info.append({"name": row[0], "type": row[1]})
    except Exception as e:
        raise HTTPException(400, f"获取表结构失败: {str(e)}")
    finally:
        conn.close()

    if not field_info:
        raise HTTPException(400, f"表 {body.table_name} 不存在或无字段")

    fields_desc = "\n".join(f"  - {f['name']} ({f['type']})" for f in field_info)
    now_str = datetime.now(_BJT).strftime("%Y-%m-%d %H:%M")

    messages = [
        {
            "role": "system",
            "content": (
                "You are a SQL expert. Generate ONLY SELECT queries. "
                "NEVER generate INSERT, UPDATE, DELETE, DROP, ALTER, CREATE or any DDL/DML.\n"
                f"Database type: {ds.db_type}\n"
                f"Current datetime: {now_str}\n"
                f"Table: {body.table_name}\n"
                f"Columns:\n{fields_desc}\n\n"
                "Respond in JSON format: {\"sql\": \"...\", \"explanation\": \"...\"}\n"
                "The sql must be a single SELECT statement. The explanation should be in Chinese."
            ),
        },
        {
            "role": "user",
            "content": body.query_text,
        },
    ]

    try:
        resp = await client.chat(messages, max_tokens=1024, temperature=0.2)
        content = resp.get("content", "")

        # Parse JSON from response
        try:
            # Try to extract JSON from markdown code blocks
            if "```" in content:
                json_match = re.search(r'```(?:json)?\s*\n?(.*?)\n?```', content, re.DOTALL)
                if json_match:
                    content = json_match.group(1)
            parsed = json.loads(content)
            generated_sql = parsed.get("sql", "").strip()
            explanation = parsed.get("explanation", "")
        except json.JSONDecodeError:
            # Fallback: treat entire content as SQL
            generated_sql = content.strip()
            explanation = ""

        if not generated_sql:
            raise HTTPException(400, "AI 未能生成有效 SQL")

        # Validate the generated SQL is safe
        try:
            _validate_sql(generated_sql)
        except HTTPException:
            raise HTTPException(400, f"AI 生成的 SQL 不安全，已被拦截: {generated_sql[:200]}")

        log_operation(db, "SQL控制台", "AI生成SQL", "success",
                      message=f"查询: {body.query_text[:100]}, SQL: {generated_sql[:100]}",
                      operator=user.username)

        return {
            "sql": generated_sql,
            "explanation": explanation,
            "engine": "llm",
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.warning("AI SQL generation failed: %s", e)
        raise HTTPException(500, f"AI 生成失败: {str(e)}")
