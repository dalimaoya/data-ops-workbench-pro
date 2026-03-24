"""File parser engine — extracts tables from Excel, Word, PDF, CSV files."""

import io
import csv
import re
from typing import List, Dict, Any, Optional, Tuple

import openpyxl
import pandas as pd


def parse_file(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    """Parse uploaded file and extract tables.

    Returns dict with keys: file_type, file_name, tables_found, tables.
    Each table dict has: table_index, source_location, title_guess,
    row_count, col_count, headers, preview_rows, parseable.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext in ("xlsx", "xls"):
        tables = _parse_excel(file_bytes, filename)
        file_type = "excel"
    elif ext == "docx":
        tables = _parse_word(file_bytes, filename)
        file_type = "word"
    elif ext == "pdf":
        tables = _parse_pdf(file_bytes, filename)
        file_type = "pdf"
    elif ext == "csv":
        tables = _parse_csv(file_bytes, filename)
        file_type = "csv"
    else:
        raise ValueError(f"Unsupported file type: .{ext}")

    return {
        "file_type": file_type,
        "file_name": filename,
        "tables_found": len(tables),
        "tables": tables,
    }


# ── Excel ──

def _parse_excel(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    tables: List[Dict[str, Any]] = []
    idx = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_raw: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows_raw.append([_cell_to_str(c) for c in row])

        # Detect data regions in this sheet
        regions = _detect_regions(rows_raw)
        for region in regions:
            header_row_idx, data_start, data_end, col_start, col_end = region
            headers = rows_raw[header_row_idx][col_start:col_end]
            data_rows = rows_raw[data_start:data_end]
            preview = [r[col_start:col_end] for r in data_rows[:5]]

            title = _guess_title(rows_raw, header_row_idx, col_start, col_end)

            tables.append({
                "table_index": idx,
                "source_location": f"Sheet: {sheet_name}",
                "title_guess": title,
                "row_count": data_end - data_start,
                "col_count": col_end - col_start,
                "headers": headers,
                "preview_rows": preview,
                "parseable": True,
                "all_rows": [r[col_start:col_end] for r in data_rows],
            })
            idx += 1

    wb.close()
    return tables


# ── Word ──

def _parse_word(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    from docx import Document

    doc = Document(io.BytesIO(file_bytes))
    tables: List[Dict[str, Any]] = []

    for i, tbl in enumerate(doc.tables):
        rows_raw: List[List[str]] = []
        for row in tbl.rows:
            cells = [cell.text.strip() for cell in row.cells]
            rows_raw.append(cells)

        if len(rows_raw) < 2:
            continue

        headers = rows_raw[0]
        data_rows = rows_raw[1:]
        # Clean empty cols
        headers, data_rows = _trim_empty_cols(headers, data_rows)
        if not headers:
            continue

        # Try to guess title from paragraph before table
        title = _guess_word_table_title(doc, tbl, i)

        tables.append({
            "table_index": i,
            "source_location": f"表格 {i + 1}",
            "title_guess": title,
            "row_count": len(data_rows),
            "col_count": len(headers),
            "headers": headers,
            "preview_rows": data_rows[:5],
            "parseable": True,
            "all_rows": data_rows,
        })

    return tables


def _guess_word_table_title(doc, table, table_index: int) -> Optional[str]:
    """Try to find a title paragraph right before the table."""
    try:
        table_elem = table._tbl
        prev_elem = table_elem.getprevious()
        if prev_elem is not None and prev_elem.tag.endswith("}p"):
            # Gather text only from <w:t> elements in runs
            parts = []
            for child in prev_elem.iter():
                if child.tag.endswith("}t") and child.text:
                    parts.append(child.text)
            text = "".join(parts).strip()
            if text and len(text) < 100:
                return text
    except Exception:
        pass
    return None


# ── PDF ──

def _parse_pdf(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    import pdfplumber

    tables: List[Dict[str, Any]] = []
    idx = 0

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            page_tables = page.extract_tables()
            if not page_tables:
                # Check if page has images but no text (image-based table)
                text = page.extract_text() or ""
                if len(page.images) > 0 and len(text.strip()) < 20:
                    tables.append({
                        "table_index": idx,
                        "source_location": f"第{page_num}页",
                        "title_guess": None,
                        "parseable": False,
                        "parse_error": "image_table",
                        "row_count": 0,
                        "col_count": 0,
                        "headers": [],
                        "preview_rows": [],
                    })
                    idx += 1
                continue

            for tbl in page_tables:
                if not tbl or len(tbl) < 2:
                    continue

                # Clean None values
                rows_raw = [[_cell_to_str(c) for c in row] for row in tbl]

                # Remove fully empty rows
                rows_raw = [r for r in rows_raw if any(c.strip() for c in r)]
                if len(rows_raw) < 2:
                    continue

                headers = rows_raw[0]
                data_rows = rows_raw[1:]
                headers, data_rows = _trim_empty_cols(headers, data_rows)
                if not headers:
                    continue

                # Guess title from text above table on same page
                title = _guess_pdf_table_title(page, tbl)

                tables.append({
                    "table_index": idx,
                    "source_location": f"第{page_num}页",
                    "title_guess": title,
                    "row_count": len(data_rows),
                    "col_count": len(headers),
                    "headers": headers,
                    "preview_rows": data_rows[:5],
                    "parseable": True,
                    "all_rows": data_rows,
                })
                idx += 1

    return tables


def _guess_pdf_table_title(page, tbl) -> Optional[str]:
    """Try to extract a title-like text near the top of the page."""
    try:
        text = page.extract_text() or ""
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        if lines:
            first = lines[0]
            # If the first line looks like a title (short, contains table-like keywords)
            if len(first) < 80:
                return first
    except Exception:
        pass
    return None


# ── CSV ──

def _parse_csv(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows_raw = [row for row in reader]

    if len(rows_raw) < 2:
        return []

    headers = rows_raw[0]
    data_rows = rows_raw[1:]
    headers, data_rows = _trim_empty_cols(headers, data_rows)
    if not headers:
        return []

    return [{
        "table_index": 0,
        "source_location": filename,
        "title_guess": filename.rsplit(".", 1)[0] if "." in filename else filename,
        "row_count": len(data_rows),
        "col_count": len(headers),
        "headers": headers,
        "preview_rows": data_rows[:5],
        "parseable": True,
        "all_rows": data_rows,
    }]


# ── Utilities ──

def _cell_to_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _trim_empty_cols(
    headers: List[str], data_rows: List[List[str]]
) -> Tuple[List[str], List[List[str]]]:
    """Remove columns that are completely empty (header + all data)."""
    if not headers:
        return headers, data_rows

    non_empty = []
    for i in range(len(headers)):
        col_vals = [headers[i]] + [r[i] if i < len(r) else "" for r in data_rows]
        if any(v.strip() for v in col_vals):
            non_empty.append(i)

    if not non_empty:
        return [], []

    new_headers = [headers[i] for i in non_empty]
    new_rows = [[r[i] if i < len(r) else "" for i in non_empty] for r in data_rows]
    return new_headers, new_rows


def _detect_regions(
    rows: List[List[str]],
) -> List[Tuple[int, int, int, int, int]]:
    """Detect data regions in a sheet.

    Returns list of (header_row, data_start, data_end, col_start, col_end).
    Uses empty-row separation to split multiple data blocks.
    """
    if not rows:
        return []

    # Find non-empty row ranges
    non_empty_indices = []
    for i, row in enumerate(rows):
        if any(c.strip() for c in row):
            non_empty_indices.append(i)

    if not non_empty_indices:
        return []

    # Group consecutive non-empty rows (allow max 1 empty row gap within a block)
    groups: List[List[int]] = []
    current_group: List[int] = [non_empty_indices[0]]

    for i in range(1, len(non_empty_indices)):
        gap = non_empty_indices[i] - non_empty_indices[i - 1]
        if gap <= 2:  # Allow 1 empty row inside a block
            current_group.append(non_empty_indices[i])
        else:
            groups.append(current_group)
            current_group = [non_empty_indices[i]]
    groups.append(current_group)

    regions = []
    for group in groups:
        if len(group) < 2:
            continue

        start = group[0]
        end = group[-1] + 1

        # Determine column range
        max_cols = max(len(rows[i]) for i in group)
        col_start = 0
        col_end = max_cols

        # Check if first row is a title row (single non-empty cell, likely merged)
        header_row = start
        first_row = rows[start]
        non_empty_first = [c for c in first_row if c.strip()]
        if len(non_empty_first) == 1 and len(group) >= 3:
            # Single-value row is likely a title; skip to next row as header
            header_row = group[1] if len(group) > 1 else start
            data_start = header_row + 1
        else:
            data_start = start + 1

        data_end = end

        if data_start >= data_end:
            continue

        regions.append((header_row, data_start, data_end, col_start, col_end))

    # If no regions found but we have data, treat whole sheet as one table
    if not regions and len(non_empty_indices) >= 2:
        header_row = non_empty_indices[0]
        data_end = non_empty_indices[-1] + 1
        max_cols = max(len(rows[i]) for i in non_empty_indices)
        regions.append((header_row, header_row + 1, data_end, 0, max_cols))

    return regions


def _guess_title(
    rows: List[List[str]], header_row: int, col_start: int, col_end: int
) -> Optional[str]:
    """Look for a title row above the header (merged cell / single-value row)."""
    if header_row == 0:
        return None

    prev_row = rows[header_row - 1]
    non_empty = [c for c in prev_row if c.strip()]
    if len(non_empty) == 1 and len(non_empty[0]) < 100:
        return non_empty[0]

    return None
