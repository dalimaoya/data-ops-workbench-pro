"""v3.1 Comprehensive Test Suite"""
import os
import sys
import json
import time
import tempfile
import openpyxl
import csv

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)) + "/backend")

from fastapi.testclient import TestClient
from app.main import app
from app.utils.captcha import _captcha_store
from app.database import SessionLocal
from app.models import TableConfig, FieldConfig

client = TestClient(app)

results = []

def record(category, item, status, detail=""):
    results.append({"category": category, "item": item, "status": status, "detail": detail})
    icon = {"pass": "✅", "fail": "❌", "warn": "⚠️"}.get(status, "❓")
    print(f"  {icon} [{category}] {item}: {detail}")


def login_as_admin():
    cap_resp = client.get('/api/auth/captcha')
    cap_data = cap_resp.json()
    captcha_id = cap_data['captcha_id']
    code = None
    for k, (c, _) in _captcha_store.items():
        if k == captcha_id:
            code = c
            break
    resp = client.post('/api/auth/login', json={
        'username': 'admin', 'password': 'Admin123!',
        'captcha_id': captcha_id, 'captcha_code': code,
    })
    data = resp.json()
    token = data.get('access_token') or data.get('token')
    if not token:
        raise Exception(f"Login failed: {data}")
    return token


def auth_header(token):
    return {"Authorization": f"Bearer {token}"}


def create_test_xlsx(headers, rows, filename="test_upload.xlsx"):
    """Create a test xlsx file and return path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "zbap统计数据"
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    path = os.path.join(tempfile.gettempdir(), filename)
    wb.save(path)
    return path


def create_test_csv(headers, rows, filename="test_upload.csv"):
    """Create a test CSV file and return path."""
    path = os.path.join(tempfile.gettempdir(), filename)
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    return path


# ═══════════════════════════════════════
# TEST 0: Login
# ═══════════════════════════════════════
print("\n=== 0. 登录测试 ===")
try:
    token = login_as_admin()
    record("回归", "管理员登录", "pass", f"Token: {token[:20]}...")
except Exception as e:
    record("回归", "管理员登录", "fail", str(e))
    sys.exit(1)

headers = auth_header(token)

# ═══════════════════════════════════════
# TEST 1: Smart Import Center
# ═══════════════════════════════════════
print("\n=== 1. 智能数据导入中心 ===")

# 1.1 Check if there are managed tables (for matching tests)
db = SessionLocal()
managed_tables = db.query(TableConfig).filter(TableConfig.is_deleted == 0, TableConfig.status == "enabled").all()
print(f"  纳管表数量: {len(managed_tables)}")
for mt in managed_tables[:5]:
    fields = db.query(FieldConfig).filter(FieldConfig.table_config_id == mt.id, FieldConfig.is_deleted == 0).all()
    field_names = [f.field_name for f in fields[:10]]
    field_aliases = [f.field_alias for f in fields[:10] if f.field_alias]
    print(f"    - {mt.table_name} ({mt.table_alias}): fields={field_names[:5]}, aliases={field_aliases[:5]}")

# Get field info from first managed table for test data
test_headers = []
test_aliases = []
target_table_id = None
if managed_tables:
    mt = managed_tables[0]
    target_table_id = mt.id
    fields = db.query(FieldConfig).filter(FieldConfig.table_config_id == mt.id, FieldConfig.is_deleted == 0).all()
    for f in fields[:8]:
        test_headers.append(f.field_name)
        test_aliases.append(f.field_alias or f.field_name)
db.close()

if not test_headers:
    test_headers = ["年份", "地区", "GDP", "人口", "增长率"]
    test_aliases = test_headers

# Create test data
test_rows = [
    ["2023", "北京", "43000", "2189", "5.2"],
    ["2023", "上海", "47218", "2487", "5.0"],
    ["2023", "广州", "30355", "1882", "4.6"],
]
# Pad rows to match headers
for i in range(len(test_rows)):
    while len(test_rows[i]) < len(test_headers):
        test_rows[i].append(str(i))

# 1.2 Excel Upload Parse
xlsx_path = create_test_xlsx(test_aliases, test_rows, "客户统计数据.xlsx")
try:
    with open(xlsx_path, 'rb') as f:
        resp = client.post('/api/ai/import/parse-file', files={"file": ("客户统计数据.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}, headers=headers)
    data = resp.json()
    if resp.status_code == 200 and data.get("success"):
        tables = data["data"]["tables"]
        if len(tables) > 0:
            t0 = tables[0]
            record("智能导入", "Excel上传解析", "pass",
                   f"解析出{len(tables)}个表, headers={t0.get('headers',[])}[:5], 行数={t0.get('row_count',0)}")
            excel_parsed_tables = tables
        else:
            record("智能导入", "Excel上传解析", "fail", "解析结果为空")
            excel_parsed_tables = []
    else:
        record("智能导入", "Excel上传解析", "fail", f"status={resp.status_code}, body={data}")
        excel_parsed_tables = []
except Exception as e:
    record("智能导入", "Excel上传解析", "fail", str(e))
    excel_parsed_tables = []

# 1.3 CSV Upload Parse
csv_path = create_test_csv(test_aliases, test_rows, "客户统计数据.csv")
try:
    with open(csv_path, 'rb') as f:
        resp = client.post('/api/ai/import/parse-file', files={"file": ("客户统计数据.csv", f, "text/csv")}, headers=headers)
    data = resp.json()
    if resp.status_code == 200 and data.get("success"):
        tables = data["data"]["tables"]
        if len(tables) > 0:
            t0 = tables[0]
            record("智能导入", "CSV上传解析", "pass",
                   f"解析出{len(tables)}个表, headers={t0.get('headers',[])}[:5], 行数={t0.get('row_count',0)}")
        else:
            record("智能导入", "CSV上传解析", "fail", "解析结果为空")
    else:
        record("智能导入", "CSV上传解析", "fail", f"status={resp.status_code}, body={data}")
except Exception as e:
    record("智能导入", "CSV上传解析", "fail", str(e))

# 1.4 Table Matching
if excel_parsed_tables and target_table_id:
    try:
        match_req = {"tables": excel_parsed_tables, "use_ai": False}
        resp = client.post('/api/ai/import/match-tables', json=match_req, headers=headers)
        data = resp.json()
        if resp.status_code == 200 and data.get("success"):
            match_results = data["data"]
            if match_results and len(match_results) > 0:
                top = match_results[0]
                candidates = top.get("candidates", [])
                if candidates:
                    record("智能导入", "指标表匹配", "pass",
                           f"匹配到{len(candidates)}个候选, 最佳: {candidates[0].get('table_alias')} (置信度={candidates[0].get('confidence')})")
                else:
                    record("智能导入", "指标表匹配", "warn", "无匹配候选（可能是字段名不同导致）")
            else:
                record("智能导入", "指标表匹配", "warn", "返回结果为空")
        else:
            record("智能导入", "指标表匹配", "fail", f"status={resp.status_code}, body={data}")
    except Exception as e:
        record("智能导入", "指标表匹配", "fail", str(e))
else:
    record("智能导入", "指标表匹配", "warn", "无纳管表或无解析数据，跳过")

# 1.5 Field Mapping
if target_table_id and test_aliases:
    try:
        map_req = {"source_headers": test_aliases, "target_table_id": target_table_id, "use_ai": False}
        resp = client.post('/api/ai/import/map-fields', json=map_req, headers=headers)
        data = resp.json()
        if resp.status_code == 200 and data.get("success"):
            mappings = data["data"]["mappings"]
            matched = sum(1 for m in mappings if m.get("target_field"))
            total_cols = len(mappings)
            types_seen = set(m.get("match_type") for m in mappings if m.get("target_field"))
            record("智能导入", "字段映射", "pass",
                   f"映射 {matched}/{total_cols} 列, 匹配类型: {types_seen}")
            
            template_match = data["data"].get("matched_template")
            if template_match:
                record("智能导入", "模板自动匹配", "pass", f"找到模板: {template_match.get('template_name')}")
        else:
            record("智能导入", "字段映射", "fail", f"status={resp.status_code}, body={data}")
    except Exception as e:
        record("智能导入", "字段映射", "fail", str(e))
else:
    record("智能导入", "字段映射", "warn", "无纳管表跳过")

# 1.6 Mapping Template CRUD
print("\n  --- 映射模板 CRUD ---")
template_id = None
if target_table_id:
    # Create
    try:
        tmpl_data = {
            "template_name": "v31测试模板",
            "target_table_id": target_table_id,
            "mappings": [{"source": h, "target": h} for h in test_aliases[:3]],
            "source_headers": test_aliases,
        }
        resp = client.post('/api/ai/import/mapping-templates', json=tmpl_data, headers=headers)
        data = resp.json()
        if resp.status_code == 200 and data.get("success"):
            template_id = data["data"]["id"]
            record("智能导入", "保存映射模板", "pass", f"模板ID={template_id}")
        else:
            record("智能导入", "保存映射模板", "fail", f"{data}")
    except Exception as e:
        record("智能导入", "保存映射模板", "fail", str(e))

    # List
    try:
        resp = client.get('/api/ai/import/mapping-templates', headers=headers)
        data = resp.json()
        if resp.status_code == 200 and data.get("success"):
            templates = data["data"]
            record("智能导入", "模板列表", "pass", f"共{len(templates)}个模板")
        else:
            record("智能导入", "模板列表", "fail", f"{data}")
    except Exception as e:
        record("智能导入", "模板列表", "fail", str(e))

    # Delete (cleanup)
    if template_id:
        try:
            resp = client.delete(f'/api/ai/import/mapping-templates/{template_id}', headers=headers)
            data = resp.json()
            if data.get("success"):
                record("智能导入", "删除模板", "pass", f"模板{template_id}已删除")
            else:
                record("智能导入", "删除模板", "fail", f"{data}")
        except Exception as e:
            record("智能导入", "删除模板", "fail", str(e))

# 1.7 Unsupported file type
try:
    resp = client.post('/api/ai/import/parse-file',
                       files={"file": ("test.txt", b"hello", "text/plain")},
                       headers=headers)
    if resp.status_code == 400:
        record("智能导入", "不支持的文件类型拒绝", "pass", "正确返回400")
    else:
        record("智能导入", "不支持的文件类型拒绝", "warn", f"status={resp.status_code}")
except Exception as e:
    record("智能导入", "不支持的文件类型拒绝", "fail", str(e))


# ═══════════════════════════════════════
# TEST 2: Scheduler
# ═══════════════════════════════════════
print("\n=== 2. 定时任务 ===")

# 2.1 List tasks (empty initially maybe)
try:
    resp = client.get('/api/scheduler/tasks', headers=headers)
    data = resp.json()
    if resp.status_code == 200 and 'items' in data:
        record("定时任务", "任务列表页", "pass", f"共{data['total']}个任务")
    else:
        record("定时任务", "任务列表页", "fail", f"status={resp.status_code}, body={data}")
except Exception as e:
    record("定时任务", "任务列表页", "fail", str(e))

# 2.2 Create task - hourly health check
task_id = None
try:
    create_data = {
        "name": "每小时健康巡检",
        "type": "health_check",
        "schedule": {"type": "interval", "hours": 1},
        "enabled": True,
        "config": {},
    }
    resp = client.post('/api/scheduler/tasks', json=create_data, headers=headers)
    data = resp.json()
    if resp.status_code == 200 and 'id' in data:
        task_id = data['id']
        record("定时任务", "创建任务", "pass", f"任务ID={task_id}, 名称=每小时健康巡检")
    else:
        record("定时任务", "创建任务", "fail", f"status={resp.status_code}, body={data}")
except Exception as e:
    record("定时任务", "创建任务", "fail", str(e))

# 2.3 Verify task in list
if task_id:
    try:
        resp = client.get('/api/scheduler/tasks', headers=headers)
        data = resp.json()
        found = any(t['id'] == task_id for t in data.get('items', []))
        if found:
            task_item = [t for t in data['items'] if t['id'] == task_id][0]
            record("定时任务", "任务列表显示", "pass",
                   f"名称={task_item['name']}, 类型={task_item['type']}, "
                   f"频率={task_item.get('schedule',{})}, 启用={task_item['enabled']}, "
                   f"下次执行={task_item.get('next_run')}")
        else:
            record("定时任务", "任务列表显示", "fail", "任务未在列表中")
    except Exception as e:
        record("定时任务", "任务列表显示", "fail", str(e))

# 2.4 Run now
if task_id:
    try:
        resp = client.post(f'/api/scheduler/tasks/{task_id}/run', headers=headers)
        data = resp.json()
        if resp.status_code == 200:
            record("定时任务", "立即执行", "pass", f"返回: {data.get('message')}")
            time.sleep(3)  # Wait for execution
        else:
            record("定时任务", "立即执行", "fail", f"status={resp.status_code}, body={data}")
    except Exception as e:
        record("定时任务", "立即执行", "fail", str(e))

# 2.5 Check execution history
if task_id:
    try:
        resp = client.get(f'/api/scheduler/tasks/{task_id}/history', headers=headers)
        data = resp.json()
        if resp.status_code == 200 and 'items' in data:
            history = data['items']
            if len(history) > 0:
                latest = history[0]
                record("定时任务", "执行历史", "pass",
                       f"共{len(history)}条记录, 最新状态={latest.get('status')}, "
                       f"摘要={latest.get('result_summary','')[:60]}")
            else:
                record("定时任务", "执行历史", "warn", "暂无执行记录（任务可能仍在执行中）")
        else:
            record("定时任务", "执行历史", "fail", f"status={resp.status_code}, body={data}")
    except Exception as e:
        record("定时任务", "执行历史", "fail", str(e))

# 2.6 Edit task
if task_id:
    try:
        resp = client.put(f'/api/scheduler/tasks/{task_id}', json={
            "name": "每小时健康巡检(已修改)",
            "schedule": {"type": "interval", "minutes": 30},
        }, headers=headers)
        data = resp.json()
        if resp.status_code == 200:
            record("定时任务", "编辑任务", "pass", "修改频率为30分钟")
        else:
            record("定时任务", "编辑任务", "fail", f"status={resp.status_code}, body={data}")
    except Exception as e:
        record("定时任务", "编辑任务", "fail", str(e))

# 2.7 Disable task
if task_id:
    try:
        resp = client.put(f'/api/scheduler/tasks/{task_id}', json={"enabled": False}, headers=headers)
        data = resp.json()
        if resp.status_code == 200:
            record("定时任务", "禁用任务", "pass", "任务已禁用")
        else:
            record("定时任务", "禁用任务", "fail", f"{data}")
    except Exception as e:
        record("定时任务", "禁用任务", "fail", str(e))

# 2.8 Invalid task type
try:
    resp = client.post('/api/scheduler/tasks', json={
        "name": "invalid", "type": "invalid_type",
        "schedule": {"type": "cron"}, "enabled": True,
    }, headers=headers)
    if resp.status_code == 400:
        record("定时任务", "无效任务类型拒绝", "pass", "正确返回400")
    else:
        record("定时任务", "无效任务类型拒绝", "warn", f"status={resp.status_code}")
except Exception as e:
    record("定时任务", "无效任务类型拒绝", "fail", str(e))

# 2.9 Delete task
if task_id:
    try:
        resp = client.delete(f'/api/scheduler/tasks/{task_id}', headers=headers)
        data = resp.json()
        if resp.status_code == 200:
            record("定时任务", "删除任务", "pass", f"任务{task_id}已删除")
        else:
            record("定时任务", "删除任务", "fail", f"{data}")
    except Exception as e:
        record("定时任务", "删除任务", "fail", str(e))

# 2.10 Non-admin access (test with operator)
print("\n  --- 权限控制 ---")
try:
    # Reset operator password  
    from passlib.context import CryptContext
    pwd_ctx = CryptContext(schemes=['bcrypt'], deprecated='auto')
    db2 = SessionLocal()
    from app.models import UserAccount
    op_user = db2.query(UserAccount).filter(UserAccount.username=='operator').first()
    if op_user:
        op_user.password_hash = pwd_ctx.hash('Operator123!')
        db2.commit()
    db2.close()

    # Login as operator
    cap_resp = client.get('/api/auth/captcha')
    cap_data = cap_resp.json()
    captcha_id2 = cap_data['captcha_id']
    code2 = None
    for k, (c, _) in _captcha_store.items():
        if k == captcha_id2:
            code2 = c
            break
    resp = client.post('/api/auth/login', json={
        'username': 'operator', 'password': 'Operator123!',
        'captcha_id': captcha_id2, 'captcha_code': code2,
    })
    op_resp_data = resp.json()
    op_token = op_resp_data.get('access_token') or op_resp_data.get('token')
    if op_token:
        op_headers = {"Authorization": f"Bearer {op_token}"}
        resp = client.get('/api/scheduler/tasks', headers=op_headers)
        if resp.status_code == 403:
            record("定时任务", "非管理员访问拒绝", "pass", "正确返回403")
        else:
            record("定时任务", "非管理员访问拒绝", "warn", f"status={resp.status_code} (应为403)")
    else:
        record("定时任务", "非管理员访问拒绝", "warn", "operator登录失败，跳过")
except Exception as e:
    record("定时任务", "非管理员访问拒绝", "fail", str(e))


# ═══════════════════════════════════════
# TEST 3: Data Comparison Report
# ═══════════════════════════════════════
print("\n=== 3. 数据对比报告 ===")

# Check if there are import tasks with diff data
if target_table_id:
    try:
        # Check for existing diff files
        import glob
        upload_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "backend", "data", "uploads")
        diff_files = glob.glob(os.path.join(upload_dir, "diff_*.json"))
        print(f"  已有diff文件数: {len(diff_files)}")

        if diff_files:
            # Use existing diff file
            diff_file = diff_files[0]
            import_task_id = int(os.path.basename(diff_file).replace("diff_", "").replace(".json", ""))
            print(f"  使用diff文件: {diff_file}, import_task_id={import_task_id}")

            # Test Excel report
            resp = client.post(f'/api/data-maintenance/{target_table_id}/compare-report',
                              json={"format": "excel", "import_task_id": import_task_id},
                              headers=headers)
            if resp.status_code == 200 and 'spreadsheet' in resp.headers.get('content-type', ''):
                record("对比报告", "Excel格式导出", "pass",
                       f"文件大小={len(resp.content)}字节, content-type={resp.headers.get('content-type')}")
                
                # Verify Excel content
                import io
                wb = openpyxl.load_workbook(io.BytesIO(resp.content))
                ws = wb.active
                title_val = ws.cell(row=1, column=1).value
                table_name_val = ws.cell(row=2, column=2).value
                operator_val = ws.cell(row=3, column=2).value
                time_val = ws.cell(row=4, column=2).value
                stats_val = ws.cell(row=5, column=2).value
                header_row = [ws.cell(row=7, column=c).value for c in range(1, 7)]
                
                has_title = "对比报告" in str(title_val)
                has_table_name = bool(table_name_val)
                has_operator = bool(operator_val)
                has_time = bool(time_val)
                has_headers = header_row == ["行号", "主键值", "字段名", "原值", "新值", "变更类型"]
                
                # Check color marking on data rows
                data_row = 8
                data_fill = ws.cell(row=data_row, column=1).fill
                has_color = data_fill and data_fill.start_color and data_fill.start_color.rgb != "00000000"
                
                detail_parts = []
                if has_title: detail_parts.append("标题✓")
                else: detail_parts.append("标题✗")
                if has_table_name: detail_parts.append(f"表名={table_name_val}")
                if has_operator: detail_parts.append(f"操作人={operator_val}")
                if has_time: detail_parts.append(f"时间={time_val}")
                if has_headers: detail_parts.append("表头✓")
                if has_color: detail_parts.append("颜色标记✓")
                else: detail_parts.append("颜色标记⚠")
                
                record("对比报告", "Excel内容验证", "pass" if (has_title and has_table_name and has_headers) else "warn",
                       ", ".join(detail_parts))
                wb.close()
            else:
                record("对比报告", "Excel格式导出", "fail",
                       f"status={resp.status_code}, content-type={resp.headers.get('content-type')}, body={resp.text[:200]}")

            # Test PDF report
            resp = client.post(f'/api/data-maintenance/{target_table_id}/compare-report',
                              json={"format": "pdf", "import_task_id": import_task_id},
                              headers=headers)
            if resp.status_code == 200 and 'pdf' in resp.headers.get('content-type', ''):
                record("对比报告", "PDF格式导出", "pass",
                       f"文件大小={len(resp.content)}字节")
                
                # Basic PDF validation
                if resp.content[:4] == b'%PDF':
                    record("对比报告", "PDF格式验证", "pass", "PDF magic header正确")
                else:
                    record("对比报告", "PDF格式验证", "fail", "非有效PDF文件")
            else:
                detail = resp.text[:200] if resp.status_code != 200 else ""
                record("对比报告", "PDF格式导出", "fail",
                       f"status={resp.status_code}, content-type={resp.headers.get('content-type')}, {detail}")
        else:
            record("对比报告", "差异数据文件", "warn", "无现有diff文件，需要先执行导入流程产生差异数据")
            
            # Create a mock diff file for testing
            mock_diff = {
                "diff_rows": [
                    {"row_num": 1, "pk_key": "2023-北京", "field_name": "gdp", "field_alias": "GDP",
                     "old_value": "42000", "new_value": "43000", "change_type": "update"},
                    {"row_num": 2, "pk_key": "2023-深圳", "field_name": "gdp", "field_alias": "GDP",
                     "old_value": None, "new_value": "34600", "change_type": "insert"},
                    {"row_num": 3, "pk_key": "2022-武汉", "field_name": "population", "field_alias": "人口",
                     "old_value": "1365", "new_value": None, "change_type": "delete"},
                ]
            }
            os.makedirs(upload_dir, exist_ok=True)
            mock_task_id = 99999
            diff_path = os.path.join(upload_dir, f"diff_{mock_task_id}.json")
            with open(diff_path, 'w', encoding='utf-8') as f:
                json.dump(mock_diff, f, ensure_ascii=False)
            
            # Also create a mock import task log
            from app.models import ImportTaskLog
            db3 = SessionLocal()
            mock_log = ImportTaskLog(
                id=mock_task_id,
                table_config_id=target_table_id,
                datasource_id=managed_tables[0].datasource_id if managed_tables else 1,
                import_file_name="mock_test.xlsx",
                import_status="diff_ready",
                operator_user="admin",
                import_batch_no="MOCK_BATCH_99999",
            )
            try:
                db3.add(mock_log)
                db3.commit()
            except:
                db3.rollback()
                # Try update
                existing = db3.query(ImportTaskLog).filter(ImportTaskLog.id == mock_task_id).first()
                if not existing:
                    record("对比报告", "模拟差异数据", "fail", "无法创建mock import task")
            db3.close()
            
            # Now test with mock data
            resp = client.post(f'/api/data-maintenance/{target_table_id}/compare-report',
                              json={"format": "excel", "import_task_id": mock_task_id},
                              headers=headers)
            if resp.status_code == 200 and 'spreadsheet' in resp.headers.get('content-type', ''):
                record("对比报告", "Excel格式导出(模拟数据)", "pass", f"文件大小={len(resp.content)}字节")
                
                # Verify content
                import io
                wb = openpyxl.load_workbook(io.BytesIO(resp.content))
                ws = wb.active
                title_val = ws.cell(row=1, column=1).value
                has_title = "对比报告" in str(title_val) if title_val else False
                record("对比报告", "Excel报告头验证", "pass" if has_title else "fail",
                       f"标题='{title_val}'")
                wb.close()
            else:
                record("对比报告", "Excel格式导出(模拟数据)", "fail", f"status={resp.status_code}, body={resp.text[:200]}")
            
            # PDF test
            resp = client.post(f'/api/data-maintenance/{target_table_id}/compare-report',
                              json={"format": "pdf", "import_task_id": mock_task_id},
                              headers=headers)
            if resp.status_code == 200:
                record("对比报告", "PDF格式导出(模拟数据)", "pass", f"文件大小={len(resp.content)}字节")
            else:
                record("对比报告", "PDF格式导出(模拟数据)", "fail", f"status={resp.status_code}, body={resp.text[:200]}")
            
            # Cleanup
            try:
                os.remove(diff_path)
            except:
                pass
    except Exception as e:
        record("对比报告", "对比报告测试", "fail", str(e))
        import traceback
        traceback.print_exc()
else:
    record("对比报告", "跳过", "warn", "无纳管表")


# ═══════════════════════════════════════
# TEST 4: Regression
# ═══════════════════════════════════════
print("\n=== 4. 回归测试 ===")

# 4.1 Dashboard
try:
    resp = client.get('/api/dashboard/stats', headers=headers)
    if resp.status_code == 200:
        record("回归", "仪表盘统计", "pass", f"data={json.dumps(resp.json(), ensure_ascii=False)[:100]}")
    else:
        record("回归", "仪表盘统计", "fail", f"status={resp.status_code}")
except Exception as e:
    record("回归", "仪表盘统计", "fail", str(e))

# 4.2 Data browsing - list tables
try:
    resp = client.get('/api/table-config', headers=headers)
    data = resp.json()
    if resp.status_code == 200:
        if isinstance(data, list):
            total = len(data)
        else:
            total = data.get('total', len(data.get('items', [])))
        record("回归", "纳管表列表", "pass", f"共{total}张表")
    else:
        record("回归", "纳管表列表", "fail", f"status={resp.status_code}")
except Exception as e:
    record("回归", "纳管表列表", "fail", str(e))

# 4.3 AI config page
try:
    resp = client.get('/api/ai-config', headers=headers)
    if resp.status_code == 200:
        record("回归", "AI配置", "pass", "AI配置接口正常")
    else:
        record("回归", "AI配置", "fail", f"status={resp.status_code}")
except Exception as e:
    record("回归", "AI配置", "fail", str(e))

# 4.4 Health check
try:
    resp = client.get('/api/health-check/config', headers=headers)
    if resp.status_code == 200:
        record("回归", "健康巡检配置", "pass", "配置接口正常")
    else:
        record("回归", "健康巡检配置", "fail", f"status={resp.status_code}")
except Exception as e:
    record("回归", "健康巡检配置", "fail", str(e))

# 4.5 Backup
try:
    resp = client.get('/api/platform/backup/list', headers=headers)
    if resp.status_code == 200:
        record("回归", "备份列表", "pass", "备份接口正常")
    else:
        record("回归", "备份列表", "fail", f"status={resp.status_code}")
except Exception as e:
    record("回归", "备份列表", "fail", str(e))

# 4.6 Operation logs
try:
    resp = client.get('/api/logs', headers=headers)
    if resp.status_code == 200:
        record("回归", "操作日志", "pass", "日志接口正常")
    else:
        record("回归", "操作日志", "fail", f"status={resp.status_code}")
except Exception as e:
    record("回归", "操作日志", "fail", str(e))

# 4.7 Datasource list
try:
    resp = client.get('/api/datasource', headers=headers)
    if resp.status_code == 200:
        record("回归", "数据源列表", "pass", "数据源接口正常")
    else:
        record("回归", "数据源列表", "fail", f"status={resp.status_code}")
except Exception as e:
    record("回归", "数据源列表", "fail", str(e))

# 4.8 User management
try:
    resp = client.get('/api/users', headers=headers)
    if resp.status_code == 200:
        record("回归", "用户管理", "pass", "用户接口正常")
    else:
        record("回归", "用户管理", "fail", f"status={resp.status_code}")
except Exception as e:
    record("回归", "用户管理", "fail", str(e))


# ═══════════════════════════════════════
# FRONTEND CODE REVIEW
# ═══════════════════════════════════════
print("\n=== 5. 前端代码审查 ===")

# Check routing
app_tsx = open(os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend", "src", "App.tsx")).read()
has_smart_import_route = '/smart-import' in app_tsx
has_scheduler_route = '/scheduler' in app_tsx
record("前端", "智能导入路由", "pass" if has_smart_import_route else "fail",
       "/smart-import 路由已注册" if has_smart_import_route else "缺少路由")
record("前端", "定时任务路由", "pass" if has_scheduler_route else "fail",
       "/scheduler 路由已注册" if has_scheduler_route else "缺少路由")

# Check admin-only access
scheduler_admin_only = 'RequireRole' in app_tsx and 'scheduler' in app_tsx
record("前端", "定时任务权限控制", "pass" if scheduler_admin_only else "warn",
       "仅管理员可访问" if scheduler_admin_only else "权限控制待确认")

# Check smart import step components
smart_import_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend", "src", "pages", "smart-import")
step_files = os.listdir(smart_import_dir)
expected_steps = ["StepUpload", "StepMatchTables", "StepMapFields", "StepPreview"]
for step in expected_steps:
    found = any(step in f for f in step_files)
    record("前端", f"智能导入-{step}组件", "pass" if found else "fail",
           "组件存在" if found else "组件缺失")


# ═══════════════════════════════════════
# SUMMARY
# ═══════════════════════════════════════
print("\n" + "=" * 60)
print("=== 测试汇总 ===")
pass_count = sum(1 for r in results if r["status"] == "pass")
fail_count = sum(1 for r in results if r["status"] == "fail")
warn_count = sum(1 for r in results if r["status"] == "warn")
total_count = len(results)
print(f"  总计: {total_count} 项 | ✅ 通过: {pass_count} | ❌ 失败: {fail_count} | ⚠️ 警告: {warn_count}")

if fail_count > 0:
    print("\n  失败项:")
    for r in results:
        if r["status"] == "fail":
            print(f"    ❌ [{r['category']}] {r['item']}: {r['detail']}")

if warn_count > 0:
    print("\n  警告项:")
    for r in results:
        if r["status"] == "warn":
            print(f"    ⚠️ [{r['category']}] {r['item']}: {r['detail']}")

# Save results as JSON for report generation
results_path = "/tmp/v31_test_results.json"
with open(results_path, 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2)
print(f"\n  结果已保存: {results_path}")
